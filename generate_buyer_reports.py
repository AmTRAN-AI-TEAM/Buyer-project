#!/usr/bin/env python3
"""由 DPS / PP 原始資料自動生成「整理后」報表。

以原始資料 (DPS原始 工作表、PP 樞紐快取) 為唯一數值來源，重建與人工整理版
相同格式的輸出檔，放在 output/ 資料夾。

用法::

    python generate_buyer_reports.py                     # 自動抓 inout/ 下的檔案
    python generate_buyer_reports.py --compare           # 額外與人工整理版逐格對帳
    python generate_buyer_reports.py --dps A.xlsx --pp B.xlsx --out-dir /tmp/out

詳見 README.md。
"""

from __future__ import annotations

import argparse
import datetime as dt
import re
import sys
import xml.etree.ElementTree as ET
import zipfile
from collections import OrderedDict, defaultdict
from pathlib import Path
from typing import Iterable, Sequence

from openpyxl import Workbook, load_workbook
from openpyxl.styles import Alignment, Font
from openpyxl.utils import get_column_letter

SPREADSHEET_NS = "http://schemas.openxmlformats.org/spreadsheetml/2006/main"
PKG_REL_NS = "http://schemas.openxmlformats.org/package/2006/relationships"

EXCEL_EPOCH = dt.date(1899, 12, 30)

MONTH_ABBR = [
    "Jan", "Feb", "Mar", "Apr", "May", "Jun",
    "Jul", "Aug", "Sep", "Oct", "Nov", "Dec",
]
MONTH_INDEX = {name.lower(): i for i, name in enumerate(MONTH_ABBR, start=1)}

# ---------------------------------------------------------------------------
# 共用小工具
# ---------------------------------------------------------------------------

_LOG_QUIET = False


def log(message: str = "") -> None:
    if not _LOG_QUIET:
        print(message)


def warn(message: str) -> None:
    print(f"[警告] {message}", file=sys.stderr)


def numeric(value) -> float:
    """把儲存格內容轉成數字；文字 / 錯誤值 (#REF!, #N/A) 一律視為 0。"""
    if value is None:
        return 0.0
    if isinstance(value, bool):
        return 0.0
    if isinstance(value, (int, float)):
        return float(value)
    text = str(value).strip().replace(",", "")
    if not text:
        return 0.0
    try:
        return float(text)
    except ValueError:
        return 0.0


def clean_number(value: float):
    """整數就回傳 int，避免輸出 1234.0 這種樣子。"""
    if abs(value - round(value)) < 1e-9:
        return int(round(value))
    return value


def serial_to_date(serial: float) -> dt.date:
    return EXCEL_EPOCH + dt.timedelta(days=int(serial))


def parse_date_arg(text: str) -> dt.date:
    for fmt in ("%Y-%m-%d", "%Y/%m/%d", "%Y%m%d"):
        try:
            return dt.datetime.strptime(text, fmt).date()
        except ValueError:
            continue
    raise argparse.ArgumentTypeError(f"無法解析日期：{text}（請用 YYYY-MM-DD）")


def autosize(ws, minimum: int = 8, maximum: int = 30) -> None:
    for col_idx, column_cells in enumerate(ws.columns, start=1):
        width = minimum
        for cell in column_cells:
            if cell.value is not None and not str(cell.value).startswith("="):
                width = max(width, min(maximum, len(str(cell.value)) + 2))
        ws.column_dimensions[get_column_letter(col_idx)].width = width


def find_input(input_dir: Path, patterns: Sequence[str], kind: str) -> Path:
    """在 inout/ 底下依關鍵字找輸入檔；多個候選時取最新修改的那個。"""
    if not input_dir.is_dir():
        raise SystemExit(f"找不到輸入資料夾：{input_dir}")
    candidates = [
        path
        for path in sorted(input_dir.glob("*.xlsx"))
        if not path.name.startswith("~$")
        and any(re.search(pattern, path.name, re.IGNORECASE) for pattern in patterns)
    ]
    if not candidates:
        raise SystemExit(
            f"在 {input_dir} 找不到 {kind} 檔（檔名需含 {' 或 '.join(patterns)}），"
            f"或請用命令列參數明確指定。"
        )
    if len(candidates) > 1:
        candidates.sort(key=lambda p: p.stat().st_mtime, reverse=True)
        warn(f"{kind} 有多個候選檔，採用最新的：{candidates[0].name}")
    return candidates[0]


# ---------------------------------------------------------------------------
# DPS
# ---------------------------------------------------------------------------

DPS_SOURCE_SHEET = "DPS原始"
DPS_TIDY_SHEET = "DPS整理后"
DPS_HEADER_KEYS = ("Line", "W/O", "AVTC P/N")


def find_header_row(ws, required: Iterable[str]) -> int:
    required_set = set(required)
    for row in ws.iter_rows():
        values = {str(cell.value).strip() for cell in row if cell.value is not None}
        if required_set.issubset(values):
            return row[0].row
    raise SystemExit(f"在 {ws.title} 找不到含有下列欄位的表頭列：{', '.join(required)}")


def header_date(cell) -> dt.date | None:
    """判斷表頭儲存格是否為日期欄。

    只認 datetime/date 物件，或「數值 + 日期格式」的儲存格；避免把 `6月`
    這類月度小計欄或普通整數誤判成日期（舊版用 40000~50000 的數值範圍猜測，
    既會誤判也會在 2036 年後失效）。
    """
    value = cell.value
    if isinstance(value, dt.datetime):
        return value.date()
    if isinstance(value, dt.date):
        return value
    if isinstance(value, (int, float)) and not isinstance(value, bool):
        try:
            is_date = bool(cell.is_date)
        except (TypeError, ValueError):
            is_date = False
        if is_date and value > 0:
            return serial_to_date(value)
    return None


def detect_dps_tail_cutoff(dps_path: Path, max_date: dt.date) -> dt.date | None:
    """從活頁簿內既有的人工「DPS整理后」推斷末欄彙總桶的起始日。

    人工版是 Excel 樞紐日期群組的產物：最後一個日期欄其實是「> 前一日」的
    彙總桶。若該表不存在（未來只拿到原始檔）就回傳 None，代表所有日期都
    各自成欄。
    """
    try:
        wb = load_workbook(dps_path, data_only=True)
    except Exception as exc:  # noqa: BLE001 - 只是推斷失敗，不該中斷主流程
        warn(f"讀取既有整理后工作表失敗，改為不使用彙總桶：{exc}")
        return None
    try:
        if DPS_TIDY_SHEET not in wb.sheetnames:
            return None
        ws = wb[DPS_TIDY_SHEET]
        dates = [d for d in (header_date(cell) for cell in ws[1]) if d is not None]
        if not dates:
            return None
        last = max(dates)
        return last if last < max_date else None
    finally:
        wb.close()


def excel_label_sort_key(label: str, is_numeric: bool):
    """模擬 Excel 樞紐的列標籤排序：數字在前（依數值），文字在後（依字串）。"""
    if is_numeric:
        try:
            return (0, float(label), "")
        except ValueError:
            pass
    return (1, 0.0, label)


def generate_dps(
    dps_path: Path,
    output_path: Path,
    include_star_parts: bool = False,
    tail_cutoff: dt.date | None = None,
) -> dict:
    wb = load_workbook(dps_path, data_only=True)
    try:
        if DPS_SOURCE_SHEET not in wb.sheetnames:
            raise SystemExit(f"{dps_path.name} 內找不到工作表：{DPS_SOURCE_SHEET}")
        ws = wb[DPS_SOURCE_SHEET]

        header_row = find_header_row(ws, DPS_HEADER_KEYS)
        cols = {
            str(cell.value).strip(): cell.column
            for cell in ws[header_row]
            if cell.value is not None
        }
        pn_col = cols["AVTC P/N"]

        date_columns: list[tuple[int, dt.date]] = []
        for cell in ws[header_row]:
            date = header_date(cell)
            if date is not None:
                date_columns.append((cell.column, date))
        if not date_columns:
            raise SystemExit(f"{DPS_SOURCE_SHEET} 內找不到任何日期欄")

        # 每個日期應該剛好有 D / N 兩班兩欄，不符就示警（資料格式可能變了）
        per_date = defaultdict(int)
        for _col, date in date_columns:
            per_date[date] += 1
        odd = {d: n for d, n in per_date.items() if n != 2}
        if odd:
            warn(
                "下列日期的欄數不是 2（D/N 兩班），請確認原始檔格式："
                + ", ".join(f"{d} x{n}" for d, n in sorted(odd.items()))
            )

        all_dates = sorted(per_date)
        date_col_set = {col for col, _date in date_columns}
        first_date_col = date_columns[0][0]

        aggregate: dict[str, dict[dt.date, float]] = defaultdict(lambda: defaultdict(float))
        label_is_numeric: dict[str, bool] = {}
        excluded: dict[str, float] = defaultdict(float)
        text_cells = 0
        orphan_cols: set[int] = set()

        for row in ws.iter_rows(min_row=header_row + 1, values_only=True):
            raw_pn = row[pn_col - 1] if pn_col - 1 < len(row) else None
            if raw_pn is None or not str(raw_pn).strip():
                continue
            pn = str(raw_pn).strip()
            is_numeric = isinstance(raw_pn, (int, float)) and not isinstance(raw_pn, bool)
            label_is_numeric.setdefault(pn, is_numeric)

            # 表頭沒被判定成日期、底下卻有數量的欄位 → 可能是漏判的日期欄
            for col_idx in range(first_date_col + 1, len(row) + 1):
                if col_idx in date_col_set:
                    continue
                value = row[col_idx - 1]
                if isinstance(value, (int, float)) and not isinstance(value, bool) and value:
                    orphan_cols.add(col_idx)

            row_total = 0.0
            for col_idx, date in date_columns:
                value = row[col_idx - 1] if col_idx - 1 < len(row) else None
                if value is None:
                    continue
                if not isinstance(value, (int, float)) or isinstance(value, bool):
                    if str(value).strip():
                        text_cells += 1
                    continue
                row_total += float(value)
                aggregate[pn][date] += float(value)

            if pn.endswith("*"):
                excluded[pn] += row_total

        if orphan_cols:
            warn(
                "下列欄位的表頭沒有被判定為日期，但底下有數量，未計入輸出："
                + ", ".join(get_column_letter(c) for c in sorted(orphan_cols))
                + "。常見原因：表頭日期是公式而檔案未經 Excel 存檔（無快取值）。"
            )

        if not include_star_parts:
            star_parts = [pn for pn in list(aggregate) + list(label_is_numeric) if pn.endswith("*")]
            for pn in set(star_parts):
                aggregate.pop(pn, None)
                label_is_numeric.pop(pn, None)
        else:
            excluded.clear()

        # 決定輸出欄：cutoff 之後的日期全部併入 cutoff 這一欄
        if tail_cutoff is not None:
            out_dates = [d for d in all_dates if d < tail_cutoff] + [tail_cutoff]
        else:
            out_dates = all_dates

        def bucket(date: dt.date) -> dt.date:
            if tail_cutoff is not None and date >= tail_cutoff:
                return tail_cutoff
            return date

        rows: list[tuple[str, bool, dict[dt.date, float]]] = []
        for pn in sorted(
            label_is_numeric, key=lambda p: excel_label_sort_key(p, label_is_numeric[p])
        ):
            values: dict[dt.date, float] = defaultdict(float)
            for date, qty in aggregate.get(pn, {}).items():
                values[bucket(date)] += qty
            rows.append((pn, label_is_numeric[pn], values))

        write_dps_workbook(output_path, out_dates, rows)

        grand_total = sum(sum(v.values()) for _pn, _n, v in rows)
        return {
            "source": dps_path,
            "header_row": header_row,
            "date_columns": len(date_columns),
            "dates": len(all_dates),
            "date_range": (all_dates[0], all_dates[-1]),
            "tail_cutoff": tail_cutoff,
            "out_columns": len(out_dates),
            "rows": len(rows),
            "grand_total": clean_number(grand_total),
            "excluded": dict(excluded),
            "text_cells": text_cells,
        }
    finally:
        wb.close()


def write_dps_workbook(
    output_path: Path,
    dates: Sequence[dt.date],
    rows: Sequence[tuple[str, bool, dict[dt.date, float]]],
) -> None:
    wb = Workbook()
    ws = wb.active
    ws.title = DPS_TIDY_SHEET

    ws.cell(1, 1, "行标签")
    for offset, date in enumerate(dates):
        cell = ws.cell(1, 2 + offset, dt.datetime(date.year, date.month, date.day))
        cell.number_format = "m/d;@"
    total_col = 2 + len(dates)
    ws.cell(1, total_col, "total")

    first_letter = get_column_letter(2)
    last_letter = get_column_letter(total_col - 1)

    for r_offset, (pn, is_numeric, values) in enumerate(rows):
        r = 2 + r_offset
        ws.cell(r, 1, float(pn) if is_numeric else pn)
        if is_numeric:
            ws.cell(r, 1).number_format = "General"
        for offset, date in enumerate(dates):
            qty = values.get(date, 0.0)
            # 人工版的空白格是留空而非填 0，這裡保持一致
            if qty:
                ws.cell(r, 2 + offset, clean_number(qty))
        ws.cell(r, total_col, f"=SUM({first_letter}{r}:{last_letter}{r})")

    for cell in ws[1]:
        cell.font = Font(bold=True)
        cell.alignment = Alignment(horizontal="center")
    ws.freeze_panes = "B2"
    autosize(ws)
    ws.column_dimensions["A"].width = 24.83

    output_path.parent.mkdir(parents=True, exist_ok=True)
    wb.save(output_path)


# ---------------------------------------------------------------------------
# PP：樞紐快取解析
# ---------------------------------------------------------------------------


def _pivot_cache_parts(xlsx: zipfile.ZipFile) -> list[tuple[str, str]]:
    """列出活頁簿內所有 pivotCacheDefinition 及其對應的 records 檔。

    舊版寫死 pivotCacheDefinition1.xml；一旦來源檔含多個樞紐快取，編號就會
    對不上。這裡改成從 zip 目錄動態列舉，並用 rels 找到對應的 records。
    """
    names = xlsx.namelist()
    parts: list[tuple[str, str]] = []
    for definition in sorted(n for n in names if re.fullmatch(
        r"xl/pivotCache/pivotCacheDefinition\d+\.xml", n
    )):
        records = None
        rels_name = definition.replace(
            "pivotCache/", "pivotCache/_rels/"
        ) + ".rels"
        if rels_name in names:
            rels = ET.fromstring(xlsx.read(rels_name))
            for rel in rels:
                target = rel.attrib.get("Target", "")
                if "pivotCacheRecords" in target:
                    records = "xl/pivotCache/" + Path(target).name
                    break
        if records is None:
            guess = definition.replace("Definition", "Records")
            records = guess if guess in names else None
        if records:
            parts.append((definition, records))
    return parts


def parse_pivot_cache(pp_path: Path) -> dict:
    with zipfile.ZipFile(pp_path) as xlsx:
        parts = _pivot_cache_parts(xlsx)
        if not parts:
            raise SystemExit(f"{pp_path.name} 內找不到樞紐快取（pivotCache），無法取得逐料號明細")

        chosen = None
        for definition_name, records_name in parts:
            definition = ET.fromstring(xlsx.read(definition_name))
            fields = [
                cf.attrib.get("name", "")
                for cf in definition.find(f"{{{SPREADSHEET_NS}}}cacheFields")
            ]
            if any(f.strip() == "AVTC FG Part Number" for f in fields):
                chosen = (definition_name, records_name, definition, fields)
                break
        if chosen is None:
            raise SystemExit(
                "樞紐快取中找不到 'AVTC FG Part Number' 欄位，來源檔格式可能已變更"
            )
        definition_name, records_name, definition, fields = chosen

        shared_by_field: list[list[str]] = []
        for cache_field in definition.find(f"{{{SPREADSHEET_NS}}}cacheFields"):
            shared_items: list[str] = []
            shared_root = cache_field.find(f"{{{SPREADSHEET_NS}}}sharedItems")
            if shared_root is not None:
                for item in shared_root:
                    tag = item.tag.split("}", 1)[-1]
                    shared_items.append("" if tag == "m" else item.attrib.get("v", ""))
            shared_by_field.append(shared_items)

        records_root = ET.fromstring(xlsx.read(records_name))
        records: list[list[str]] = []
        for record in records_root.findall(f"{{{SPREADSHEET_NS}}}r"):
            row: list[str] = []
            for index, child in enumerate(record):
                tag = child.tag.split("}", 1)[-1]
                if tag == "x":
                    shared_index = int(child.attrib.get("v", "0"))
                    values = shared_by_field[index]
                    row.append(values[shared_index] if shared_index < len(values) else "")
                else:
                    row.append("" if tag == "m" else child.attrib.get("v", ""))
            records.append(row)

        refreshed_raw = definition.attrib.get("refreshedDate")
        refreshed_by = definition.attrib.get("refreshedBy", "")
        refreshed_date = None
        if refreshed_raw:
            try:
                refreshed_date = serial_to_date(float(refreshed_raw))
            except ValueError:
                refreshed_date = None

        source = definition.find(f"{{{SPREADSHEET_NS}}}cacheSource")
        source_sheet = ""
        if source is not None:
            ws_source = source.find(f"{{{SPREADSHEET_NS}}}worksheetSource")
            if ws_source is not None:
                source_sheet = ws_source.attrib.get("sheet", "")

        return {
            "fields": fields,
            "records": records,
            "definition_part": definition_name,
            "refreshed_date": refreshed_date,
            "refreshed_by": refreshed_by,
            "source_sheet": source_sheet,
        }


# ---------------------------------------------------------------------------
# PP：期間欄位推導
# ---------------------------------------------------------------------------

WEEK_LABEL_RE = re.compile(r"^WK\s*(\d{1,2})(?:\s+([A-Za-z]{3,9}))?\s*'?$", re.IGNORECASE)
MONTH_FCST_LABEL_RE = re.compile(r"^([A-Za-z]{3})\s*'?\s*(\d{2})\s*FCST$", re.IGNORECASE)
MONTH_PLAIN_LABEL_RE = re.compile(r"^([A-Za-z]{3})\s*-\s*(\d{2})\s*'?$", re.IGNORECASE)
TOTAL_LABEL_RE = re.compile(r"total", re.IGNORECASE)

CACHE_WEEK_RE = re.compile(r"^WK\s*(\d{1,2})\s+(\d{2})'([A-Za-z]{3})", re.IGNORECASE)
CACHE_MONTH_RE = re.compile(r"^([A-Za-z]{3})\s*(?:'|-)\s*(\d{2})\s*(?:FCST)?$", re.IGNORECASE)


def normalize_field(name: str) -> str:
    return name.replace("_x000a_", " ").replace("\n", " ").strip()


def index_cache_periods(fields: Sequence[str]) -> tuple[dict, dict]:
    """建立 (年,週) -> 欄位索引清單、(年,月) -> 欄位索引 兩張對照表。"""
    weeks: dict[tuple[str, int], list[int]] = defaultdict(list)
    months: dict[tuple[str, str], int] = {}
    for idx, raw in enumerate(fields):
        name = normalize_field(raw)
        m = CACHE_WEEK_RE.match(name)
        if m:
            weeks[(m.group(2), int(m.group(1)))].append(idx)
            continue
        m = CACHE_MONTH_RE.match(name)
        if m and m.group(1).lower() in MONTH_INDEX:
            months.setdefault((m.group(2), m.group(1).title()), idx)
    return dict(weeks), months


def find_layout_row(wb) -> tuple[object, int, int] | None:
    """找出可見樞紐報表中，描述期間欄位版面的那一列。

    這一列（例：`WK27 Jul | WK28 | ... | Oct-26 | Nov26FCST | 2026 TOTAL | Jan'27 FCST`）
    就是人工整理版的欄位藍本，靠它推導「週明細到哪個月為止、之後改用月預測」，
    不必把 30~44 週、Nov/Dec、'27 這些寫死在程式裡。
    """
    best = None
    for ws in wb.worksheets:
        for row in ws.iter_rows(min_row=1, max_row=40):
            hits = 0
            first_col = None
            for cell in row:
                text = str(cell.value).strip() if cell.value is not None else ""
                if not text:
                    continue
                if (
                    WEEK_LABEL_RE.match(text)
                    or MONTH_FCST_LABEL_RE.match(text)
                    or MONTH_PLAIN_LABEL_RE.match(text)
                ):
                    hits += 1
                    if first_col is None:
                        first_col = cell.column
            if hits >= 8 and (best is None or hits > best[3]):
                best = (ws, row[0].row, first_col, hits)
    if best is None:
        return None
    return best[0], best[1], best[2]


def build_pp_periods(
    fields: Sequence[str],
    layout: list[str] | None,
    base_year: str,
    start_week: int,
) -> list[tuple[str, list[int]]]:
    """回傳 [(輸出欄名, [快取欄位索引...])]。"""
    weeks, months = index_cache_periods(fields)

    if not layout:
        # 沒有可見版面可參考時的退路：週明細只取到最後一個「有對應月份且
        # 月份在起始週所屬月份之後 3 個月內」的週；其餘用月預測。
        warn("找不到可見樞紐報表的欄位版面，改用推導模式（週明細取 15 週）。")
        layout = [f"WK{w:02d}" for w in range(start_week, start_week + 15)]
        covered_months = set()
        for w in range(start_week, start_week + 15):
            for idx in weeks.get((base_year, w), []):
                m = CACHE_WEEK_RE.match(normalize_field(fields[idx]))
                if m:
                    covered_months.add(m.group(3).title())
        for month in MONTH_ABBR:
            if (base_year, month) in months and month not in covered_months:
                if MONTH_INDEX[month.lower()] > max(
                    (MONTH_INDEX[c.lower()] for c in covered_months), default=0
                ):
                    layout.append(f"{month}{base_year}FCST")
        next_year = f"{int(base_year) + 1:02d}"
        for month in MONTH_ABBR:
            if (next_year, month) in months:
                layout.append(f"{month}'{next_year} FCST")

    # 第一輪：走一遍版面，判斷每個週欄屬於哪個月、哪些月份有週明細
    tokens: list[tuple[str, object, str]] = []  # (kind, key, label)
    current_month: str | None = None
    weekly_months: set[tuple[str, str]] = set()

    for label in layout:
        text = label.strip()
        if not text:
            continue
        if TOTAL_LABEL_RE.search(text) and not WEEK_LABEL_RE.match(text):
            tokens.append(("total", None, text))
            current_month = None
            continue
        m = WEEK_LABEL_RE.match(text)
        if m:
            week = int(m.group(1))
            hint = m.group(2)
            month = None
            if hint and hint[:3].lower() in MONTH_INDEX:
                month = hint[:3].title()
            else:
                candidates = []
                for idx in weeks.get((base_year, week), []):
                    cm = CACHE_WEEK_RE.match(normalize_field(fields[idx]))
                    if cm:
                        candidates.append(cm.group(3).title())
                if candidates:
                    month = current_month if current_month in candidates else candidates[0]
            if month:
                current_month = month
                weekly_months.add((base_year, month))
            tokens.append(("week", week, f"WK{week:02d}"))
            continue
        m = MONTH_FCST_LABEL_RE.match(text) or MONTH_PLAIN_LABEL_RE.match(text)
        if m:
            month = m.group(1).title()
            year = m.group(2)
            tokens.append(("month", (year, month), text))
            current_month = None
            continue
        tokens.append(("other", None, text))

    # 第二輪：從起始週開始輸出，跳過月小計欄與年度合計欄
    periods: "OrderedDict[str, list[int]]" = OrderedDict()
    started = False
    for kind, key, label in tokens:
        if not started:
            if kind == "week" and key == start_week:
                started = True
            else:
                continue
        if kind == "week":
            idxs = weeks.get((base_year, key), [])
            if not idxs:
                warn(f"快取中找不到 {base_year}' 年的 WK{key:02d}，該欄以 0 輸出。")
            periods.setdefault(label, [])
            for idx in idxs:
                if idx not in periods[label]:
                    periods[label].append(idx)
        elif kind == "month":
            if key in weekly_months:
                continue  # 該月已有週明細，這是小計欄，不重複輸出
            idx = months.get(key)
            if idx is None:
                warn(f"快取中找不到月份欄位 {key[1]}-{key[0]}，略過。")
                continue
            periods.setdefault(label, [idx])

    if not started:
        raise SystemExit(
            f"版面中找不到起始週 WK{start_week:02d}；請用 --pp-start-week 指定正確的起始週。"
        )
    return list(periods.items())


def read_layout(pp_path: Path) -> tuple[list[str] | None, list[str]]:
    """回傳 (期間欄位標籤序列, 客戶顯示順序)。"""
    wb = load_workbook(pp_path, data_only=True)
    try:
        found = find_layout_row(wb)
        if found is None:
            return None, []
        ws, row_idx, first_col = found
        labels = [
            str(cell.value).strip()
            for cell in ws[row_idx]
            if cell.column >= first_col and cell.value is not None and str(cell.value).strip()
        ]
        customers: list[str] = []
        for row in ws.iter_rows(min_row=row_idx + 1, max_col=max(1, first_col - 1)):
            for cell in row:
                text = str(cell.value).strip() if cell.value is not None else ""
                if text and "TTL" not in text.upper() and text not in customers:
                    customers.append(text)
        return labels, customers
    finally:
        wb.close()


def generate_pp(
    pp_path: Path,
    output_path: Path,
    plan: str = "Production Input",
    start_week: int | None = None,
    base_year: str | None = None,
    report_date: dt.date | None = None,
) -> dict:
    cache = parse_pivot_cache(pp_path)
    fields = cache["fields"]
    records = cache["records"]

    field_index = {normalize_field(name): idx for idx, name in enumerate(fields)}
    required = ["Customer", "Model", "AVTC FG Part Number", "Plan"]
    missing = [name for name in required if name not in field_index]
    if missing:
        raise SystemExit(f"樞紐快取缺少必要欄位：{', '.join(missing)}")

    customer_idx = field_index["Customer"]
    model_idx = field_index["Model"]
    pn_idx = field_index["AVTC FG Part Number"]
    plan_idx = field_index["Plan"]

    if report_date is None:
        report_date = cache["refreshed_date"] or dt.date.today()
    if base_year is None:
        base_year = f"{report_date.year % 100:02d}"
    if start_week is None:
        # 取報表日所在週的前一週開始（與人工整理慣例一致：保留上一週作參照）
        start_week = max(1, report_date.isocalendar()[1] - 1)

    layout_labels, customer_order = read_layout(pp_path)
    periods = build_pp_periods(fields, layout_labels, base_year, start_week)
    if not periods:
        raise SystemExit("推導不出任何輸出期間欄位")

    plans_seen = {r[plan_idx].strip() for r in records if len(r) > plan_idx}
    if plan not in plans_seen:
        raise SystemExit(
            f"快取中沒有 Plan = {plan!r}；可選值：{', '.join(sorted(plans_seen))}"
        )

    aggregate: dict[str, dict[str, float]] = defaultdict(lambda: defaultdict(float))
    metadata: dict[str, tuple[str, str]] = {}
    source_order: list[str] = []
    plan_rows = 0

    for record in records:
        if len(record) <= max(customer_idx, model_idx, pn_idx, plan_idx):
            continue
        if record[plan_idx].strip() != plan:
            continue
        pn = record[pn_idx].strip()
        if not pn:
            continue
        plan_rows += 1
        if pn not in metadata:
            metadata[pn] = (record[customer_idx].strip(), record[model_idx].strip())
            source_order.append(pn)
        for label, source_indexes in periods:
            aggregate[pn][label] += sum(
                numeric(record[i]) for i in source_indexes if i < len(record)
            )

    labels = [label for label, _ in periods]
    kept = [pn for pn in source_order if any(aggregate[pn][label] for label in labels)]

    rank = {name: i for i, name in enumerate(customer_order)}
    fallback = len(rank)
    kept.sort(key=lambda pn: (rank.get(metadata[pn][0], fallback), metadata[pn][0], pn))

    rows = []
    for pn in kept:
        customer, model = metadata[pn]
        rows.append(
            [customer, pn, model]
            + [clean_number(aggregate[pn][label]) for label in labels]
        )

    write_pp_workbook(output_path, labels, rows)

    return {
        "source": pp_path,
        "cache_part": cache["definition_part"],
        "cache_source_sheet": cache["source_sheet"],
        "refreshed_date": cache["refreshed_date"],
        "refreshed_by": cache["refreshed_by"],
        "records": len(records),
        "plan": plan,
        "plan_rows": plan_rows,
        "base_year": base_year,
        "start_week": start_week,
        "report_date": report_date,
        "layout_found": layout_labels is not None,
        "periods": labels,
        "rows": len(rows),
        "dropped_zero": len(metadata) - len(kept),
        "grand_total": clean_number(sum(sum(r[3:]) for r in rows)),
    }


PP_TIDY_SHEET = "整理后PP"
PP_SPACER_COLS = 3  # 人工版在 total 前留了三個空白欄，這裡照樣保留以維持版面一致


def write_pp_workbook(output_path: Path, labels: Sequence[str], rows: Sequence[list]) -> None:
    wb = Workbook()
    ws = wb.active
    ws.title = PP_TIDY_SHEET

    ws.cell(1, 1, "Customer")
    ws.cell(1, 2, "AVTC FG Part Number")
    ws.cell(1, 3, "Model")
    for offset, label in enumerate(labels):
        ws.cell(1, 4 + offset, label)
    first_col = 4
    last_col = 3 + len(labels)
    total_col = last_col + PP_SPACER_COLS + 1
    ws.cell(1, total_col, "total")

    first_letter = get_column_letter(first_col)
    last_letter = get_column_letter(last_col)

    for r_offset, row in enumerate(rows):
        r = 2 + r_offset
        ws.cell(r, 1, row[0])
        ws.cell(r, 2, row[1])
        ws.cell(r, 3, row[2])
        for offset, value in enumerate(row[3:]):
            cell = ws.cell(r, 4 + offset, value)
            cell.number_format = "#,##0_ "
        total = ws.cell(r, total_col, f"=SUM({first_letter}{r}:{last_letter}{r})")
        total.number_format = "#,##0_ "

    for cell in ws[1]:
        if cell.value is not None:
            cell.font = Font(bold=True)
            cell.alignment = Alignment(horizontal="center")
    ws.freeze_panes = "D2"
    autosize(ws)
    ws.column_dimensions["A"].width = 17.0
    ws.column_dimensions["B"].width = 28.6
    ws.column_dimensions["C"].width = 20.7
    for offset in range(len(labels)):
        ws.column_dimensions[get_column_letter(4 + offset)].width = 13.0

    output_path.parent.mkdir(parents=True, exist_ok=True)
    wb.save(output_path)


# ---------------------------------------------------------------------------
# 對帳（--compare）
# ---------------------------------------------------------------------------


def _load_grid(path: Path, sheet: str, key_col: int, first_data_col: int, label_row: int = 1):
    wb = load_workbook(path, data_only=True)
    try:
        if sheet not in wb.sheetnames:
            return None
        ws = wb[sheet]
        labels = {}
        for cell in ws[label_row]:
            if cell.column < first_data_col or cell.value is None:
                continue
            text = str(cell.value).strip()
            if text.lower() == "total":
                continue
            date = header_date(cell)
            labels[cell.column] = date if date is not None else text
        grid = {}
        for row in ws.iter_rows(min_row=label_row + 1):
            key = row[key_col - 1].value
            if key is None or not str(key).strip():
                continue
            grid[str(key).strip()] = {
                labels[cell.column]: numeric(cell.value)
                for cell in row
                if cell.column in labels
            }
        return grid
    finally:
        wb.close()


def compare(
    title: str,
    manual_path: Path,
    manual_sheet: str,
    generated_path: Path,
    generated_sheet: str,
    key_col: int,
    first_data_col_manual: int,
    first_data_col_generated: int,
    limit: int = 30,
) -> None:
    log(f"\n=== 對帳：{title} ===")
    manual = _load_grid(manual_path, manual_sheet, key_col, first_data_col_manual)
    if manual is None:
        log(f"  來源檔內沒有 {manual_sheet} 工作表，略過對帳。")
        return
    generated = _load_grid(generated_path, generated_sheet, key_col, first_data_col_generated)
    if generated is None:
        log("  產出檔讀取失敗，略過對帳。")
        return

    only_manual = sorted(set(manual) - set(generated))
    only_generated = sorted(set(generated) - set(manual))
    log(f"  人工版 {len(manual)} 列 / 本次產出 {len(generated)} 列")
    if only_generated:
        log(f"  ▲ 人工版漏記（原始有、人工無）共 {len(only_generated)} 筆：")
        for key in only_generated[:limit]:
            values = {k: v for k, v in generated[key].items() if v}
            log(f"      + {key}  {values}")
    if only_manual:
        log(f"  ▼ 人工版多出（原始無）共 {len(only_manual)} 筆：")
        for key in only_manual[:limit]:
            log(f"      - {key}")

    diffs = []
    for key in sorted(set(manual) & set(generated)):
        for label in sorted(set(manual[key]) | set(generated[key]), key=str):
            a = manual[key].get(label, 0.0)
            b = generated[key].get(label, 0.0)
            if abs(a - b) > 1e-6:
                diffs.append((key, label, a, b))
    log(f"  數值差異：{len(diffs)} 格")
    for key, label, a, b in diffs[:limit]:
        log(f"      ! {key} | {label} | 人工={clean_number(a)} 原始={clean_number(b)}")
    if len(diffs) > limit:
        log(f"      ...（其餘 {len(diffs) - limit} 筆省略）")


# ---------------------------------------------------------------------------
# CLI
# ---------------------------------------------------------------------------


def build_parser(project_root: Path) -> argparse.ArgumentParser:
    parser = argparse.ArgumentParser(
        description="由 DPS / PP 原始資料生成整理后報表（數值一律以原始檔為準）。",
        formatter_class=argparse.ArgumentDefaultsHelpFormatter,
    )
    parser.add_argument("--input-dir", type=Path, default=project_root / "inout",
                        help="輸入資料夾")
    parser.add_argument("--out-dir", type=Path, default=project_root / "output",
                        help="輸出資料夾（不存在會自動建立）")
    parser.add_argument("--dps", type=Path, default=None,
                        help="DPS 來源檔；省略時於輸入資料夾自動尋找")
    parser.add_argument("--pp", type=Path, default=None,
                        help="PP 來源檔；省略時於輸入資料夾自動尋找")
    parser.add_argument("--skip-dps", action="store_true", help="不產生 DPS 報表")
    parser.add_argument("--skip-pp", action="store_true", help="不產生 PP 報表")

    parser.add_argument("--include-star-parts", action="store_true",
                        help="保留結尾帶 * 的 DPS 料號（預設排除，與人工整理規則一致）")
    parser.add_argument("--dps-tail-cutoff", default="auto",
                        help="DPS 末欄彙總桶起始日：auto / none / YYYY-MM-DD")

    parser.add_argument("--pp-plan", default="Production Input", help="PP 的 Plan 篩選值")
    parser.add_argument("--pp-start-week", default="auto",
                        help="PP 起始週：auto（依報表日推算）或週數")
    parser.add_argument("--pp-base-year", default=None,
                        help="PP 主年度兩位數（例 26）；省略時依報表日判斷")
    parser.add_argument("--pp-report-date", type=parse_date_arg, default=None,
                        help="PP 報表基準日；省略時取樞紐快取的更新日期")

    parser.add_argument("--compare", action="store_true",
                        help="與來源檔內既有的人工整理版逐格對帳並列出差異")
    parser.add_argument("--quiet", action="store_true", help="只輸出錯誤訊息")
    return parser


def main() -> None:
    global _LOG_QUIET

    project_root = Path(__file__).resolve().parent
    args = build_parser(project_root).parse_args()
    _LOG_QUIET = args.quiet

    out_dir: Path = args.out_dir
    out_dir.mkdir(parents=True, exist_ok=True)

    log("=" * 72)
    log("Buyer Reports 產生器")
    log(f"輸入資料夾：{args.input_dir}")
    log(f"輸出資料夾：{out_dir}")
    log("=" * 72)

    if not args.skip_dps:
        dps_path = args.dps or find_input(args.input_dir, [r"DPS"], "DPS")
        if not dps_path.is_file():
            raise SystemExit(f"找不到 DPS 檔案：{dps_path}")

        if args.dps_tail_cutoff.lower() == "none":
            cutoff = None
        elif args.dps_tail_cutoff.lower() == "auto":
            wb = load_workbook(dps_path, data_only=True)
            try:
                ws = wb[DPS_SOURCE_SHEET] if DPS_SOURCE_SHEET in wb.sheetnames else None
                max_date = None
                if ws is not None:
                    header_row = find_header_row(ws, DPS_HEADER_KEYS)
                    dates = [d for d in (header_date(c) for c in ws[header_row]) if d]
                    max_date = max(dates) if dates else None
            finally:
                wb.close()
            cutoff = detect_dps_tail_cutoff(dps_path, max_date) if max_date else None
        else:
            cutoff = parse_date_arg(args.dps_tail_cutoff)

        dps_out = out_dir / "DPS整理后.xlsx"
        info = generate_dps(
            dps_path,
            dps_out,
            include_star_parts=args.include_star_parts,
            tail_cutoff=cutoff,
        )
        log("\n--- DPS ---")
        log(f"  來源            ：{info['source'].name}")
        log(f"  表頭列          ：第 {info['header_row']} 列")
        log(f"  日期欄          ：{info['date_columns']} 欄 / {info['dates']} 個日期"
            f"（{info['date_range'][0]} ~ {info['date_range'][1]}，D+N 兩班合併）")
        if info["tail_cutoff"]:
            log(f"  末欄彙總桶      ：{info['tail_cutoff']} 起之日期併入同一欄"
                f"（沿用既有整理后版面）")
        else:
            log("  末欄彙總桶      ：未使用，每個日期各自成欄")
        log(f"  輸出            ：{info['rows']} 列 x {info['out_columns']} 個日期欄，"
            f"合計 {info['grand_total']:,} pcs")
        if info["excluded"]:
            total = clean_number(sum(info["excluded"].values()))
            log(f"  已排除 * 料號   ：{len(info['excluded'])} 個，合計 {total:,} pcs"
                f"（用 --include-star-parts 可保留）")
            for pn, qty in sorted(info["excluded"].items()):
                log(f"      - {pn}  {clean_number(qty):,}")
        if info["text_cells"]:
            log(f"  日期區文字格    ：{info['text_cells']} 格（已當 0 計）")
        log(f"  產出檔          ：{dps_out}")

        if args.compare:
            compare(
                "DPS", dps_path, DPS_TIDY_SHEET, dps_out, DPS_TIDY_SHEET,
                key_col=1, first_data_col_manual=2, first_data_col_generated=2,
            )

    if not args.skip_pp:
        pp_path = args.pp or find_input(args.input_dir, [r"\bPP\b", r"PP"], "PP")
        if not pp_path.is_file():
            raise SystemExit(f"找不到 PP 檔案：{pp_path}")

        start_week = None if args.pp_start_week.lower() == "auto" else int(args.pp_start_week)
        pp_out = out_dir / "整理后PP.xlsx"
        info = generate_pp(
            pp_path,
            pp_out,
            plan=args.pp_plan,
            start_week=start_week,
            base_year=args.pp_base_year,
            report_date=args.pp_report_date,
        )
        log("\n--- PP ---")
        log(f"  來源            ：{info['source'].name}")
        log(f"  樞紐快取        ：{info['cache_part']}"
            f"（原始表 {info['cache_source_sheet'] or '未知'}，{info['records']} 筆）")
        log(f"  快取更新        ：{info['refreshed_date']} by {info['refreshed_by'] or '未知'}")
        log(f"  報表基準日      ：{info['report_date']}"
            f"（主年度 20{info['base_year']}，起始週 WK{info['start_week']:02d}）")
        log(f"  欄位版面        ：{'取自可見樞紐報表' if info['layout_found'] else '推導模式'}")
        log(f"  Plan 篩選       ：{info['plan']}（{info['plan_rows']} 筆料號）")
        log(f"  期間欄          ：{len(info['periods'])} 欄 → {', '.join(info['periods'])}")
        log(f"  輸出            ：{info['rows']} 列"
            f"（已略過期間內全為 0 的 {info['dropped_zero']} 個料號），"
            f"合計 {info['grand_total']:,} pcs")
        log(f"  產出檔          ：{pp_out}")

        if info["refreshed_date"] and info["refreshed_date"] < dt.date.today() - dt.timedelta(days=45):
            warn(
                f"PP 樞紐快取最後更新於 {info['refreshed_date']}，距今已超過 45 天，"
                "數字可能是舊快照，請向提供者確認是否已 refresh。"
            )

        if args.compare:
            compare(
                "PP", pp_path, PP_TIDY_SHEET, pp_out, PP_TIDY_SHEET,
                key_col=2, first_data_col_manual=4, first_data_col_generated=4,
            )

    log("\n完成。")


if __name__ == "__main__":
    main()
