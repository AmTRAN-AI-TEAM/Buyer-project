"""Common runtime and Excel helpers for Buyer Reports."""

from __future__ import annotations

import argparse
import configparser
import datetime as dt
import os
import re
import sys
import tempfile
from copy import copy
from pathlib import Path
from typing import Iterable, Sequence

from openpyxl.utils import get_column_letter

try:
    from tqdm import tqdm
except ImportError:  # pragma: no cover - 打包時 requirements 會安裝；這裡保留退路
    tqdm = None

EXCEL_EPOCH = dt.date(1899, 12, 30)
CONFIG_FILE_NAME = "buyer_reports.ini"
CTB_ETA_CONFIG_FILE_NAME = "ctb_eta_days.ini"
CTB_ETA_CONFIG_SECTION = "ctb_eta"
CTB_ETA_SITE_SECTION = "supplier_site"
CTB_ETA_NEW_SITE_SECTION = "supplier_site_new"
DEFAULT_CTB_ETA_LEAD_DAYS = 15
DEFAULT_DPS_SHEET_KEYWORDS = ("DPS",)
DEFAULT_PP_SHEET_KEYWORDS = ("PP", "Data")
DEFAULT_DPS_PART_NUMBER_HEADERS = ("AVTC P/N", "P/N", "Model")
DEFAULT_PP_PART_NUMBER_FIELD_KEYWORDS = ("Part Number",)
DEFAULT_CUSTOMERS = ("AVTC", "RAKEN")
DEFAULT_CUSTOMER_MODES = {
    "AVTC": {
        "dps_mode": "first_valid",
        "pp_mode": "first_valid",
        "dps_pp_dps_weeks_ahead": 5,
        "dps_pp_late_dps_mode": "merge_to_cutoff",
        "dps_drop_zero_total_rows": False,
        "dps_trim_trailing_zero_dates": False,
    },
    "RAKEN": {
        "dps_mode": "merge_all",
        "pp_mode": "first_valid",
        "dps_pp_dps_weeks_ahead": 2,
        "dps_pp_late_dps_mode": "drop",
        "dps_drop_zero_total_rows": True,
        "dps_trim_trailing_zero_dates": True,
    },
}
VALID_REPORT_MODES = ("first_valid", "merge_all")
VALID_DPS_PP_LATE_DPS_MODES = ("merge_to_cutoff", "drop")

# ---------------------------------------------------------------------------
# 共用小工具
# ---------------------------------------------------------------------------

_LOG_QUIET = False
_LOG_FILES = []
_PROGRESS_BAR = None


def set_log_quiet(enabled: bool) -> None:
    global _LOG_QUIET
    _LOG_QUIET = enabled


def log(message: str = "") -> None:
    if not _LOG_QUIET:
        if _PROGRESS_BAR is not None:
            _PROGRESS_BAR.write(message)
        else:
            print(message, flush=True)
    for log_file in _LOG_FILES:
        log_file.write(message + "\n")
        log_file.flush()


def warn(message: str) -> None:
    text = f"[警告] {message}"
    if _PROGRESS_BAR is not None:
        _PROGRESS_BAR.write(text, file=sys.stderr)
    else:
        print(text, file=sys.stderr, flush=True)
    for log_file in _LOG_FILES:
        log_file.write(text + "\n")
        log_file.flush()


class Progress:
    def __init__(self, total: int):
        self.total = total
        self.bar = None

    def __enter__(self):
        global _PROGRESS_BAR
        if tqdm is not None and not _LOG_QUIET and self.total > 0 and sys.stdout.isatty():
            self.bar = tqdm(
                total=self.total,
                desc="Buyer Reports",
                unit="step",
                dynamic_ncols=True,
                leave=True,
            )
            _PROGRESS_BAR = self.bar
        return self

    def step(self, label: str) -> None:
        if self.bar is not None:
            self.bar.set_description_str(label)
            self.bar.update(1)

    def __exit__(self, exc_type, exc, tb):
        global _PROGRESS_BAR
        if self.bar is not None:
            self.bar.close()
        _PROGRESS_BAR = None


def is_frozen_app() -> bool:
    return bool(getattr(sys, "frozen", False))


def project_root() -> Path:
    if is_frozen_app():
        return Path(sys.executable).resolve().parent
    return Path(__file__).resolve().parent.parent


def parse_keyword_list(value: str) -> tuple[str, ...]:
    keywords = [part.strip() for part in re.split(r"[,;\n]+", value) if part.strip()]
    return tuple(dict.fromkeys(keywords))


def parse_report_mode(value: str, fallback: str, label: str) -> str:
    mode = value.strip().lower()
    if not mode:
        return fallback
    if mode not in VALID_REPORT_MODES:
        warn(
            f"{label} 模式 {value!r} 不支援，已改用 {fallback!r}。"
            f"可用模式：{', '.join(VALID_REPORT_MODES)}"
        )
        return fallback
    return mode


def parse_dps_pp_late_dps_mode(value: str, fallback: str, label: str) -> str:
    mode = value.strip().lower().replace("-", "_")
    aliases = {
        "merge": "merge_to_cutoff",
        "merge_to_cutoff": "merge_to_cutoff",
        "cutoff": "merge_to_cutoff",
        "drop": "drop",
        "ignore": "drop",
        "truncate": "drop",
    }
    if not mode:
        return fallback
    parsed = aliases.get(mode)
    if parsed is None:
        warn(
            f"{label} 設定 {value!r} 不支援，已改用 {fallback!r}。"
            f"可用模式：{', '.join(VALID_DPS_PP_LATE_DPS_MODES)}"
        )
        return fallback
    return parsed


def parse_positive_int(value: str, fallback: int, label: str) -> int:
    try:
        parsed = int(value.strip())
    except ValueError:
        warn(f"{label} 設定 {value!r} 不是整數，已改用 {fallback}。")
        return fallback
    if parsed < 0:
        warn(f"{label} 設定 {value!r} 小於 0，已改用 {fallback}。")
        return fallback
    return parsed


def parse_bool(value: str, fallback: bool, label: str) -> bool:
    normalized = value.strip().casefold()
    if not normalized:
        return fallback
    if normalized in {"1", "yes", "true", "on", "y"}:
        return True
    if normalized in {"0", "no", "false", "off", "n"}:
        return False
    warn(f"{label} 設定 {value!r} 不是 true/false，已改用 {fallback}。")
    return fallback


def _supplier_site_config_key(value) -> str:
    return "" if value is None else str(value).strip().casefold()


def _supplier_site_config_display(value) -> str:
    return "" if value is None else str(value).strip()


def _read_ctb_eta_site_section(
    parser: configparser.ConfigParser,
    section: str,
    default_days: int,
) -> tuple[dict[str, tuple[str, int]], bool]:
    entries: dict[str, tuple[str, int]] = {}
    had_invalid_value = False
    if not parser.has_section(section):
        return entries, had_invalid_value
    for raw_site, raw_days in parser.items(section):
        site = _supplier_site_config_display(raw_site)
        key = _supplier_site_config_key(site)
        if not key:
            had_invalid_value = True
            warn(f"CTB ETA 設定忽略空白 Supplier Site。")
            continue
        days = parse_positive_int(
            raw_days,
            default_days,
            f"CTB ETA Supplier Site {site}",
        )
        if str(raw_days).strip() != str(days):
            had_invalid_value = True
        entries[key] = (site, days)
    return entries, had_invalid_value


def _read_ctb_eta_config_file(
    path: Path,
) -> tuple[int, dict[str, tuple[str, int]], dict[str, tuple[str, int]], bool]:
    default_days = DEFAULT_CTB_ETA_LEAD_DAYS
    main_entries: dict[str, tuple[str, int]] = {}
    new_entries: dict[str, tuple[str, int]] = {}
    needs_normalization = not path.is_file()
    if not path.is_file():
        return default_days, main_entries, new_entries, needs_normalization

    parser = configparser.ConfigParser(interpolation=None)
    parser.optionxform = str
    try:
        parser.read(path, encoding="utf-8")
    except configparser.Error as exc:
        raise SystemExit(f"{path.name} 格式錯誤，請修正後再執行：{exc}") from exc

    raw_default = None
    if parser.has_section(CTB_ETA_CONFIG_SECTION):
        raw_default = parser.get(
            CTB_ETA_CONFIG_SECTION,
            "default_lead_days",
            fallback=None,
        )
    default_days = parse_positive_int(
        raw_default if raw_default is not None else str(DEFAULT_CTB_ETA_LEAD_DAYS),
        DEFAULT_CTB_ETA_LEAD_DAYS,
        "CTB ETA default_lead_days",
    )
    if raw_default is None or str(raw_default).strip() != str(default_days):
        needs_normalization = True

    main_entries, main_invalid = _read_ctb_eta_site_section(
        parser,
        CTB_ETA_SITE_SECTION,
        default_days,
    )
    new_entries, new_invalid = _read_ctb_eta_site_section(
        parser,
        CTB_ETA_NEW_SITE_SECTION,
        default_days,
    )
    return default_days, main_entries, new_entries, needs_normalization or main_invalid or new_invalid


def _write_ctb_eta_config(
    path: Path,
    default_days: int,
    main_entries: dict[str, tuple[str, int]],
    new_entries: dict[str, tuple[str, int]],
) -> None:
    divider = "# " + "＝" * 34
    lines = [
        f"[{CTB_ETA_CONFIG_SECTION}]",
        "# CTB ETA 預設提前天數；未列出的 Supplier Site 使用此值。",
        f"default_lead_days = {default_days}",
        "",
        f"[{CTB_ETA_SITE_SECTION}]",
        "# 已確認的 Supplier Site。",
    ]
    lines.extend(
        f"{display} = {days}"
        for display, days in main_entries.values()
    )
    lines.extend(
        [
            "",
            divider,
            "# 本次新偵測的 Supplier Site；下一次執行時會自動移到上方。",
            divider,
            f"[{CTB_ETA_NEW_SITE_SECTION}]",
        ]
    )
    lines.extend(
        f"{display} = {days}"
        for display, days in new_entries.values()
    )
    path.parent.mkdir(parents=True, exist_ok=True)
    temp_path = path.with_name(f".{path.name}.tmp")
    temp_path.write_text("\n".join(lines) + "\n", encoding="utf-8")
    os.replace(temp_path, path)


def sync_ctb_eta_config(root: Path, supplier_sites: Sequence[str]) -> dict:
    """Sync detected Supplier Sites while preserving saved ETA lead times."""
    path = root / CTB_ETA_CONFIG_FILE_NAME
    default_days, main_entries, pending_entries, needs_normalization = _read_ctb_eta_config_file(path)
    promoted_sites = []
    for key, entry in pending_entries.items():
        if key not in main_entries:
            main_entries[key] = entry
            promoted_sites.append(entry[0])
    pending_entries = {}

    detected: dict[str, str] = {}
    for raw_site in supplier_sites:
        display = _supplier_site_config_display(raw_site)
        key = _supplier_site_config_key(display)
        if key:
            detected.setdefault(key, display)
    for key, display in sorted(detected.items()):
        if key not in main_entries:
            pending_entries[key] = (display, default_days)

    new_sites = [display for display, _days in pending_entries.values()]
    changed = (
        needs_normalization
        or bool(promoted_sites)
        or bool(new_sites)
        or not path.is_file()
    )
    if changed:
        _write_ctb_eta_config(path, default_days, main_entries, pending_entries)

    active_entries = dict(pending_entries)
    active_entries.update(main_entries)
    return {
        "path": path,
        "default_lead_days": default_days,
        "lead_days_by_supplier_site": {
            key: days for key, (_display, days) in active_entries.items()
        },
        "detected_supplier_sites": tuple(detected.values()),
        "new_supplier_sites": tuple(new_sites),
        "promoted_supplier_sites": tuple(promoted_sites),
        "changed": changed,
    }


def load_ctb_eta_config(path: Path) -> dict:
    """Read the current CTB ETA settings after optional user editing."""
    default_days, main_entries, pending_entries, _needs_normalization = _read_ctb_eta_config_file(path)
    active_entries = dict(pending_entries)
    active_entries.update(main_entries)
    return {
        "path": path,
        "default_lead_days": default_days,
        "lead_days_by_supplier_site": {
            key: days for key, (_display, days) in active_entries.items()
        },
        "new_supplier_sites": tuple(display for display, _days in pending_entries.values()),
        "promoted_supplier_sites": (),
        "detected_supplier_sites": (),
        "changed": False,
    }


def load_sheet_detection_config(root: Path) -> dict:
    config_path = root / CONFIG_FILE_NAME
    result = {
        "path": config_path,
        "loaded": False,
        "dps_sheet_keywords": DEFAULT_DPS_SHEET_KEYWORDS,
        "pp_sheet_keywords": DEFAULT_PP_SHEET_KEYWORDS,
        "dps_part_number_headers": DEFAULT_DPS_PART_NUMBER_HEADERS,
        "pp_part_number_field_keywords": DEFAULT_PP_PART_NUMBER_FIELD_KEYWORDS,
        "customers": [
            {
                "name": name,
                "dps_mode": DEFAULT_CUSTOMER_MODES[name]["dps_mode"],
                "pp_mode": DEFAULT_CUSTOMER_MODES[name]["pp_mode"],
                "dps_pp_dps_weeks_ahead": DEFAULT_CUSTOMER_MODES[name][
                    "dps_pp_dps_weeks_ahead"
                ],
                "dps_pp_late_dps_mode": DEFAULT_CUSTOMER_MODES[name][
                    "dps_pp_late_dps_mode"
                ],
                "dps_drop_zero_total_rows": DEFAULT_CUSTOMER_MODES[name][
                    "dps_drop_zero_total_rows"
                ],
                "dps_trim_trailing_zero_dates": DEFAULT_CUSTOMER_MODES[name][
                    "dps_trim_trailing_zero_dates"
                ],
            }
            for name in DEFAULT_CUSTOMERS
        ],
    }
    if not config_path.is_file():
        return result

    parser = configparser.ConfigParser()
    parser.read(config_path, encoding="utf-8")
    result["loaded"] = True
    if parser.has_section("sheet_detection"):
        dps_keywords = parse_keyword_list(
            parser.get("sheet_detection", "dps_sheet_keywords", fallback="")
        )
        pp_keywords = parse_keyword_list(
            parser.get("sheet_detection", "pp_sheet_keywords", fallback="")
        )
        if dps_keywords:
            result["dps_sheet_keywords"] = dps_keywords
        if pp_keywords:
            result["pp_sheet_keywords"] = pp_keywords
    if parser.has_section("dps"):
        part_headers = parse_keyword_list(
            parser.get("dps", "part_number_headers", fallback="")
        )
        if part_headers:
            result["dps_part_number_headers"] = part_headers
    if parser.has_section("pp"):
        part_keywords = parse_keyword_list(
            parser.get("pp", "part_number_field_keywords", fallback="")
        )
        if part_keywords:
            result["pp_part_number_field_keywords"] = part_keywords
    customer_names = DEFAULT_CUSTOMERS
    if parser.has_section("customers"):
        configured_names = parse_keyword_list(parser.get("customers", "names", fallback=""))
        if configured_names:
            customer_names = configured_names
    customers = []
    section_map = {section.casefold(): section for section in parser.sections()}
    for name in customer_names:
        defaults = DEFAULT_CUSTOMER_MODES.get(
            name.upper(),
            {
                "dps_mode": "first_valid",
                "pp_mode": "first_valid",
                "dps_pp_dps_weeks_ahead": 5,
                "dps_pp_late_dps_mode": "merge_to_cutoff",
                "dps_drop_zero_total_rows": False,
                "dps_trim_trailing_zero_dates": False,
            },
        )
        section = section_map.get(f"customer.{name}".casefold())
        dps_mode = defaults["dps_mode"]
        pp_mode = defaults["pp_mode"]
        dps_pp_dps_weeks_ahead = defaults["dps_pp_dps_weeks_ahead"]
        dps_pp_late_dps_mode = defaults["dps_pp_late_dps_mode"]
        dps_drop_zero_total_rows = defaults["dps_drop_zero_total_rows"]
        dps_trim_trailing_zero_dates = defaults["dps_trim_trailing_zero_dates"]
        if section:
            dps_mode = parse_report_mode(
                parser.get(section, "dps_mode", fallback=dps_mode),
                defaults["dps_mode"],
                f"{name} DPS",
            )
            pp_mode = parse_report_mode(
                parser.get(section, "pp_mode", fallback=pp_mode),
                defaults["pp_mode"],
                f"{name} PP",
            )
            dps_pp_dps_weeks_ahead = parse_positive_int(
                parser.get(
                    section,
                    "dps_pp_dps_weeks_ahead",
                    fallback=str(dps_pp_dps_weeks_ahead),
                ),
                defaults["dps_pp_dps_weeks_ahead"],
                f"{name} DPS+PP 的 DPS 保留週數",
            )
            dps_pp_late_dps_mode = parse_dps_pp_late_dps_mode(
                parser.get(
                    section,
                    "dps_pp_late_dps_mode",
                    fallback=dps_pp_late_dps_mode,
                ),
                defaults["dps_pp_late_dps_mode"],
                f"{name} DPS+PP 截止日後 DPS 處理模式",
            )
            dps_drop_zero_total_rows = parse_bool(
                parser.get(
                    section,
                    "dps_drop_zero_total_rows",
                    fallback=str(dps_drop_zero_total_rows),
                ),
                defaults["dps_drop_zero_total_rows"],
                f"{name} DPS 零數量列略過",
            )
            dps_trim_trailing_zero_dates = parse_bool(
                parser.get(
                    section,
                    "dps_trim_trailing_zero_dates",
                    fallback=str(dps_trim_trailing_zero_dates),
                ),
                defaults["dps_trim_trailing_zero_dates"],
                f"{name} DPS 尾端空白日期欄略過",
            )
        customers.append({
            "name": name,
            "dps_mode": dps_mode,
            "pp_mode": pp_mode,
            "dps_pp_dps_weeks_ahead": dps_pp_dps_weeks_ahead,
            "dps_pp_late_dps_mode": dps_pp_late_dps_mode,
            "dps_drop_zero_total_rows": dps_drop_zero_total_rows,
            "dps_trim_trailing_zero_dates": dps_trim_trailing_zero_dates,
        })
    if customers:
        result["customers"] = customers
    return result


def keyword_label(keywords: Sequence[str]) -> str:
    return " 或 ".join(keywords)


def normalize_label(value) -> str:
    return re.sub(r"\s+", " ", str(value).strip()).casefold()


EXCEL_ESCAPED_WHITESPACE_RE = re.compile(
    r"_x(0009|000a|000d|0020|00a0)_",
    re.IGNORECASE,
)


def normalize_part_number(value) -> str:
    if value is None:
        return ""

    text = str(value)
    text = EXCEL_ESCAPED_WHITESPACE_RE.sub(
        lambda match: chr(int(match.group(1), 16)),
        text,
    )
    return text.strip()


def sheet_name_matches_keywords(sheet_name: str, keywords: Sequence[str]) -> bool:
    return any(re.search(re.escape(keyword), sheet_name, re.IGNORECASE) for keyword in keywords)


def path_is_under(path: Path, parent: Path) -> bool:
    try:
        path.resolve().relative_to(parent.resolve())
        return True
    except ValueError:
        return False


def running_from_windows_temp() -> bool:
    if not is_frozen_app() or not sys.platform.startswith("win"):
        return False
    exe_path = Path(sys.executable)
    temp_paths = {
        tempfile.gettempdir(),
        os.environ.get("TEMP", ""),
        os.environ.get("TMP", ""),
    }
    return any(
        temp and path_is_under(exe_path, Path(temp))
        for temp in temp_paths
    )


def show_windows_error(title: str, message: str) -> None:
    if not sys.platform.startswith("win"):
        return
    try:
        import ctypes
        ctypes.windll.user32.MessageBoxW(None, message, title, 0x10)
    except Exception:
        pass


def prevent_temp_execution() -> None:
    if not running_from_windows_temp():
        return
    message = (
        "BuyerReports.exe 目前看起來是從 Windows 暫存資料夾執行。\n\n"
        "常見原因是直接在 release.zip 壓縮檔裡雙擊執行檔。\n"
        "請先右鍵 release.zip 選擇「全部解壓縮」，再到解壓後的 "
        "BuyerReports 資料夾內執行 BuyerReports.exe。\n\n"
        "若仍被 Microsoft Defender SmartScreen 阻擋，請確認檔案來源可信後，"
        "點「更多資訊」再點「仍要執行」。"
    )
    show_windows_error("Buyer Reports 需要先解壓縮", message)
    raise SystemExit(message.replace("\n", " "))


def setup_run_log(out_dirs: Path | Sequence[Path]) -> tuple[Path, ...]:
    global _LOG_FILES
    close_run_log()
    if isinstance(out_dirs, Path):
        dirs = [out_dirs]
    else:
        dirs = list(out_dirs)

    log_paths = []
    seen = set()
    for out_dir in dirs:
        log_path = out_dir / "log"
        key = log_path.resolve()
        if key in seen:
            continue
        seen.add(key)
        out_dir.mkdir(parents=True, exist_ok=True)
        _LOG_FILES.append(log_path.open("w", encoding="utf-8"))
        log_paths.append(log_path)
    return tuple(log_paths)


def close_run_log() -> None:
    global _LOG_FILES
    for log_file in _LOG_FILES:
        log_file.close()
    _LOG_FILES = []


def write_traceback(detail: str) -> None:
    if _LOG_FILES:
        for log_file in _LOG_FILES:
            log_file.write("\n--- traceback ---\n")
            log_file.write(detail)
            log_file.flush()
    else:
        print(detail, file=sys.stderr)


def pause_for_windows_exe(enabled: bool) -> None:
    if not enabled:
        return
    try:
        input("\n按 Enter 關閉視窗...")
    except EOFError:
        pass


def numeric(value) -> float:
    """把儲存格內容轉成數字；文字 / 錯誤值 (#REF!, #N/A) 一律視為 0。"""
    if value is None:
        return 0.0
    if isinstance(value, bool):
        return 0.0
    if isinstance(value, (int, float)):
        return float(value)
    text = str(value).strip().replace(",", "")
    if not text:
        return 0.0
    try:
        return float(text)
    except ValueError:
        return 0.0


def clean_number(value: float):
    """整數就回傳 int，避免輸出 1234.0 這種樣子。"""
    if abs(value - round(value)) < 1e-9:
        return int(round(value))
    return value


def text_value(value) -> str:
    """把輸出值固定轉成 Excel 文字欄位使用的字串。"""
    if value is None:
        return ""
    if isinstance(value, bool):
        return ""
    if isinstance(value, (int, float)):
        return str(clean_number(float(value)))
    return str(value)


def write_text_cell(cell, value) -> None:
    cell.value = text_value(value)
    cell.number_format = "@"


def number_value(value):
    """把輸出數據固定轉成 Excel 數值欄位使用的值。"""
    if value is None:
        return None
    if isinstance(value, bool):
        return None
    if isinstance(value, (int, float)):
        return clean_number(float(value))
    text = str(value).strip().replace(",", "")
    if not text:
        return None
    try:
        return clean_number(float(text))
    except ValueError:
        return None


def write_number_cell(cell, value) -> None:
    cell.value = number_value(value)
    cell.number_format = "General"


def enforce_output_types(
    ws,
    max_row: int,
    max_col: int,
    text_rows: Iterable[int] = (),
    text_cols: Iterable[int] = (),
    number_cols: Iterable[int] = (),
) -> None:
    text_row_set = set(text_rows)
    text_col_set = set(text_cols)
    number_col_set = set(number_cols)
    for row in ws.iter_rows(min_row=1, max_row=max_row, max_col=max_col):
        for cell in row:
            if cell.row in text_row_set or cell.column in text_col_set:
                if cell.value is not None:
                    cell.value = text_value(cell.value)
                cell.number_format = "@"
            elif cell.column in number_col_set:
                if cell.value is not None:
                    cell.value = number_value(cell.value)
                cell.number_format = "General"


def enforce_text_range(ws, max_row: int, max_col: int) -> None:
    for row in ws.iter_rows(min_row=1, max_row=max_row, max_col=max_col):
        for cell in row:
            if cell.value is not None:
                cell.value = text_value(cell.value)
            cell.number_format = "@"


def date_text(date: dt.date) -> str:
    return f"{date.month}/{date.day}"


def serial_to_date(serial: float) -> dt.date:
    return EXCEL_EPOCH + dt.timedelta(days=int(serial))


def parse_date_arg(text: str) -> dt.date:
    for fmt in ("%Y-%m-%d", "%Y/%m/%d", "%Y%m%d"):
        try:
            return dt.datetime.strptime(text, fmt).date()
        except ValueError:
            continue
    raise argparse.ArgumentTypeError(f"無法解析日期：{text}（請用 YYYY-MM-DD）")


def autosize(ws, minimum: int = 8, maximum: int = 30) -> None:
    for col_idx, column_cells in enumerate(ws.columns, start=1):
        width = minimum
        for cell in column_cells:
            if cell.value is not None and not str(cell.value).startswith("="):
                width = max(width, min(maximum, len(str(cell.value)) + 2))
        ws.column_dimensions[get_column_letter(col_idx)].width = width


def first_existing_sheet(wb, candidates: Sequence[str]) -> str | None:
    return next((name for name in candidates if name in wb.sheetnames), None)


def first_sheet_matching_keywords(wb, keywords: Sequence[str]):
    for ws in wb.worksheets:
        if sheet_name_matches_keywords(ws.title, keywords):
            return ws
    names = "、".join(wb.sheetnames)
    raise SystemExit(f"找不到名稱包含 {keyword_label(keywords)} 的工作表；目前工作表：{names}")


def find_total_col(ws, header_row: int = 1) -> int | None:
    for cell in ws[header_row]:
        text = str(cell.value).strip().lower() if cell.value is not None else ""
        if text == "total" or text.startswith("total(") or text.startswith("total（"):
            return cell.column
    return None


def copy_cell_format(source, target) -> None:
    if source.has_style:
        target.font = copy(source.font)
        target.fill = copy(source.fill)
        target.border = copy(source.border)
        target.alignment = copy(source.alignment)
        target.protection = copy(source.protection)
        target.number_format = source.number_format


def copy_column_layout(source_ws, target_ws, source_col: int, target_col: int) -> None:
    source_letter = get_column_letter(source_col)
    target_letter = get_column_letter(target_col)
    source_dim = source_ws.column_dimensions[source_letter]
    target_dim = target_ws.column_dimensions[target_letter]
    if source_dim.width is not None:
        target_dim.width = source_dim.width
    target_dim.hidden = source_dim.hidden
    target_dim.outlineLevel = source_dim.outlineLevel
    target_dim.collapsed = source_dim.collapsed


def copy_row_layout(source_ws, target_ws, source_row: int, target_row: int) -> None:
    source_dim = source_ws.row_dimensions[source_row]
    target_dim = target_ws.row_dimensions[target_row]
    if source_dim.height is not None:
        target_dim.height = source_dim.height
    target_dim.hidden = source_dim.hidden
    target_dim.outlineLevel = source_dim.outlineLevel
    target_dim.collapsed = source_dim.collapsed


def set_filter_to_used_range(ws, last_col: int, last_row: int) -> None:
    ws.auto_filter.ref = f"A1:{get_column_letter(last_col)}{last_row}"

def first_sheet_containing(wb, keyword: str):
    return first_sheet_matching_keywords(wb, (keyword,))


def find_header_row(ws, required: Iterable[str]) -> int:
    required_set = set(required)
    for row in ws.iter_rows():
        values = {str(cell.value).strip() for cell in row if cell.value is not None}
        if required_set.issubset(values):
            return row[0].row
    raise SystemExit(f"在 {ws.title} 找不到含有下列欄位的表頭列：{', '.join(required)}")


def header_date(cell) -> dt.date | None:
    """判斷表頭儲存格是否為日期欄。

    只認 datetime/date 物件，或「數值 + 日期格式」的儲存格；避免把 `6月`
    這類月度小計欄或普通整數誤判成日期（舊版用 40000~50000 的數值範圍猜測，
    既會誤判也會在 2036 年後失效）。
    """
    value = cell.value
    if isinstance(value, dt.datetime):
        return value.date()
    if isinstance(value, dt.date):
        return value
    if isinstance(value, (int, float)) and not isinstance(value, bool):
        try:
            is_date = bool(cell.is_date)
        except (TypeError, ValueError):
            is_date = False
        if is_date and value > 0:
            return serial_to_date(value)
    return None
