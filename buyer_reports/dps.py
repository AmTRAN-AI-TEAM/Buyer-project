"""DPS report generation."""

from __future__ import annotations

import datetime as dt
from collections import defaultdict
from pathlib import Path
from typing import Sequence

from openpyxl import Workbook, load_workbook
from openpyxl.styles import Alignment, Font
from openpyxl.utils import get_column_letter

from .common import (
    autosize,
    clean_number,
    copy_cell_format,
    copy_column_layout,
    copy_row_layout,
    date_text,
    DEFAULT_DPS_SHEET_KEYWORDS,
    DEFAULT_DPS_PART_NUMBER_HEADERS,
    enforce_output_types,
    find_total_col,
    first_existing_sheet,
    first_sheet_matching_keywords,
    header_date,
    keyword_label,
    normalize_label,
    set_filter_to_used_range,
    warn,
    write_number_cell,
    write_text_cell,
)

# ---------------------------------------------------------------------------
# DPS
# ---------------------------------------------------------------------------

DPS_SOURCE_SHEET_KEYWORDS = DEFAULT_DPS_SHEET_KEYWORDS
DPS_PART_NUMBER_HEADERS = DEFAULT_DPS_PART_NUMBER_HEADERS
DPS_TIDY_SHEET = "DPS整理後"
DPS_COMPARE_SHEETS = (DPS_TIDY_SHEET, "DPS整理后")
DPS_BASE_HEADER_KEYS = ("Line", "W/O")
DPS_HEADER_KEYS = (*DPS_BASE_HEADER_KEYS, "AVTC P/N")



def detect_dps_tail_cutoff(dps_path: Path, max_date: dt.date) -> dt.date | None:
    """從活頁簿內既有的人工「DPS整理後」或舊版「DPS整理后」推斷末欄彙總桶的起始日。

    人工版是 Excel 樞紐日期群組的產物：最後一個日期欄其實是「> 前一日」的
    彙總桶。若該表不存在（未來只拿到原始檔）就回傳 None，代表所有日期都
    各自成欄。
    """
    try:
        wb = load_workbook(dps_path, data_only=True)
    except Exception as exc:  # noqa: BLE001 - 只是推斷失敗，不該中斷主流程
        warn(f"讀取既有整理後工作表失敗，改為不使用彙總桶：{exc}")
        return None
    try:
        sheet_name = next((name for name in DPS_COMPARE_SHEETS if name in wb.sheetnames), None)
        if sheet_name is None:
            return None
        ws = wb[sheet_name]
        dates = [d for d in (header_date(cell) for cell in ws[1]) if d is not None]
        if not dates:
            return None
        last = max(dates)
        return last if last < max_date else None
    finally:
        wb.close()

def excel_label_sort_key(label: str, is_numeric: bool):
    """模擬 Excel 樞紐的列標籤排序：數字在前（依數值），文字在後（依字串）。"""
    if is_numeric:
        try:
            return (0, float(label), "")
        except ValueError:
            pass
    return (1, 0.0, label)


def find_dps_header(ws, part_number_headers: Sequence[str] = DPS_PART_NUMBER_HEADERS) -> tuple[int, int, str]:
    required = {normalize_label(name) for name in DPS_BASE_HEADER_KEYS}
    part_aliases = [(normalize_label(name), name) for name in part_number_headers]

    for row in ws.iter_rows():
        headers = {
            normalize_label(cell.value): cell.column
            for cell in row
            if cell.value is not None and str(cell.value).strip()
        }
        if not required.issubset(headers):
            continue
        for normalized, label in part_aliases:
            if normalized in headers:
                return row[0].row, headers[normalized], label

    raise SystemExit(
        f"在 {ws.title} 找不到含有 Line、W/O，且料號欄名稱包含 "
        f"{keyword_label(part_number_headers)} 的表頭列。"
    )


def read_dps_data(
    dps_path: Path,
    include_star_parts: bool = False,
    sheet_keywords: Sequence[str] = DPS_SOURCE_SHEET_KEYWORDS,
    part_number_headers: Sequence[str] = DPS_PART_NUMBER_HEADERS,
) -> dict:
    wb = load_workbook(dps_path, data_only=True)
    try:
        ws = first_sheet_matching_keywords(wb, sheet_keywords)

        header_row, pn_col, pn_header = find_dps_header(ws, part_number_headers)
        date_columns: list[tuple[int, dt.date]] = []
        for cell in ws[header_row]:
            date = header_date(cell)
            if date is not None:
                date_columns.append((cell.column, date))
        if not date_columns:
            raise SystemExit(f"{ws.title} 內找不到任何日期欄")

        # 每個日期應該剛好有 D / N 兩班兩欄，不符就示警（資料格式可能變了）
        per_date = defaultdict(int)
        for _col, date in date_columns:
            per_date[date] += 1
        odd = {d: n for d, n in per_date.items() if n != 2}
        if odd:
            warn(
                "下列日期的欄數不是 2（D/N 兩班），請確認原始檔格式："
                + ", ".join(f"{d} x{n}" for d, n in sorted(odd.items()))
            )

        all_dates = sorted(per_date)
        date_col_set = {col for col, _date in date_columns}
        first_date_col = date_columns[0][0]

        aggregate: dict[str, dict[dt.date, float]] = defaultdict(lambda: defaultdict(float))
        label_is_numeric: dict[str, bool] = {}
        excluded: dict[str, float] = defaultdict(float)
        text_cells = 0
        orphan_cols: set[int] = set()

        for row in ws.iter_rows(min_row=header_row + 1, values_only=True):
            raw_pn = row[pn_col - 1] if pn_col - 1 < len(row) else None
            if raw_pn is None or not str(raw_pn).strip():
                continue
            pn = str(raw_pn).strip()
            is_numeric = isinstance(raw_pn, (int, float)) and not isinstance(raw_pn, bool)
            label_is_numeric.setdefault(pn, is_numeric)

            # 表頭沒被判定成日期、底下卻有數量的欄位 → 可能是漏判的日期欄
            for col_idx in range(first_date_col + 1, len(row) + 1):
                if col_idx in date_col_set:
                    continue
                value = row[col_idx - 1]
                if isinstance(value, (int, float)) and not isinstance(value, bool) and value:
                    orphan_cols.add(col_idx)

            row_total = 0.0
            for col_idx, date in date_columns:
                value = row[col_idx - 1] if col_idx - 1 < len(row) else None
                if value is None:
                    continue
                if not isinstance(value, (int, float)) or isinstance(value, bool):
                    if str(value).strip():
                        text_cells += 1
                    continue
                row_total += float(value)
                aggregate[pn][date] += float(value)

            if pn.endswith("*"):
                excluded[pn] += row_total

        if orphan_cols:
            warn(
                "下列欄位的表頭沒有被判定為日期，但底下有數量，未計入輸出："
                + ", ".join(get_column_letter(c) for c in sorted(orphan_cols))
                + "。常見原因：表頭日期是公式而檔案未經 Excel 存檔（無快取值）。"
            )

        if not include_star_parts:
            star_parts = [pn for pn in list(aggregate) + list(label_is_numeric) if pn.endswith("*")]
            for pn in set(star_parts):
                aggregate.pop(pn, None)
                label_is_numeric.pop(pn, None)
        else:
            excluded.clear()

        return {
            "source": dps_path,
            "source_sheet": ws.title,
            "header_row": header_row,
            "part_number_header": pn_header,
            "aggregate": aggregate,
            "label_is_numeric": label_is_numeric,
            "all_dates": all_dates,
            "date_columns": len(date_columns),
            "dates": len(all_dates),
            "date_range": (all_dates[0], all_dates[-1]),
            "excluded": dict(excluded),
            "text_cells": text_cells,
        }
    finally:
        wb.close()


def combine_dps_data(items: Sequence[dict]) -> dict:
    aggregate: dict[str, dict[dt.date, float]] = defaultdict(lambda: defaultdict(float))
    label_is_numeric: dict[str, bool] = {}
    all_dates: set[dt.date] = set()
    excluded: dict[str, float] = defaultdict(float)
    text_cells = 0
    date_columns = 0
    sources = []

    for item in items:
        sources.append(item)
        date_columns += item["date_columns"]
        text_cells += item["text_cells"]
        all_dates.update(item["all_dates"])
        for pn, is_numeric in item["label_is_numeric"].items():
            label_is_numeric.setdefault(pn, is_numeric)
        for pn, values in item["aggregate"].items():
            for date, qty in values.items():
                aggregate[pn][date] += qty
        for pn, qty in item["excluded"].items():
            excluded[pn] += qty

    if not all_dates:
        raise SystemExit("DPS 來源中找不到任何日期欄")

    sorted_dates = sorted(all_dates)
    return {
        "sources": sources,
        "aggregate": aggregate,
        "label_is_numeric": label_is_numeric,
        "all_dates": sorted_dates,
        "date_columns": date_columns,
        "dates": len(sorted_dates),
        "date_range": (sorted_dates[0], sorted_dates[-1]),
        "excluded": dict(excluded),
        "text_cells": text_cells,
    }


def build_dps_rows(
    aggregate: dict[str, dict[dt.date, float]],
    label_is_numeric: dict[str, bool],
    all_dates: Sequence[dt.date],
    tail_cutoff: dt.date | None = None,
) -> tuple[list[dt.date], list[tuple[str, bool, dict[dt.date, float]]]]:
    # 決定輸出欄：cutoff 之後的日期全部併入 cutoff 這一欄
    if tail_cutoff is not None:
        out_dates = [d for d in all_dates if d < tail_cutoff] + [tail_cutoff]
    else:
        out_dates = list(all_dates)

    def bucket(date: dt.date) -> dt.date:
        if tail_cutoff is not None and date >= tail_cutoff:
            return tail_cutoff
        return date

    rows: list[tuple[str, bool, dict[dt.date, float]]] = []
    for pn in sorted(label_is_numeric, key=lambda p: excel_label_sort_key(p, label_is_numeric[p])):
        values: dict[dt.date, float] = defaultdict(float)
        for date, qty in aggregate.get(pn, {}).items():
            values[bucket(date)] += qty
        rows.append((pn, label_is_numeric[pn], values))
    return out_dates, rows


def generate_dps(
    dps_path: Path,
    output_path: Path,
    include_star_parts: bool = False,
    tail_cutoff: dt.date | None = None,
    sheet_keywords: Sequence[str] = DPS_SOURCE_SHEET_KEYWORDS,
    part_number_headers: Sequence[str] = DPS_PART_NUMBER_HEADERS,
) -> dict:
    data = read_dps_data(
        dps_path,
        include_star_parts=include_star_parts,
        sheet_keywords=sheet_keywords,
        part_number_headers=part_number_headers,
    )
    out_dates, rows = build_dps_rows(
        data["aggregate"],
        data["label_is_numeric"],
        data["all_dates"],
        tail_cutoff,
    )
    write_dps_workbook(output_path, out_dates, rows, template_path=dps_path)

    grand_total = sum(sum(v.values()) for _pn, _n, v in rows)
    return {
        "source": dps_path,
        "source_sheet": data["source_sheet"],
        "header_row": data["header_row"],
        "part_number_header": data["part_number_header"],
        "date_columns": data["date_columns"],
        "dates": data["dates"],
        "date_range": data["date_range"],
        "tail_cutoff": tail_cutoff,
        "out_columns": len(out_dates),
        "rows": len(rows),
        "grand_total": clean_number(grand_total),
        "excluded": data["excluded"],
        "text_cells": data["text_cells"],
    }


def generate_merged_dps(
    dps_paths: Sequence[Path],
    output_path: Path,
    include_star_parts: bool = False,
    tail_cutoff: dt.date | None = None,
    sheet_keywords: Sequence[str] = DPS_SOURCE_SHEET_KEYWORDS,
    part_number_headers: Sequence[str] = DPS_PART_NUMBER_HEADERS,
) -> dict:
    items = []
    skipped = []
    for dps_path in dps_paths:
        try:
            items.append(
                read_dps_data(
                    dps_path,
                    include_star_parts=include_star_parts,
                    sheet_keywords=sheet_keywords,
                    part_number_headers=part_number_headers,
                )
            )
        except (SystemExit, Exception) as exc:  # noqa: BLE001 - 合併模式需略過壞檔
            skipped.append((dps_path, str(exc)))
            warn(f"DPS 來源檔 {dps_path.name} 無法併入，已略過。原因：{exc}")

    if not items:
        raise SystemExit("所有 DPS 來源檔都無法產出，未建立合併檔。")

    combined = combine_dps_data(items)
    out_dates, rows = build_dps_rows(
        combined["aggregate"],
        combined["label_is_numeric"],
        combined["all_dates"],
        tail_cutoff,
    )
    template_path = items[0]["source"]
    write_dps_workbook(output_path, out_dates, rows, template_path=template_path)

    grand_total = sum(sum(v.values()) for _pn, _n, v in rows)
    return {
        "sources": [item["source"] for item in items],
        "source_details": [
            {
                "source": item["source"],
                "source_sheet": item["source_sheet"],
                "header_row": item["header_row"],
                "part_number_header": item["part_number_header"],
                "date_columns": item["date_columns"],
                "dates": item["dates"],
                "date_range": item["date_range"],
            }
            for item in items
        ],
        "skipped": skipped,
        "tail_cutoff": tail_cutoff,
        "date_columns": combined["date_columns"],
        "dates": combined["dates"],
        "date_range": combined["date_range"],
        "out_columns": len(out_dates),
        "rows": len(rows),
        "grand_total": clean_number(grand_total),
        "excluded": combined["excluded"],
        "text_cells": combined["text_cells"],
    }


def write_dps_workbook(
    output_path: Path,
    dates: Sequence[dt.date],
    rows: Sequence[tuple[str, bool, dict[dt.date, float]]],
    template_path: Path | None = None,
) -> None:
    wb = Workbook()
    ws = wb.active
    ws.title = DPS_TIDY_SHEET

    write_text_cell(ws.cell(1, 1), "行标签")
    for offset, date in enumerate(dates):
        write_text_cell(ws.cell(1, 2 + offset), date_text(date))
    total_col = 2 + len(dates)
    write_text_cell(ws.cell(1, total_col), "total")

    for r_offset, (pn, is_numeric, values) in enumerate(rows):
        r = 2 + r_offset
        write_text_cell(ws.cell(r, 1), pn)
        total_qty = 0.0
        for offset, date in enumerate(dates):
            qty = values.get(date, 0.0)
            # 人工版的空白格是留空而非填 0，這裡保持一致
            if qty:
                write_number_cell(ws.cell(r, 2 + offset), clean_number(qty))
                total_qty += qty
        write_number_cell(ws.cell(r, total_col), clean_number(total_qty))

    for cell in ws[1]:
        cell.font = Font(bold=True)
        cell.alignment = Alignment(horizontal="center")
    ws.freeze_panes = "B2"
    autosize(ws)
    ws.column_dimensions["A"].width = 24.83
    last_row = len(rows) + 1
    apply_dps_layout(ws, template_path, total_col, last_row)
    enforce_output_types(
        ws,
        last_row,
        total_col,
        text_rows={1},
        text_cols={1},
        number_cols=range(2, total_col + 1),
    )

    output_path.parent.mkdir(parents=True, exist_ok=True)
    wb.save(output_path)


def apply_dps_layout(
    ws,
    template_path: Path | None,
    total_col: int,
    last_row: int,
) -> None:
    set_filter_to_used_range(ws, total_col, last_row)
    if template_path is None:
        return

    template_wb = None
    try:
        template_wb = load_workbook(template_path, data_only=False)
        sheet_name = first_existing_sheet(template_wb, DPS_COMPARE_SHEETS)
        if sheet_name is None:
            return
        template_ws = template_wb[sheet_name]
        template_total_col = find_total_col(template_ws) or min(total_col, template_ws.max_column)
        template_data_col = max(2, template_total_col - 1)

        def source_col(target_col: int) -> int:
            if target_col == total_col:
                return template_total_col
            if target_col <= template_data_col:
                return min(target_col, template_ws.max_column)
            return min(template_data_col, template_ws.max_column)

        copy_row_layout(template_ws, ws, 1, 1)
        for col in range(1, total_col + 1):
            src_col = source_col(col)
            copy_column_layout(template_ws, ws, src_col, col)
            copy_cell_format(template_ws.cell(1, src_col), ws.cell(1, col))

        if template_ws.max_row < 2:
            return
        for row_idx in range(2, last_row + 1):
            copy_row_layout(template_ws, ws, 2, row_idx)
            for col in range(1, total_col + 1):
                src_col = source_col(col)
                copy_cell_format(template_ws.cell(2, src_col), ws.cell(row_idx, col))
    except Exception as exc:  # noqa: BLE001 - 樣式套用失敗不應阻斷數值報表產生
        warn(f"DPS 樣式樣板套用失敗，已改用內建版面：{exc}")
    finally:
        if template_wb is not None:
            template_wb.close()
