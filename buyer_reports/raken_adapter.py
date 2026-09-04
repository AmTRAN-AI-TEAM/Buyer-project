"""RAKEN-specific CTB input adapters.

RAKEN does not provide the generic ``BOM1`` / ``open po`` workbook layout used
by the AVTC flow.  Its BOM relationship and usage come from the reference CTB
workbook, while PO and shortage data come from separate sheets/files.  This
module translates those sources into the common CTB data classes without
changing the AVTC readers.
"""

from __future__ import annotations

import datetime as dt
import posixpath
import re
from collections import defaultdict
from copy import copy
from dataclasses import dataclass
from io import BytesIO
from pathlib import Path
from typing import Any, Sequence
from zipfile import ZIP_DEFLATED, ZipFile
from xml.etree import ElementTree as ET

from openpyxl import Workbook, load_workbook
from openpyxl.formatting.rule import CellIsRule
from openpyxl.styles import Font
from openpyxl.utils import get_column_letter

from .common import (
    copy_cell_format,
    copy_column_layout,
    copy_row_layout,
    clean_number,
    normalize_label,
    normalize_part_number,
    numeric,
    unhide_workbook_columns,
)
from .ctb import (
    BomRow,
    CTB_SHEET,
    CtbPart,
    OpenPoRecord,
    Period,
    ShortageRecord,
    _cell_ref,
    _enable_formula_recalculation,
    _initial_sum_cols_for_cutoff,
    _sum_rows_in_col_expression,
    build_part_map,
    eta_schedule_for_records,
    filter_ctb_parts,
    read_dps_pp,
    read_over_shortage,
    workbook_has_sheet,
)


RAKEN_REFERENCE_NAME = "光学 CTB 20260701.xlsx"
RAKEN_DEMAND_SHEET = "demand"
RAKEN_REFERENCE_CTB_SHEET = "CTB"
RAKEN_PO_SHEET = "PO"
RAKEN_SHORTAGE_SHEET = "over shortage"

NON_MATERIAL_LABELS = (
    "不用lbr",
    "不用fiml",
    "客供料",
    "已用在单独36阶",
)


@dataclass(frozen=True)
class RakenGroup:
    source_row: int
    part: str
    model: str
    remark: str
    use_value: Any
    org: str
    control_pn: str
    moq: str


@dataclass(frozen=True)
class RakenDetail:
    source_row: int
    part: str
    use: float
    model: str
    remark: str
    org: str
    control_pn: str
    moq: str
    allocation: Any = None


@dataclass(frozen=True)
class RakenAllocation:
    source_row: int
    part: str
    allocation: Any


@dataclass(frozen=True)
class RakenChildSpec:
    part: str
    use: float
    source_row: int
    model: str
    remark: str
    org: str
    control_pn: str
    moq: str
    allocation: Any = None


def _part_key(value: Any) -> str:
    return normalize_part_number(value).casefold()


def _text(value: Any) -> str:
    return "" if value is None else str(value).strip()


def _is_non_material(value: Any) -> bool:
    text = _part_key(value)
    return not text or any(label in text for label in NON_MATERIAL_LABELS)


def _is_numeric_cell(value: Any) -> bool:
    if isinstance(value, bool) or value is None:
        return False
    if isinstance(value, (int, float)):
        return True
    text = _text(value).replace(",", "")
    if not text:
        return False
    try:
        float(text)
    except ValueError:
        return False
    return True


def _optional_numeric(value: Any) -> float | None:
    if value is None or isinstance(value, bool):
        return None
    if isinstance(value, (int, float)):
        return float(value)
    text = _text(value).replace(",", "")
    if not text:
        return None
    if text.endswith("%"):
        try:
            return float(text[:-1]) / 100.0
        except ValueError:
            return None
    try:
        return float(text)
    except ValueError:
        return None


def _raken_allocation_value(value: Any) -> float | None:
    number = _optional_numeric(value)
    if number is None:
        return None
    return number / 100.0 if abs(number) > 1 else number


def _raken_ctb_allocation_value(value: Any) -> Any:
    number = _raken_allocation_value(value)
    if number is not None:
        return clean_number(number)
    text = _text(value)
    return text or None


def _use_tokens(value: Any) -> list[float]:
    """Read numeric use values from ``1*1`` / ``*10*1*`` strings."""
    if _is_numeric_cell(value):
        return [numeric(value)]
    text = _text(value)
    if "*" not in text:
        return []
    values = []
    for token in text.split("*"):
        token = token.strip()
        if not token or not _is_numeric_cell(token):
            continue
        values.append(numeric(token))
    return values


def _expand_compound_part(value: Any) -> list[str]:
    """Expand CTB shorthand such as ``M01/02`` and ``FFCN4/N5``."""
    text = normalize_part_number(value)
    if not text:
        return []
    pieces = [piece.strip() for piece in text.split("/") if piece.strip()]
    if len(pieces) <= 1:
        return pieces

    first = pieces[0]
    result = [first]
    for suffix in pieces[1:]:
        if suffix.casefold().startswith(first.casefold()):
            result.append(suffix)
            continue

        # Numeric suffixes replace the complete numeric tail:
        # 098101044C44/45 -> 098101044C45 (not ...C445).
        if suffix[0].isdigit():
            prefix = re.sub(r"\d+$", "", first)
            if prefix != first:
                result.append(prefix + suffix)
                continue

        # CTB shorthand keeps the overlapping tail before the slash:
        # 098101084M01/02 -> 098101084M02
        # 0981010FFCN4/N5 -> 0981010FFCN5
        overlap = 0
        max_overlap = min(len(first), len(suffix))
        for size in range(max_overlap, 0, -1):
            if first[-size:].casefold() == suffix[:size].casefold():
                overlap = size
                break
        if overlap:
            result.append(first[:-overlap] + suffix)
            continue

        # For forms such as 0981010FFCN4/N5, the slash suffix starts with
        # the final alphabetic segment of the first item.
        leading_letters = re.match(r"[A-Za-z]+", suffix)
        if leading_letters:
            shared_text = leading_letters.group(0)
            shared_at = first.casefold().rfind(shared_text.casefold())
            if shared_at >= 0:
                result.append(first[:shared_at] + suffix)
                continue

        result.append(first + suffix)
    return result


def _find_header_row(rows: Sequence[Sequence[Any]], required: Sequence[str]) -> int | None:
    targets = [normalize_label(label) for label in required]
    for offset, values in enumerate(rows, start=1):
        normalized = {normalize_label(value) for value in values if value is not None}
        if all(target in normalized for target in targets):
            return offset
    return None


def _header_columns(values: Sequence[Any]) -> dict[str, int]:
    return {
        normalize_label(value): index
        for index, value in enumerate(values, start=1)
        if value is not None and normalize_label(value)
    }


def _find_column(headers: dict[str, int], aliases: Sequence[str]) -> int | None:
    normalized = [normalize_label(alias) for alias in aliases]
    for alias in normalized:
        if alias in headers:
            return headers[alias]
    for alias in normalized:
        for header, column in headers.items():
            if alias in header:
                return column
    return None


def _reference_candidates(input_dir: Path) -> list[Path]:
    candidates = []
    for path in input_dir.glob("*.xlsx"):
        if path.name.startswith("~$"):
            continue
        if all(
            workbook_has_sheet(path, sheet)
            for sheet in (RAKEN_DEMAND_SHEET, RAKEN_REFERENCE_CTB_SHEET, RAKEN_PO_SHEET)
        ):
            candidates.append(path)
    return sorted(candidates, key=lambda path: path.stat().st_mtime, reverse=True)


def find_raken_reference_workbook(input_dir: Path) -> Path:
    preferred = input_dir / RAKEN_REFERENCE_NAME
    if preferred.is_file() and all(
        workbook_has_sheet(preferred, sheet)
        for sheet in (RAKEN_DEMAND_SHEET, RAKEN_REFERENCE_CTB_SHEET, RAKEN_PO_SHEET)
    ):
        return preferred
    candidates = _reference_candidates(input_dir)
    if not candidates:
        raise SystemExit(
            f"找不到 RAKEN 光學 CTB 參考檔：{input_dir} 內需有 demand、CTB、PO 工作表"
        )
    return candidates[0]


def find_raken_shortage_workbook(input_dir: Path) -> Path:
    preferred = input_dir / "shortage.xlsx"
    if preferred.is_file() and workbook_has_sheet(preferred, RAKEN_SHORTAGE_SHEET):
        return preferred
    candidates = [
        path
        for path in input_dir.glob("*.xlsx")
        if not path.name.startswith("~$")
        and workbook_has_sheet(path, RAKEN_SHORTAGE_SHEET)
    ]
    candidates.sort(key=lambda path: path.stat().st_mtime, reverse=True)
    if not candidates:
        raise SystemExit(
            f"找不到 RAKEN shortage 來源：{input_dir} 內需有 {RAKEN_SHORTAGE_SHEET!r} 工作表"
        )
    return candidates[0]


def has_raken_ctb_input_candidates(input_dir: Path) -> bool:
    if not input_dir.is_dir():
        return False
    try:
        find_raken_reference_workbook(input_dir)
        find_raken_shortage_workbook(input_dir)
    except SystemExit:
        return False
    return True


def _read_raken_demand_mapping(
    reference_path: Path,
) -> tuple[dict[str, list[dict[str, str]]], dict[str, list[dict[str, str]]], dict[str, int]]:
    wb = load_workbook(reference_path, read_only=True, data_only=True)
    try:
        sheet_name = next(
            (name for name in wb.sheetnames if normalize_label(name) == normalize_label(RAKEN_DEMAND_SHEET)),
            None,
        )
        if sheet_name is None:
            raise SystemExit(f"{reference_path.name} 內找不到 {RAKEN_DEMAND_SHEET} 工作表")
        ws = wb[sheet_name]
        header_rows = list(ws.iter_rows(min_row=1, max_row=min(ws.max_row, 10), values_only=True))
        header_row = _find_header_row(header_rows, ("FG PN", "PART_NO"))
        if header_row is None:
            raise SystemExit(f"{reference_path.name} 的 demand 找不到 FG PN / PART_NO 表頭")
        headers = _header_columns(header_rows[header_row - 1])
        fg_col = _find_column(headers, ("FG PN",))
        part_col = _find_column(headers, ("PART_NO", "PART NO"))
        model_col = _find_column(headers, ("Model",))
        vendor_col = _find_column(headers, ("VENDOR", "Vendor"))
        if fg_col is None or part_col is None:
            raise SystemExit(f"{reference_path.name} 的 demand 缺少 FG PN 或 PART_NO 欄")

        by_part: dict[str, list[dict[str, str]]] = defaultdict(list)
        by_parent: dict[str, list[dict[str, str]]] = defaultdict(list)
        skipped = 0
        rows = 0
        max_col = max(fg_col, part_col, model_col or 0, vendor_col or 0)
        for row_idx, values in enumerate(
            ws.iter_rows(min_row=header_row + 1, max_col=max_col, values_only=True),
            start=header_row + 1,
        ):
            fg = normalize_part_number(values[fg_col - 1] if fg_col <= len(values) else None)
            part = normalize_part_number(values[part_col - 1] if part_col <= len(values) else None)
            if not fg or _is_non_material(part):
                if part and _is_non_material(part):
                    skipped += 1
                continue
            model = _text(values[model_col - 1] if model_col and model_col <= len(values) else "")
            vendor = _text(values[vendor_col - 1] if vendor_col and vendor_col <= len(values) else "")
            entry = {
                "parent": fg,
                "part": part,
                "model": model,
                "vendor": vendor,
                "source_row": str(row_idx),
            }
            part_key = _part_key(part)
            parent_key = _part_key(fg)
            if not any(
                old["parent"].casefold() == fg.casefold()
                and old["model"] == model
                and old["vendor"] == vendor
                for old in by_part[part_key]
            ):
                by_part[part_key].append(entry)
            if not any(
                old["part"].casefold() == part.casefold()
                and old["model"] == model
                and old["vendor"] == vendor
                for old in by_parent[parent_key]
            ):
                by_parent[parent_key].append(entry)
            rows += 1
        return by_part, by_parent, {
            "rows": rows,
            "skipped": skipped,
            "groups": len(by_part),
            "parents": len(by_parent),
        }
    finally:
        wb.close()


def _read_raken_ctb_rows(reference_path: Path) -> list[tuple[int, list[Any]]]:
    wb = load_workbook(reference_path, read_only=True, data_only=True)
    try:
        sheet_name = next(
            (name for name in wb.sheetnames if normalize_label(name) == normalize_label(RAKEN_REFERENCE_CTB_SHEET)),
            None,
        )
        if sheet_name is None:
            raise SystemExit(f"{reference_path.name} 內找不到 {RAKEN_REFERENCE_CTB_SHEET} 工作表")
        ws = wb[sheet_name]
        return [
            (row_idx, list(values))
            for row_idx, values in enumerate(
                ws.iter_rows(min_row=1, max_col=11, values_only=True),
                start=1,
            )
        ]
    finally:
        wb.close()


def _group_children(
    groups: Sequence[RakenGroup],
    details: Sequence[RakenDetail],
    allocations: Sequence[RakenAllocation],
    *,
    warnings: list[str],
) -> dict[int, list[RakenChildSpec]]:
    expected = {
        index: {_part_key(part): part for part in _expand_compound_part(group.part)}
        for index, group in enumerate(groups)
    }
    allocation_by_child: dict[int, dict[str, Any]] = defaultdict(dict)
    allocation_by_group: dict[int, Any] = {}
    for allocation in allocations:
        if allocation.allocation in (None, ""):
            continue
        allocation_key = _part_key(allocation.part)
        allocation_parts = {_part_key(part) for part in _expand_compound_part(allocation.part)}
        for index, group in enumerate(groups):
            if allocation_key == _part_key(group.part):
                allocation_by_group.setdefault(index, allocation.allocation)
                continue
            for key in allocation_parts:
                if key in expected[index]:
                    allocation_by_child[index].setdefault(key, allocation.allocation)

    result: dict[int, list[RakenChildSpec]] = defaultdict(list)
    for detail in details:
        matches = [index for index, values in expected.items() if _part_key(detail.part) in values]
        if not matches:
            warnings.append(
                f"CTB row {detail.source_row} 子件 {detail.part} 無法對應目前 demand 群組，已略過"
            )
            continue
        for index in matches:
            detail_key = _part_key(detail.part)
            display_part = expected[index][detail_key]
            result[index].append(
                RakenChildSpec(
                    part=display_part,
                    use=detail.use,
                    source_row=detail.source_row,
                    model=detail.model or groups[index].model,
                    remark=detail.remark or groups[index].remark,
                    org=detail.org or groups[index].org,
                    control_pn=detail.control_pn or groups[index].control_pn,
                    moq=detail.moq or groups[index].moq,
                    allocation=detail.allocation
                    if detail.allocation not in (None, "")
                    else allocation_by_child[index].get(detail_key, allocation_by_group.get(index)),
                )
            )

    for index, group in enumerate(groups):
        group_expected = expected[index]
        existing = {_part_key(item.part) for item in result.get(index, [])}
        missing = [part for key, part in group_expected.items() if key not in existing]
        if not missing:
            continue
        tokens = _use_tokens(group.use_value)
        if len(tokens) == 1 and len(group_expected) == 1:
            tokens = tokens * len(missing)
        if len(tokens) != len(group_expected):
            warnings.append(
                f"CTB row {group.source_row} 群組 {group.part} 的 F 用量與子件數量不一致，"
                f"用量={_text(group.use_value)!r}，子件={len(group_expected)}"
            )
            if len(tokens) < len(group_expected):
                tokens.extend([1.0] * (len(group_expected) - len(tokens)))
        token_by_key = {
            key: tokens[pos] if pos < len(tokens) else 1.0
            for pos, key in enumerate(group_expected)
        }
        for key in [_part_key(part) for part in missing]:
            result[index].append(
                RakenChildSpec(
                    part=group_expected[key],
                    use=token_by_key.get(key, 1.0),
                    source_row=group.source_row,
                    model=group.model,
                    remark=group.remark,
                    org=group.org,
                    control_pn=group.control_pn,
                    moq=group.moq,
                    allocation=allocation_by_child[index].get(key, allocation_by_group.get(index)),
                )
            )
    return result


def _append_summary_warning(
    warnings: list[str],
    prefix: str,
    values: Sequence[str],
    *,
    limit: int = 10,
    action: str = "已略過",
) -> None:
    if not values:
        return
    sample = "、".join(values[:limit])
    if len(values) > limit:
        sample += f" ...（另 {len(values) - limit} 筆）"
    warnings.append(f"{prefix}{len(values)} 筆，{action}：{sample}")


def _build_ctb_child_indexes(
    sections: Sequence[tuple[list[RakenGroup], list[RakenDetail], list[RakenAllocation]]],
    warnings: list[str],
) -> tuple[dict[str, list[RakenChildSpec]], dict[str, list[RakenChildSpec]]]:
    by_group: dict[str, list[RakenChildSpec]] = defaultdict(list)
    by_child: dict[str, list[RakenChildSpec]] = defaultdict(list)
    for section_groups, section_details, section_allocations in sections:
        children_by_group = _group_children(
            section_groups,
            section_details,
            section_allocations,
            warnings=warnings,
        )
        for group_index, group in enumerate(section_groups):
            children = children_by_group.get(group_index, [])
            if not children:
                warnings.append(f"CTB row {group.source_row} 群組 {group.part} 找不到可用子件")
                continue
            by_group[_part_key(group.part)].extend(children)
            for child in children:
                by_child[_part_key(child.part)].append(child)
    return by_group, by_child


def _lookup_ctb_child_specs(
    part: str,
    by_group: dict[str, list[RakenChildSpec]],
    by_child: dict[str, list[RakenChildSpec]],
) -> list[RakenChildSpec]:
    key = _part_key(part)
    return by_group.get(key) or by_child.get(key) or []


def read_raken_bom_rows(
    reference_path: Path,
    periods: Sequence[Period],
    demand_by_parent: dict[str, list[float]],
) -> tuple[list[BomRow], dict[str, Any]]:
    mapping_by_part, mapping_by_parent, demand_stats = _read_raken_demand_mapping(reference_path)
    raw_rows = _read_raken_ctb_rows(reference_path)
    header_row = None
    for row_idx, values in raw_rows[:20]:
        if len(values) >= 6 and normalize_label(values[1]) in {"乐轩料号", "乐轩料號"} and normalize_label(values[5]) == "用量":
            header_row = row_idx
            break
    if header_row is None:
        raise SystemExit(f"{reference_path.name} 的 CTB 找不到 乐轩料号 / 用量 表頭")

    sections: list[tuple[list[RakenGroup], list[RakenDetail], list[RakenAllocation]]] = []
    groups: list[RakenGroup] = []
    details: list[RakenDetail] = []
    allocations: list[RakenAllocation] = []
    for row_idx, values in raw_rows[header_row:]:
        values = values + [None] * max(0, 11 - len(values))
        part = normalize_part_number(values[1])
        model = _text(values[3])
        remark = _text(values[4])
        use_value = values[5]
        row_type = normalize_label(values[10]) if values[10] is not None else ""
        if row_type == "demand":
            groups.append(
                RakenGroup(
                    row_idx,
                    part,
                    model,
                    remark,
                    use_value,
                    _text(values[0]),
                    _text(values[2]),
                    _text(values[9]),
                )
            )
            continue
        if row_type == "balance":
            if groups:
                sections.append((groups, details, allocations))
            groups = []
            details = []
            allocations = []
            continue
        if groups and part and row_type == "" and _is_numeric_cell(use_value):
            details.append(
                RakenDetail(
                    row_idx,
                    part,
                    numeric(use_value),
                    model,
                    remark,
                    _text(values[0]),
                    _text(values[2]),
                    _text(values[9]),
                    _raken_ctb_allocation_value(values[8]),
                )
            )
            continue
        if groups and part and row_type:
            allocations.append(
                RakenAllocation(
                    row_idx,
                    part,
                    _raken_ctb_allocation_value(values[8]),
                )
            )
    if groups:
        sections.append((groups, details, allocations))

    warnings: list[str] = []
    ctb_by_group, ctb_by_child = _build_ctb_child_indexes(sections, warnings)
    rows: list[BomRow] = []
    active_parent_count = 0
    mapped_demand_links = 0
    missing_demand_parents: list[str] = []
    missing_ctb_parts: list[str] = []
    for parent, parent_demand in demand_by_parent.items():
        if not any(parent_demand):
            continue
        active_parent_count += 1
        mappings = mapping_by_parent.get(_part_key(parent), [])
        if not mappings:
            missing_demand_parents.append(parent)
            continue
        for item in mappings:
            child_specs = _lookup_ctb_child_specs(item["part"], ctb_by_group, ctb_by_child)
            if not child_specs:
                missing_ctb_parts.append(
                    f"demand row {item['source_row']} FG {parent} PART_NO {item['part']}"
                )
                continue
            mapped_demand_links += 1
            for spec in child_specs:
                rows.append(
                    BomRow(
                        source_row=spec.source_row,
                        category=item["model"] or spec.model,
                        parent=parent,
                        child=spec.part,
                        use=spec.use,
                        remark=spec.remark,
                        vendor=item["vendor"],
                        demand=[value * spec.use for value in parent_demand],
                        org=spec.org,
                        control_pn=spec.control_pn,
                        moq=spec.moq,
                        allocation=spec.allocation,
                    )
                )

    _append_summary_warning(
        warnings,
        "DPS+PP 有需求成品在 demand sheet 找不到 FG PN 對應，共 ",
        missing_demand_parents,
        action="已以空白計算列輸出",
    )
    _append_summary_warning(
        warnings,
        "demand PART_NO 找不到 CTB 用量/展開來源，共 ",
        missing_ctb_parts,
    )

    return rows, {
        "groups": sum(len(groups) for groups, _details, _allocations in sections),
        "mapped_groups": mapped_demand_links,
        "bom_rows": len(rows),
        "warnings": warnings,
        "demand_rows": demand_stats["rows"],
        "demand_skipped": demand_stats["skipped"],
        "demand_groups": len(mapping_by_part),
        "demand_parents": demand_stats["parents"],
        "active_dps_pp_parents": active_parent_count,
        "mapped_demand_links": mapped_demand_links,
        "ctb_group_keys": len(ctb_by_group),
        "ctb_child_keys": len(ctb_by_child),
        "missing_demand_parents": missing_demand_parents,
        "missing_ctb_parts": missing_ctb_parts,
    }


def read_raken_open_po(reference_path: Path) -> list[OpenPoRecord]:
    wb = load_workbook(reference_path, read_only=True, data_only=True)
    try:
        sheet_name = next(
            (name for name in wb.sheetnames if normalize_label(name) == normalize_label(RAKEN_PO_SHEET)),
            None,
        )
        if sheet_name is None:
            raise SystemExit(f"{reference_path.name} 內找不到 {RAKEN_PO_SHEET} 工作表")
        ws = wb[sheet_name]
        header_rows = list(ws.iter_rows(min_row=1, max_row=min(ws.max_row, 10), max_col=2, values_only=True))
        header_row = None
        for offset, values in enumerate(header_rows, start=1):
            if len(values) >= 2 and normalize_label(values[0]) == "行标签" and "quantity due" in normalize_label(values[1]):
                header_row = offset
                break
        if header_row is None:
            raise SystemExit(f"{reference_path.name} 的 PO 找不到 行标签 / Quantity Due 表頭")

        records: list[OpenPoRecord] = []
        summary_labels = {"total", "總計", "总计", "(空白)", "(blank)"}
        for row_idx, values in enumerate(
            ws.iter_rows(min_row=header_row + 1, max_col=2, values_only=True),
            start=header_row + 1,
        ):
            item = normalize_part_number(values[0] if values else None)
            if not item or item.casefold() in {label.casefold() for label in summary_labels}:
                continue
            quantity_due = numeric(values[1] if len(values) > 1 else None)
            if quantity_due == 0:
                continue
            records.append(
                OpenPoRecord(
                    source_row=row_idx,
                    key=item,
                    item=item,
                    supplier="",
                    supplier_site="",
                    quantity_due=quantity_due,
                    need_by_date=None,
                )
            )
        return records
    finally:
        wb.close()


def read_raken_erp_price(reference_path: Path) -> tuple[dict[str, dict[str, float]], str | None]:
    """Read optional Price values for the template columns."""
    wb = load_workbook(reference_path, read_only=True, data_only=True, keep_links=False)
    try:
        sheet_name = next(
            (name for name in wb.sheetnames if normalize_label(name) == normalize_label("ERP Price")),
            None,
        )
        if sheet_name is None:
            return {}, "參考檔內找不到 ERP Price，G Price 欄留白"
        ws = wb[sheet_name]
        header_rows = list(
            ws.iter_rows(
                min_row=1,
                max_row=min(ws.max_row, 10),
                max_col=min(ws.max_column, 17),
                values_only=True,
            )
        )
        header_row = _find_header_row(header_rows, ("Part No.", "Price"))
        if header_row is None:
            return {}, "ERP Price 找不到 Part No. / Price 表頭，G Price 欄留白"
        headers = _header_columns(header_rows[header_row - 1])
        part_col = _find_column(headers, ("Part No.", "Part No"))
        price_col = _find_column(headers, ("Price",))
        if part_col is None or price_col is None:
            return {}, "ERP Price 缺少 Part No. 或 Price，G Price 欄留白"

        result: dict[str, dict[str, float]] = {}
        max_col = max(part_col, price_col)
        for values in ws.iter_rows(
            min_row=header_row + 1,
            max_col=max_col,
            values_only=True,
        ):
            part = normalize_part_number(values[part_col - 1] if part_col <= len(values) else None)
            if not part or _is_non_material(part):
                continue
            key = _part_key(part)
            # ERP Price may contain multiple vendor/site rows.  Match the
            # workbook's first-match VLOOKUP behavior instead of summing them.
            if key in result:
                continue
            result[key] = {
                "price": _optional_numeric(values[price_col - 1] if price_col <= len(values) else None),
            }
        return result, None
    finally:
        wb.close()


def _merge_shortage(existing: ShortageRecord, incoming: ShortageRecord) -> None:
    if not existing.description:
        existing.description = incoming.description
    if not existing.buyer:
        existing.buyer = incoming.buyer
    if not existing.planner:
        existing.planner = incoming.planner
    if not existing.lead_time:
        existing.lead_time = incoming.lead_time
    existing.po_remain += incoming.po_remain
    existing.over_shortage += incoming.over_shortage
    existing.hld += incoming.hld
    existing.bor_mm += incoming.bor_mm
    existing.overshortage1 += incoming.overshortage1


def _align_raken_part_names(
    bom_rows: Sequence[BomRow],
    open_po: Sequence[OpenPoRecord],
    shortage: dict[str, ShortageRecord],
) -> dict[str, ShortageRecord]:
    """Join RAKEN sources case-insensitively while retaining BOM spelling."""
    display_by_key: dict[str, str] = {}

    def register(value: str) -> str:
        key = _part_key(value)
        if key and key not in display_by_key:
            display_by_key[key] = value
        return display_by_key.get(key, value)

    for row in bom_rows:
        row.child = register(row.child)
    for record in open_po:
        record.item = register(record.item)
        record.key = record.item
    normalized_shortage: dict[str, ShortageRecord] = {}
    for record in shortage.values():
        record.part = register(record.part)
        existing = normalized_shortage.get(record.part)
        if existing is None:
            normalized_shortage[record.part] = record
        else:
            _merge_shortage(existing, record)
    return normalized_shortage


def _raken_template_ctb_stream(template_path: Path) -> BytesIO:
    """Build an in-memory CTB-only workbook for reading layout metadata.

    The optical workbook contains several large source sheets, especially
    ``ERP Price``.  Loading the complete workbook in normal mode just to copy
    CTB styles is unnecessarily expensive and can exhaust memory.  Keep the
    original input untouched, but expose only its CTB sheet plus shared style
    resources to openpyxl.
    """
    main_ns = "http://schemas.openxmlformats.org/spreadsheetml/2006/main"
    rel_ns = "http://schemas.openxmlformats.org/officeDocument/2006/relationships"
    package_rel_ns = "http://schemas.openxmlformats.org/package/2006/relationships"
    qname = lambda namespace, name: f"{{{namespace}}}{name}"

    with ZipFile(template_path, "r") as source_zip:
        workbook_root = ET.fromstring(source_zip.read("xl/workbook.xml"))
        source_workbook_rels = ET.fromstring(source_zip.read("xl/_rels/workbook.xml.rels"))
        sheets = workbook_root.find(qname(main_ns, "sheets"))
        if sheets is None:
            raise SystemExit(f"{template_path.name} 找不到 workbook sheets 設定")
        ctb_sheet = None
        for sheet in list(sheets):
            if normalize_label(sheet.attrib.get("name")) == normalize_label(RAKEN_REFERENCE_CTB_SHEET):
                ctb_sheet = sheet
                break
        if ctb_sheet is None:
            raise SystemExit(f"{template_path.name} 找不到 CTB 工作表")
        ctb_rel_id = ctb_sheet.attrib.get(qname(rel_ns, "id"))
        ctb_target = next(
            (
                rel.attrib.get("Target")
                for rel in source_workbook_rels
                if rel.attrib.get("Id") == ctb_rel_id
            ),
            None,
        )
        if not ctb_target:
            raise SystemExit(f"{template_path.name} 找不到 CTB 工作表的檔案關聯")
        ctb_sheet_path = (
            ctb_target.lstrip("/")
            if ctb_target.startswith("/")
            else posixpath.normpath(posixpath.join("xl", ctb_target))
        )
        for sheet in list(sheets):
            if sheet is not ctb_sheet:
                sheets.remove(sheet)
        ctb_sheet.set(qname(rel_ns, "id"), "rId1")
        for element_name in ("definedNames", "customWorkbookViews"):
            element = workbook_root.find(qname(main_ns, element_name))
            if element is not None:
                workbook_root.remove(element)

        workbook_rels = ET.Element(qname(package_rel_ns, "Relationships"))
        relationships = (
            ("rId1", "http://schemas.openxmlformats.org/officeDocument/2006/relationships/worksheet", "worksheets/sheet1.xml"),
            ("rId2", "http://schemas.openxmlformats.org/officeDocument/2006/relationships/styles", "styles.xml"),
            ("rId3", "http://schemas.openxmlformats.org/officeDocument/2006/relationships/sharedStrings", "sharedStrings.xml"),
            ("rId4", "http://schemas.openxmlformats.org/officeDocument/2006/relationships/theme", "theme/theme1.xml"),
        )
        for rel_id, rel_type, target in relationships:
            ET.SubElement(
                workbook_rels,
                qname(package_rel_ns, "Relationship"),
                {"Id": rel_id, "Type": rel_type, "Target": target},
            )

        sheet_root = ET.fromstring(source_zip.read(ctb_sheet_path))
        for parent in list(sheet_root.iter()):
            for child in list(parent):
                if child.tag.rsplit("}", 1)[-1] in {"drawing", "legacyDrawing", "extLst"}:
                    parent.remove(child)

        content_types_root = ET.fromstring(source_zip.read("[Content_Types].xml"))
        for override in list(content_types_root):
            part_name = override.attrib.get("PartName", "")
            if part_name.startswith("/xl/worksheets/"):
                if part_name == f"/{ctb_sheet_path}":
                    override.set("PartName", "/xl/worksheets/sheet1.xml")
                else:
                    content_types_root.remove(override)

        stream = BytesIO()
        with ZipFile(stream, "w", ZIP_DEFLATED) as target_zip:
            target_zip.writestr("[Content_Types].xml", ET.tostring(content_types_root, encoding="utf-8", xml_declaration=True))
            target_zip.writestr("_rels/.rels", source_zip.read("_rels/.rels"))
            for name in source_zip.namelist():
                if name.startswith("docProps/"):
                    target_zip.writestr(name, source_zip.read(name))
            target_zip.writestr("xl/workbook.xml", ET.tostring(workbook_root, encoding="utf-8", xml_declaration=True))
            target_zip.writestr("xl/_rels/workbook.xml.rels", ET.tostring(workbook_rels, encoding="utf-8", xml_declaration=True))
            target_zip.writestr("xl/styles.xml", source_zip.read("xl/styles.xml"))
            target_zip.writestr("xl/sharedStrings.xml", source_zip.read("xl/sharedStrings.xml"))
            target_zip.writestr("xl/theme/theme1.xml", source_zip.read("xl/theme/theme1.xml"))
            target_zip.writestr("xl/worksheets/sheet1.xml", ET.tostring(sheet_root, encoding="utf-8", xml_declaration=True))
        stream.seek(0)
        return stream


def _template_layout_last_col(template_ws) -> int:
    last_col = 11
    for row_idx in (1, 2, 3):
        for col_idx in range(1, template_ws.max_column + 1):
            if template_ws.cell(row_idx, col_idx).value not in (None, ""):
                last_col = max(last_col, col_idx)
    return last_col


def _copy_raken_layout(template_ws, target_ws, last_col: int) -> None:
    for col_idx in range(1, last_col + 1):
        copy_column_layout(template_ws, target_ws, col_idx, col_idx)
    for row_idx in (1, 2, 3):
        copy_row_layout(template_ws, target_ws, row_idx, row_idx)
        for col_idx in range(1, last_col + 1):
            _copy_raken_cell_style(template_ws.cell(row_idx, col_idx), target_ws.cell(row_idx, col_idx))

    target_ws.sheet_view.showGridLines = template_ws.sheet_view.showGridLines
    target_ws.sheet_view.zoomScale = template_ws.sheet_view.zoomScale
    target_ws.freeze_panes = None
    target_ws.sheet_format = copy(template_ws.sheet_format)
    target_ws.sheet_properties = copy(template_ws.sheet_properties)
    target_ws.page_margins = copy(template_ws.page_margins)
    target_ws.page_setup = copy(template_ws.page_setup)
    target_ws.print_options = copy(template_ws.print_options)


def _copy_raken_data_row(template_ws, target_ws, source_row: int, target_row: int, last_col: int) -> None:
    copy_row_layout(template_ws, target_ws, source_row, target_row)
    for col_idx in range(1, last_col + 1):
        _copy_raken_cell_style(template_ws.cell(source_row, col_idx), target_ws.cell(target_row, col_idx))
        target_ws.cell(target_row, col_idx).value = None


def _copy_raken_cell_style(source_cell, target_cell) -> None:
    """Copy a style array after the source style tables are shared."""
    target_cell._style = copy(source_cell._style) if source_cell.has_style else None


def _copy_raken_style_tables(source_wb, target_wb) -> None:
    """Make source style IDs valid in the newly created output workbook."""
    for attribute in (
        "_fonts",
        "_fills",
        "_borders",
        "_alignments",
        "_protections",
        "_number_formats",
        "_cell_styles",
        "_named_styles",
    ):
        source_values = getattr(source_wb, attribute)
        setattr(target_wb, attribute, type(source_values)(list(source_values)))


def _raken_template_style_rows(template_ws) -> tuple[int, int, int]:
    demand_row = None
    detail_row = None
    balance_row = None
    for row_idx in range(1, min(template_ws.max_row, 2000) + 1):
        row_type = normalize_label(template_ws.cell(row_idx, 11).value)
        if demand_row is None and row_type == "demand":
            demand_row = row_idx
        if balance_row is None and row_type == "balance":
            balance_row = row_idx
        if (
            detail_row is None
            and demand_row is not None
            and row_idx > demand_row
            and template_ws.cell(row_idx, 2).value not in (None, "")
            and template_ws.cell(row_idx, 11).value in (None, "")
        ):
            detail_row = row_idx
        if demand_row and detail_row and balance_row:
            break
    if demand_row is None or balance_row is None:
        raise SystemExit("光學 CTB 的 CTB sheet 找不到 demand / Balance 列樣式")
    if detail_row is None:
        detail_row = demand_row
    return demand_row, detail_row, balance_row


def _raken_model(part: CtbPart) -> str:
    values = []
    seen = set()
    for row in part.bom_rows:
        model = row.category.strip()
        if not model or model.casefold() == "#n/a" or model in seen:
            continue
        seen.add(model)
        values.append(model)
    if values:
        return "/".join(values)
    return part.model.strip()


def _raken_bom_text(part: CtbPart, attribute: str) -> str:
    values = []
    for row in part.bom_rows:
        value = str(getattr(row, attribute, "") or "").strip()
        if value and value not in values:
            values.append(value)
    return "/".join(values)


def _raken_description(part: CtbPart) -> str:
    for row in part.bom_rows:
        if row.remark.strip():
            return row.remark.strip()
    return part.shortage.description.strip() if part.shortage else ""


def _raken_use(part: CtbPart) -> float | str | None:
    values = []
    for row in part.bom_rows:
        if not any(abs(row.use - old) < 1e-9 for old in values):
            values.append(row.use)
    if not values:
        return None
    if len(values) == 1:
        return clean_number(values[0])
    return "*".join(str(clean_number(value)) for value in values)


def _raken_allocation_label(value: Any) -> str:
    if isinstance(value, bool):
        return ""
    if isinstance(value, (int, float)):
        return f"{clean_number(float(value) * 100)}%"
    return _text(value)


def _raken_allocation(part: CtbPart) -> Any:
    values: list[Any] = []
    seen: set[str] = set()
    for row in part.bom_rows:
        value = getattr(row, "allocation", None)
        if value in (None, ""):
            continue
        key = f"number:{float(value):.10g}" if isinstance(value, (int, float)) else f"text:{value}"
        if key in seen:
            continue
        seen.add(key)
        values.append(value)
    if not values:
        return None
    if len(values) == 1:
        return values[0]
    return "/".join(label for label in (_raken_allocation_label(value) for value in values) if label)


def _raken_sum_formula(row_indices: Sequence[int], col_idx: int) -> str:
    return _sum_rows_in_col_expression(row_indices, col_idx)


def _raken_first_balance_formula(
    balance_row: int,
    demand_row: int,
    period_columns: Sequence[tuple[int, Period]],
    over_shortage: float,
    initial_sum_cols: tuple[int, int] | None,
) -> str:
    if initial_sum_cols is None:
        selected = list(period_columns[1:])
    else:
        selected = [
            (col, period)
            for col, period in period_columns
            if initial_sum_cols[0] <= col <= initial_sum_cols[1]
        ]
    if selected:
        demand_ref = f"{_cell_ref(demand_row, selected[0][0])}:{_cell_ref(demand_row, selected[-1][0])}"
        demand_expr = f"SUM({demand_ref})"
    else:
        demand_expr = "0"
    return f"={clean_number(over_shortage)}+{demand_expr}"


def _raken_write_balance_formulas(
    ws,
    balance_row: int,
    demand_row: int,
    eta_rows: Sequence[int],
    period_columns: Sequence[tuple[int, Period]],
    over_shortage: float,
    initial_sum_cols: tuple[int, int] | None,
) -> None:
    for index, (col_idx, _period) in enumerate(period_columns):
        if index == 0:
            formula = _raken_first_balance_formula(
                balance_row,
                demand_row,
                period_columns,
                over_shortage,
                initial_sum_cols,
            )
        else:
            previous_col = period_columns[index - 1][0]
            eta_expr = _raken_sum_formula(eta_rows, previous_col)
            formula = (
                f"={_cell_ref(balance_row, previous_col)}+{eta_expr}"
                f"-{_cell_ref(demand_row, col_idx)}"
            )
        ws.cell(balance_row, col_idx).value = formula


def write_raken_ctb_sheet(
    wb: Workbook,
    template_path: Path,
    periods: Sequence[Period],
    parts: Sequence[CtbPart],
    price_by_part: dict[str, dict[str, float]],
    *,
    dps_cutoff_end: dt.date | None = None,
    default_eta_lead_days: int,
    eta_lead_days_by_supplier_site: dict[str, int] | None = None,
) -> dict[str, int | str]:
    template_stream = _raken_template_ctb_stream(template_path)
    template_wb = load_workbook(template_stream, data_only=False, keep_links=False)
    try:
        template_name = next(
            (name for name in template_wb.sheetnames if normalize_label(name) == normalize_label(CTB_SHEET)),
            None,
        )
        if template_name is None:
            raise SystemExit(f"{template_path.name} 內找不到 CTB 工作表")
        template_ws = template_wb[template_name]
        _copy_raken_style_tables(template_wb, wb)
        target_ws = wb.active
        target_ws.title = CTB_SHEET
        layout_last_col = _template_layout_last_col(template_ws)
        _copy_raken_layout(template_ws, target_ws, layout_last_col)
        demand_style_row, detail_style_row, balance_style_row = _raken_template_style_rows(template_ws)

        # Clear all template values, then restore only column labels.  No
        # original formulas or material values are carried into the output.
        for row_idx in range(1, 4):
            for col_idx in range(1, layout_last_col + 1):
                target_ws.cell(row_idx, col_idx).value = None
        for col_idx in range(1, 11):
            target_ws.cell(3, col_idx).value = template_ws.cell(3, col_idx).value

        first_period_col = 12
        last_period_col = first_period_col + len(periods) - 1
        if last_period_col > layout_last_col:
            for col_idx in range(layout_last_col + 1, last_period_col + 1):
                copy_column_layout(template_ws, target_ws, layout_last_col, col_idx)
                for row_idx in (1, 2, 3):
                    copy_cell_format(
                        template_ws.cell(row_idx, layout_last_col),
                        target_ws.cell(row_idx, col_idx),
                    )
            layout_last_col = last_period_col
        for col_idx in range(first_period_col, layout_last_col + 1):
            target_ws.cell(1, col_idx).value = None
            target_ws.cell(2, col_idx).value = None
            target_ws.cell(3, col_idx).value = None
        for index, period in enumerate(periods):
            col_idx = first_period_col + index
            if period.start is not None:
                target_ws.cell(1, col_idx).value = period.header1 or period.start.strftime("%b").upper()
                target_ws.cell(2, col_idx).value = period.header2 or f"WK{period.start.isocalendar().week:02d}"
                target_ws.cell(3, col_idx).value = period.start
            else:
                target_ws.cell(1, col_idx).value = period.header1
                target_ws.cell(2, col_idx).value = period.header2
                target_ws.cell(3, col_idx).value = period.header4 or period.label

        period_columns = [
            (first_period_col + index, period)
            for index, period in enumerate(periods)
        ]
        initial_sum_cols = _initial_sum_cols_for_cutoff(period_columns, dps_cutoff_end)
        row_idx = 4
        demand_rows = detail_rows = balance_rows = placeholder_parts = 0
        balance_row_indices: list[int] = []
        for part in parts:
            calculation_blank = _raken_calculation_blank(part)
            if calculation_blank:
                placeholder_parts += 1
            demand_row = row_idx
            _copy_raken_data_row(template_ws, target_ws, demand_style_row, demand_row, layout_last_col)
            _write_raken_row_values(
                target_ws,
                demand_row,
                part,
                row_type="demand",
                use_value=_raken_use(part),
                price_by_part=price_by_part,
                calculation_blank=calculation_blank,
                last_period_col=last_period_col,
            )
            if not calculation_blank:
                for index, value in enumerate(part.demand):
                    target_ws.cell(demand_row, first_period_col + index).value = clean_number(value)
            demand_rows += 1
            row_idx += 1

            schedules = (
                {}
                if calculation_blank
                else eta_schedule_for_records(
                    periods,
                    part.open_po,
                    demand=part.demand,
                    over_shortage=part.shortage.over_shortage if part.shortage else 0.0,
                    default_lead_days=default_eta_lead_days,
                    lead_days_by_supplier_site=eta_lead_days_by_supplier_site,
                    period_start_col=first_period_col,
                )
            )
            po_by_key = {record.key: record for record in part.open_po}
            eta_rows: list[int] = []
            eta_items = list(schedules.items()) or [("", [0.0] * len(periods))]
            for key, schedule in eta_items:
                eta_row = row_idx
                eta_rows.append(eta_row)
                _copy_raken_data_row(template_ws, target_ws, detail_style_row, eta_row, layout_last_col)
                record = po_by_key.get(key)
                _write_raken_row_values(
                    target_ws,
                    eta_row,
                    part,
                    row_type="ETA",
                    use_value=_raken_use(part),
                    price_by_part=price_by_part,
                    open_po=None if calculation_blank else record.quantity_due if record else None,
                    calculation_blank=calculation_blank,
                    last_period_col=last_period_col,
                )
                if not calculation_blank:
                    for index, value in enumerate(schedule):
                        target_ws.cell(eta_row, first_period_col + index).value = clean_number(value)
                detail_rows += 1
                row_idx += 1

            balance_row = row_idx
            _copy_raken_data_row(template_ws, target_ws, balance_style_row, balance_row, layout_last_col)
            _write_raken_row_values(
                target_ws,
                balance_row,
                part,
                row_type="Balance",
                use_value=_raken_use(part),
                price_by_part=price_by_part,
                calculation_blank=calculation_blank,
                last_period_col=last_period_col,
            )
            if not calculation_blank:
                _raken_write_balance_formulas(
                    target_ws,
                    balance_row,
                    demand_row,
                    eta_rows,
                    period_columns,
                    part.shortage.over_shortage if part.shortage else 0.0,
                    initial_sum_cols,
                )
                balance_row_indices.append(balance_row)
            balance_rows += 1
            row_idx += 1

        target_ws.auto_filter.ref = f"A3:{get_column_letter(layout_last_col)}{row_idx - 1}"
        if balance_row_indices and periods:
            for balance_row in balance_row_indices:
                target_ws.conditional_formatting.add(
                    f"{get_column_letter(first_period_col)}{balance_row}:"
                    f"{get_column_letter(last_period_col)}{balance_row}",
                    CellIsRule(
                        operator="lessThan",
                        formula=["0"],
                        font=Font(color="FFFF0000"),
                    ),
                )
        return {
            "mode": "raken-template",
            "parts": len(parts),
            "demand_rows": demand_rows,
            "eta_rows": detail_rows,
            "other_rows": 0,
            "balance_rows": balance_rows,
            "placeholder_parts": placeholder_parts,
            "periods": len(periods),
            "template_rows": template_ws.max_row,
            "output_rows": row_idx - 1,
            "template_source": template_path,
        }
    finally:
        template_wb.close()
        template_stream.close()


def _write_raken_row_values(
    ws,
    row_idx: int,
    part: CtbPart,
    *,
    row_type: str,
    use_value: float | str | None,
    price_by_part: dict[str, dict[str, float]],
    open_po: float | None = None,
    calculation_blank: bool = False,
    last_period_col: int,
) -> None:
    values: dict[int, Any] = {
        1: None,
        2: part.part,
        3: None,
        4: _raken_model(part),
        5: None,
        6: None if calculation_blank else use_value,
        7: None if calculation_blank else price_by_part.get(_part_key(part.part), {}).get("price"),
        8: None if calculation_blank else open_po,
        9: None,
        10: None,
        11: row_type,
    }
    for col_idx in range(1, last_period_col + 1):
        if col_idx not in values:
            ws.cell(row_idx, col_idx).value = None
    for col_idx, value in values.items():
        ws.cell(row_idx, col_idx).value = value if value not in (None, "") else None


def _raken_bom_part_keys(bom_rows: Sequence[BomRow]) -> set[str]:
    return {_part_key(row.child) for row in bom_rows if _part_key(row.child)}


def _filter_raken_open_po(
    open_po: Sequence[OpenPoRecord],
    allowed_part_keys: set[str],
) -> list[OpenPoRecord]:
    return [record for record in open_po if _part_key(record.item) in allowed_part_keys]


def _filter_raken_shortage(
    shortage: dict[str, ShortageRecord],
    allowed_part_keys: set[str],
) -> dict[str, ShortageRecord]:
    return {
        part_no: record
        for part_no, record in shortage.items()
        if _part_key(record.part or part_no) in allowed_part_keys
    }


def _raken_placeholder_parts(
    periods: Sequence[Period],
    missing_parents: Sequence[str],
    existing_part_keys: set[str],
) -> list[CtbPart]:
    placeholders: list[CtbPart] = []
    seen = set(existing_part_keys)
    for parent in missing_parents:
        parent = normalize_part_number(parent)
        key = _part_key(parent)
        if not key or key in seen:
            continue
        seen.add(key)
        placeholders.append(
            CtbPart(
                part=parent,
                demand=[0.0] * len(periods),
                calculation_blank=True,
                placeholder_reason="missing demand FG PN",
            )
        )
    return placeholders


def _raken_calculation_blank(part: CtbPart) -> bool:
    return bool(getattr(part, "calculation_blank", False))


def _raken_ctb_source_order(part: CtbPart) -> tuple[int, str]:
    source_rows = [
        row.source_row
        for row in part.bom_rows
        if isinstance(row.source_row, int) and row.source_row > 0
    ]
    source_row = min(source_rows) if source_rows else 10**9
    return source_row, _part_key(part.part)


def _sort_raken_calculated_parts(parts: Sequence[CtbPart]) -> list[CtbPart]:
    return sorted(parts, key=_raken_ctb_source_order)


def generate_raken_ctb(
    dps_pp_path: Path,
    reference_path: Path,
    shortage_path: Path,
    output_path: Path,
    *,
    dps_cutoff_end: dt.date | None = None,
    default_eta_lead_days: int,
    eta_lead_days_by_supplier_site: dict[str, int] | None = None,
) -> dict[str, Any]:
    periods, demand_by_parent = read_dps_pp(dps_pp_path)
    bom_rows, bom_info = read_raken_bom_rows(reference_path, periods, demand_by_parent)
    bom_part_keys = _raken_bom_part_keys(bom_rows)
    shortage_source = read_over_shortage(shortage_path)
    open_po_source = read_raken_open_po(reference_path)
    shortage = _filter_raken_shortage(shortage_source, bom_part_keys)
    open_po = _filter_raken_open_po(open_po_source, bom_part_keys)
    filter_warnings: list[str] = []
    if len(shortage_source) != len(shortage):
        filter_warnings.append(
            "shortage 有 "
            f"{len(shortage_source) - len(shortage)} 筆不屬於 DPS+PP→demand 展開料號，已略過"
        )
    if len(open_po_source) != len(open_po):
        filter_warnings.append(
            "Open PO 有 "
            f"{len(open_po_source) - len(open_po)} 筆不屬於 DPS+PP→demand 展開料號，已略過"
        )
    shortage = _align_raken_part_names(bom_rows, open_po, shortage)
    parts_by_part, part_order = build_part_map(periods, bom_rows, shortage, open_po)
    parts = _sort_raken_calculated_parts(filter_ctb_parts(parts_by_part, part_order))
    placeholder_parts = _raken_placeholder_parts(
        periods,
        bom_info.get("missing_demand_parents", []),
        {_part_key(part.part) for part in parts},
    )
    parts.extend(placeholder_parts)
    price_by_part, price_warning = read_raken_erp_price(reference_path)

    wb = Workbook()
    # RAKEN output intentionally contains only CTB.  Rebuild the values using
    # the optical CTB's layout and styles without copying its source content.
    stats = write_raken_ctb_sheet(
        wb,
        reference_path,
        periods,
        parts,
        price_by_part,
        dps_cutoff_end=dps_cutoff_end,
        default_eta_lead_days=default_eta_lead_days,
        eta_lead_days_by_supplier_site=eta_lead_days_by_supplier_site,
    )
    if wb.sheetnames != [CTB_SHEET]:
        raise SystemExit("RAKEN CTB 輸出應只包含 CTB 工作表")
    _enable_formula_recalculation(wb)
    output_path.parent.mkdir(parents=True, exist_ok=True)
    unhide_workbook_columns(wb)
    wb.save(output_path)
    return {
        **stats,
        "mode": "raken-template",
        "dps_pp_source": dps_pp_path,
        "reference_source": reference_path,
        "shortage_source": shortage_path,
        "bom_rows": len(bom_rows),
        "over_shortage_rows": len(shortage),
        "over_shortage_source_rows": len(shortage_source),
        "open_po_rows": len(open_po),
        "open_po_source_rows": len(open_po_source),
        "placeholder_parts": len(placeholder_parts),
        "template_source": reference_path,
        "warnings": bom_info["warnings"] + filter_warnings + ([price_warning] if price_warning else []),
        "raken_bom": bom_info,
        "output": output_path,
    }
