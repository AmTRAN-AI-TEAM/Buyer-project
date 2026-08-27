"""DPS + PP integrated report generation."""

from __future__ import annotations

import datetime as dt
import re
from collections import defaultdict
from pathlib import Path
from typing import Sequence

from openpyxl import Workbook, load_workbook
from openpyxl.styles import Alignment, Font
from openpyxl.utils import get_column_letter

from .common import (
    autosize,
    clean_number,
    DEFAULT_DPS_PART_NUMBER_HEADERS,
    DEFAULT_DPS_SHEET_KEYWORDS,
    DEFAULT_PP_PART_NUMBER_FIELD_KEYWORDS,
    DEFAULT_PP_SHEET_KEYWORDS,
    enforce_output_types,
    numeric,
    set_filter_to_used_range,
    warn,
    write_number_cell,
    write_text_cell,
    VALID_DPS_PP_LATE_DPS_MODES,
)
from .dps import (
    combine_dps_data,
    excel_label_sort_key,
    read_dps_data,
)
from .pp import (
    build_pp_periods,
    MONTH_FCST_LABEL_RE,
    MONTH_INDEX,
    MONTH_PLAIN_LABEL_RE,
    normalize_field,
    parse_pivot_cache,
    read_layout,
    select_pp_pivot_source,
)

DPS_PP_TIDY_SHEET = "DPS+PP"
DPS_PP_OUTPUT_NAME = "DPS+PP.xlsx"
DAY_ABBR = ("Mon", "Tue", "Wed", "Thu", "Fri", "Sat", "Sun")


def buyer_week_for_date(date: dt.date) -> tuple[int, int]:
    """Return buyer week year/week, where each week runs Saturday-Friday."""
    iso = (date + dt.timedelta(days=2)).isocalendar()
    return iso.year, iso.week


def dps_pp_base_date_for_run_date(date: dt.date) -> dt.date:
    """Return the date used to auto-detect the DPS+PP current week."""
    if date.weekday() == 4:  # Friday: prepare the report as the next buyer week.
        return date + dt.timedelta(days=1)
    return date


def buyer_week_range(year: int, week: int) -> tuple[dt.date, dt.date]:
    """Return Saturday-Friday date range for a buyer week."""
    start = dt.date.fromisocalendar(year, week, 1) - dt.timedelta(days=2)
    return start, start + dt.timedelta(days=6)


def add_buyer_weeks(year: int, week: int, weeks: int) -> tuple[int, int]:
    start, _end = buyer_week_range(year, week)
    return buyer_week_for_date(start + dt.timedelta(days=weeks * 7))


def week_label_for_date(date: dt.date) -> str:
    _year, week = buyer_week_for_date(date)
    return f"WK{week:02d}"


def month_label(date: dt.date) -> str:
    return date.strftime("%b").upper()


def week_range_label(start: dt.date, end: dt.date) -> str:
    return f"{start.month}/{start.day}-{end.month}/{end.day}"


def parse_pp_week_label(label: str) -> int | None:
    match = re.fullmatch(r"WK\s*(\d{1,2})", str(label).strip(), re.IGNORECASE)
    if not match:
        return None
    return int(match.group(1))


def pp_period_header(label: str, base_year: int) -> dict | None:
    week = parse_pp_week_label(label)
    if week is not None:
        start, end = buyer_week_range(base_year, week)
        return {
            "label": f"WK{week:02d}",
            "start": start,
            "row1": month_label(start),
            "row2": f"WK{week:02d}",
            "row3": week_range_label(start, end),
            "row4": start.isoformat(),
        }

    text = str(label).strip()
    match = MONTH_FCST_LABEL_RE.match(text) or MONTH_PLAIN_LABEL_RE.match(text)
    if not match:
        return None
    month = match.group(1).title()
    year = 2000 + int(match.group(2))
    start = dt.date(year, MONTH_INDEX[month.lower()], 1)
    return {
        "label": text,
        "start": start,
        "row1": month_label(start),
        "row2": text,
        "row3": "",
        "row4": start.isoformat(),
    }


def read_dps_sources(
    dps_paths: Sequence[Path],
    dps_mode: str,
    include_star_parts: bool,
    sheet_keywords: Sequence[str],
    part_number_headers: Sequence[str],
    drop_zero_total_rows: bool = False,
) -> tuple[dict, list[tuple[Path, str]]]:
    items = []
    skipped = []
    for dps_path in dps_paths:
        try:
            item = read_dps_data(
                dps_path,
                include_star_parts=include_star_parts,
                sheet_keywords=sheet_keywords,
                part_number_headers=part_number_headers,
                drop_zero_total_rows=drop_zero_total_rows,
            )
            items.append(item)
            if dps_mode != "merge_all":
                break
        except (SystemExit, Exception) as exc:  # noqa: BLE001 - 單一壞檔不阻斷候選嘗試
            skipped.append((dps_path, str(exc)))
            warn(f"DPS+PP 的 DPS 來源檔 {dps_path.name} 無法併入，已略過。原因：{exc}")

    if not items:
        raise SystemExit("DPS+PP 找不到任何格式正確的 DPS 來源。")
    return combine_dps_data(items), skipped


def read_pp_plan_data(
    pp_path: Path,
    plan: str,
    start_week: int | None,
    base_year: str | None,
    report_date: dt.date | None,
    sheet_keywords: Sequence[str],
    part_number_keywords: Sequence[str],
) -> dict:
    pivot_source = select_pp_pivot_source(
        pp_path,
        sheet_keywords=sheet_keywords,
        part_number_keywords=part_number_keywords,
    )
    cache = parse_pivot_cache(
        pp_path,
        definition_part=pivot_source["cache_definition"],
        part_number_keywords=part_number_keywords,
    )
    fields = cache["fields"]
    records = cache["records"]

    field_index = {normalize_field(name): idx for idx, name in enumerate(fields)}
    part_number_field = cache["part_number_field"]
    required = ["Customer", "Model", "Plan"]
    missing = [name for name in required if name not in field_index]
    if missing:
        raise SystemExit(f"樞紐快取缺少必要欄位：{', '.join(missing)}")
    if part_number_field not in field_index:
        raise SystemExit(f"樞紐快取找不到選定的料號欄位：{part_number_field}")

    if report_date is None:
        report_date = cache["refreshed_date"] or dt.date.today()
    if base_year is None:
        base_year = f"{report_date.year % 100:02d}"
    if start_week is None:
        start_week = max(1, report_date.isocalendar()[1] - 1)

    layout = read_layout(pp_path, sheet_name=pivot_source["sheet_name"])
    periods = build_pp_periods(fields, layout.labels, base_year, start_week)
    if not periods:
        raise SystemExit("推導不出任何 PP 期間欄位")

    customer_idx = field_index["Customer"]
    model_idx = field_index["Model"]
    pn_idx = field_index[part_number_field]
    plan_idx = field_index["Plan"]

    plans_seen = {r[plan_idx].strip() for r in records if len(r) > plan_idx}
    if plan not in plans_seen:
        raise SystemExit(
            f"快取中沒有 Plan = {plan!r}；可選值：{', '.join(sorted(plans_seen))}"
        )

    aggregate: dict[str, dict[str, float]] = defaultdict(lambda: defaultdict(float))
    metadata: dict[str, tuple[str, str]] = {}
    source_order: list[str] = []
    plan_rows = 0
    for record in records:
        if len(record) <= max(customer_idx, model_idx, pn_idx, plan_idx):
            continue
        if record[plan_idx].strip() != plan:
            continue
        pn = record[pn_idx].strip()
        if not pn:
            continue
        plan_rows += 1
        if pn not in metadata:
            metadata[pn] = (record[customer_idx].strip(), record[model_idx].strip())
            source_order.append(pn)
        for label, source_indexes in periods:
            aggregate[pn][label] += sum(
                numeric(record[i]) for i in source_indexes if i < len(record)
            )

    return {
        "source": pp_path,
        "layout_sheet": pivot_source["sheet_name"],
        "cache_part": cache["definition_part"],
        "part_number_field": part_number_field,
        "refreshed_date": cache["refreshed_date"],
        "refreshed_by": cache["refreshed_by"],
        "records": len(records),
        "plan": plan,
        "plan_rows": plan_rows,
        "base_year": base_year,
        "start_week": start_week,
        "periods": [label for label, _source_indexes in periods],
        "aggregate": aggregate,
        "metadata": metadata,
        "source_order": source_order,
    }


def build_dps_buckets(
    data: dict,
    cutoff_end: dt.date,
    late_dps_mode: str = "merge_to_cutoff",
) -> tuple[list[dt.date], dict, float]:
    values: dict[str, dict[dt.date, float]] = defaultdict(lambda: defaultdict(float))
    out_dates = {date for date in data["all_dates"] if date <= cutoff_end}
    late_total = 0.0

    for pn, by_date in data["aggregate"].items():
        for date, qty in by_date.items():
            if date > cutoff_end:
                late_total += qty
                if late_dps_mode == "drop":
                    continue
                bucket = cutoff_end
            else:
                bucket = date
            values[pn][bucket] += qty
            out_dates.add(bucket)

    return sorted(out_dates), values, late_total


def trim_dps_bucket_dates(
    dates: Sequence[dt.date],
    values: dict[str, dict[dt.date, float]],
) -> tuple[list[dt.date], list[dt.date]]:
    last_nonzero_index = None
    for index, date in enumerate(dates):
        if any(by_date.get(date, 0.0) for by_date in values.values()):
            last_nonzero_index = index

    if last_nonzero_index is None:
        return list(dates), []
    kept = list(dates[:last_nonzero_index + 1])
    trimmed = list(dates[last_nonzero_index + 1:])
    return kept, trimmed


def select_pp_periods(pp_data: dict, cutoff_end: dt.date) -> list[dict]:
    base_year = 2000 + int(pp_data["base_year"])
    selected = []
    for label in pp_data["periods"]:
        header = pp_period_header(label, base_year)
        if header is None:
            continue
        if header["start"] > cutoff_end:
            selected.append(header)
    return selected


def read_bom_map(paths: Sequence[Path]) -> tuple[dict[str, str], bool]:
    bom = {}
    found = False
    for path in paths:
        try:
            wb = load_workbook(path, data_only=True)
        except Exception:  # noqa: BLE001 - BOM 是輔助欄位，不阻斷主報表
            continue
        try:
            if "BOM1" not in wb.sheetnames:
                continue
            found = True
            ws = wb["BOM1"]
            for row in ws.iter_rows(min_row=1):
                key = row[1].value if len(row) > 1 else None
                value = row[7].value if len(row) > 7 else None
                if key is None or not str(key).strip():
                    continue
                bom.setdefault(str(key).strip(), "" if value is None else str(value).strip())
        finally:
            wb.close()
    return bom, found


def build_output_rows(
    dps_data: dict,
    dps_values: dict,
    dps_dates: Sequence[dt.date],
    pp_data: dict,
    pp_periods: Sequence[dict],
    bom: dict[str, str],
) -> list[list]:
    seen = set()
    ordered_parts = []

    def has_values(part: str) -> bool:
        return any(dps_values[part].get(date, 0) for date in dps_dates) or any(
            pp_data["aggregate"][part].get(period["label"], 0) for period in pp_periods
        )

    for pn in pp_data["source_order"]:
        if pn in seen or not has_values(pn):
            continue
        seen.add(pn)
        ordered_parts.append(pn)

    dps_only = [
        pn for pn in dps_data["label_is_numeric"]
        if pn not in seen and has_values(pn)
    ]
    dps_only.sort(
        key=lambda pn: excel_label_sort_key(pn, dps_data["label_is_numeric"].get(pn, False))
    )
    ordered_parts.extend(dps_only)

    rows = []
    for pn in ordered_parts:
        period_values = [
            clean_number(dps_values[pn].get(date, 0))
            for date in dps_dates
        ]
        period_values.extend(
            clean_number(pp_data["aggregate"][pn].get(period["label"], 0))
            for period in pp_periods
        )
        total = clean_number(sum(numeric(value) for value in period_values))
        rows.append([pn, *period_values, total, bom.get(pn, "")])
    return rows


def write_dps_pp_workbook(
    output_path: Path,
    dps_dates: Sequence[dt.date],
    pp_periods: Sequence[dict],
    rows: Sequence[list],
) -> None:
    wb = Workbook()
    ws = wb.active
    ws.title = DPS_PP_TIDY_SHEET

    for row in range(1, 5):
        write_text_cell(ws.cell(row, 1), "成品料号")

    col = 2
    for date in dps_dates:
        write_text_cell(ws.cell(1, col), month_label(date))
        write_text_cell(ws.cell(2, col), week_label_for_date(date))
        write_text_cell(ws.cell(3, col), DAY_ABBR[date.weekday()])
        write_text_cell(ws.cell(4, col), date.isoformat())
        col += 1

    for period in pp_periods:
        write_text_cell(ws.cell(1, col), period["row1"])
        write_text_cell(ws.cell(2, col), period["row2"])
        write_text_cell(ws.cell(3, col), period["row3"])
        write_text_cell(ws.cell(4, col), period["row4"])
        col += 1

    total_col = col
    bom_col = col + 1
    write_text_cell(ws.cell(4, total_col), "total")
    write_text_cell(ws.cell(4, bom_col), "BOM")

    for offset, row_values in enumerate(rows, start=5):
        for col_idx, value in enumerate(row_values, start=1):
            if col_idx == 1 or col_idx == bom_col:
                write_text_cell(ws.cell(offset, col_idx), value)
            else:
                write_number_cell(ws.cell(offset, col_idx), value)

    for row in ws.iter_rows(min_row=1, max_row=4, max_col=bom_col):
        for cell in row:
            if cell.value is not None:
                cell.font = Font(bold=True)
                cell.alignment = Alignment(horizontal="center")

    ws.freeze_panes = "B5"
    set_filter_to_used_range(ws, bom_col, len(rows) + 4)
    autosize(ws, maximum=18)
    ws.column_dimensions["A"].width = 28.0
    ws.column_dimensions[get_column_letter(total_col)].width = 14.0
    ws.column_dimensions[get_column_letter(bom_col)].width = 20.0
    enforce_output_types(
        ws,
        len(rows) + 4,
        bom_col,
        text_rows=range(1, 5),
        text_cols={1, bom_col},
        number_cols=range(2, total_col + 1),
    )

    output_path.parent.mkdir(parents=True, exist_ok=True)
    wb.save(output_path)


def generate_dps_pp(
    dps_paths: Sequence[Path],
    pp_path: Path,
    output_path: Path,
    dps_mode: str,
    dps_weeks_ahead: int,
    current_week: int | None = None,
    current_date: dt.date | None = None,
    include_star_parts: bool = False,
    pp_plan: str = "Production Input",
    pp_start_week: int | None = None,
    pp_base_year: str | None = None,
    pp_report_date: dt.date | None = None,
    dps_sheet_keywords: Sequence[str] = DEFAULT_DPS_SHEET_KEYWORDS,
    dps_part_number_headers: Sequence[str] = DEFAULT_DPS_PART_NUMBER_HEADERS,
    pp_sheet_keywords: Sequence[str] = DEFAULT_PP_SHEET_KEYWORDS,
    pp_part_number_keywords: Sequence[str] = DEFAULT_PP_PART_NUMBER_FIELD_KEYWORDS,
    drop_zero_total_rows: bool = False,
    trim_trailing_zero_date_columns: bool = False,
    late_dps_mode: str = "merge_to_cutoff",
) -> dict:
    if late_dps_mode not in VALID_DPS_PP_LATE_DPS_MODES:
        raise SystemExit(
            f"DPS+PP 截止日後 DPS 處理模式 {late_dps_mode!r} 不支援；"
            f"可用模式：{', '.join(VALID_DPS_PP_LATE_DPS_MODES)}"
        )
    if current_date is None:
        current_date = dt.date.today()
    current_week_auto = current_week is None
    current_week_base_date = dps_pp_base_date_for_run_date(current_date)
    current_year, auto_current_week = buyer_week_for_date(current_week_base_date)
    if current_week is None:
        current_week = auto_current_week
    else:
        current_year, _auto_current_week = buyer_week_for_date(current_date)
    current_week_range = buyer_week_range(current_year, current_week)
    cutoff_year, cutoff_week = add_buyer_weeks(
        current_year,
        current_week,
        dps_weeks_ahead - 1,
    )
    cutoff_start, cutoff_end = buyer_week_range(cutoff_year, cutoff_week)
    pp_start_year, pp_start_week_number = buyer_week_for_date(cutoff_end + dt.timedelta(days=1))

    dps_data, skipped_dps = read_dps_sources(
        dps_paths,
        dps_mode=dps_mode,
        include_star_parts=include_star_parts,
        sheet_keywords=dps_sheet_keywords,
        part_number_headers=dps_part_number_headers,
        drop_zero_total_rows=drop_zero_total_rows,
    )
    pp_data = read_pp_plan_data(
        pp_path,
        plan=pp_plan,
        start_week=pp_start_week,
        base_year=pp_base_year,
        report_date=pp_report_date,
        sheet_keywords=pp_sheet_keywords,
        part_number_keywords=pp_part_number_keywords,
    )

    dps_dates, dps_values, dps_late_total = build_dps_buckets(
        dps_data,
        cutoff_end,
        late_dps_mode=late_dps_mode,
    )
    dps_trimmed_trailing_zero_dates = []
    if trim_trailing_zero_date_columns:
        dps_dates, dps_trimmed_trailing_zero_dates = trim_dps_bucket_dates(
            dps_dates,
            dps_values,
        )
    pp_periods = select_pp_periods(pp_data, cutoff_end)
    if not pp_periods:
        raise SystemExit(
            f"PP 中找不到 {cutoff_end} 之後的期間欄位，無法產出 DPS+PP。"
        )

    source_paths = [item["source"] for item in dps_data["sources"]] + [pp_path]
    bom, bom_found = read_bom_map(source_paths)
    rows = build_output_rows(dps_data, dps_values, dps_dates, pp_data, pp_periods, bom)
    write_dps_pp_workbook(output_path, dps_dates, pp_periods, rows)

    grand_total = sum(sum(numeric(value) for value in row[1:-2]) for row in rows)
    return {
        "dps_sources": [item["source"] for item in dps_data["sources"]],
        "dps_source_details": [
            {
                "source": item["source"],
                "source_sheet": item["source_sheet"],
                "part_number_header": item["part_number_header"],
                "date_range": item["date_range"],
                "dropped_zero_rows": item["dropped_zero_rows"],
            }
            for item in dps_data["sources"]
        ],
        "skipped_dps": skipped_dps,
        "dps_dropped_zero_rows": dps_data["dropped_zero_rows"],
        "pp_source": pp_path,
        "pp_sheet": pp_data["layout_sheet"],
        "pp_cache": pp_data["cache_part"],
        "pp_part_number_field": pp_data["part_number_field"],
        "current_date": current_date,
        "current_week_base_date": current_week_base_date,
        "current_week_auto": current_week_auto,
        "current_week": current_week,
        "current_week_year": current_year,
        "current_week_range": current_week_range,
        "dps_weeks_ahead": dps_weeks_ahead,
        "dps_cutoff_week": cutoff_week,
        "dps_cutoff_year": cutoff_year,
        "dps_cutoff_range": (cutoff_start, cutoff_end),
        "pp_start_week": pp_start_week_number,
        "pp_start_year": pp_start_year,
        "dps_dates": len(dps_dates),
        "dps_trimmed_trailing_zero_dates": dps_trimmed_trailing_zero_dates,
        "pp_periods": [period["label"] for period in pp_periods],
        "dps_late_mode": late_dps_mode,
        "dps_late_total": clean_number(dps_late_total),
        "rows": len(rows),
        "grand_total": clean_number(grand_total),
        "bom_found": bom_found,
        "bom_count": len(bom),
    }
