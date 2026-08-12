"""Workbook comparison helpers for --compare."""

from __future__ import annotations

from pathlib import Path
from typing import Sequence

from openpyxl import load_workbook

from .common import clean_number, header_date, log, numeric

# ---------------------------------------------------------------------------
# 對帳（--compare）
# ---------------------------------------------------------------------------


def _load_grid(
    path: Path,
    sheet: str | Sequence[str],
    key_col: int,
    first_data_col: int,
    label_row: int = 1,
):
    wb = load_workbook(path, data_only=True)
    try:
        sheet_names = [sheet] if isinstance(sheet, str) else list(sheet)
        sheet_name = next((name for name in sheet_names if name in wb.sheetnames), None)
        if sheet_name is None:
            return None
        ws = wb[sheet_name]
        labels = {}
        for cell in ws[label_row]:
            if cell.column < first_data_col or cell.value is None:
                continue
            text = str(cell.value).strip()
            if text.lower() == "total":
                continue
            date = header_date(cell)
            labels[cell.column] = date if date is not None else text
        grid = {}
        for row in ws.iter_rows(min_row=label_row + 1):
            key = row[key_col - 1].value
            if key is None or not str(key).strip():
                continue
            grid[str(key).strip()] = {
                labels[cell.column]: numeric(cell.value)
                for cell in row
                if cell.column in labels
            }
        return grid
    finally:
        wb.close()


def compare(
    title: str,
    manual_path: Path,
    manual_sheet: str | Sequence[str],
    generated_path: Path,
    generated_sheet: str | Sequence[str],
    key_col: int,
    first_data_col_manual: int,
    first_data_col_generated: int,
    limit: int = 30,
) -> None:
    log(f"\n=== 對帳：{title} ===")
    manual = _load_grid(manual_path, manual_sheet, key_col, first_data_col_manual)
    if manual is None:
        manual_sheet_label = manual_sheet if isinstance(manual_sheet, str) else " / ".join(manual_sheet)
        log(f"  來源檔內沒有 {manual_sheet_label} 工作表，略過對帳。")
        return
    generated = _load_grid(generated_path, generated_sheet, key_col, first_data_col_generated)
    if generated is None:
        log("  產出檔讀取失敗，略過對帳。")
        return

    only_manual = sorted(set(manual) - set(generated))
    only_generated = sorted(set(generated) - set(manual))
    log(f"  人工版 {len(manual)} 列 / 本次產出 {len(generated)} 列")
    if only_generated:
        log(f"  ▲ 人工版漏記（原始有、人工無）共 {len(only_generated)} 筆：")
        for key in only_generated[:limit]:
            values = {k: v for k, v in generated[key].items() if v}
            log(f"      + {key}  {values}")
    if only_manual:
        log(f"  ▼ 人工版多出（原始無）共 {len(only_manual)} 筆：")
        for key in only_manual[:limit]:
            log(f"      - {key}")

    diffs = []
    for key in sorted(set(manual) & set(generated)):
        for label in sorted(set(manual[key]) | set(generated[key]), key=str):
            a = manual[key].get(label, 0.0)
            b = generated[key].get(label, 0.0)
            if abs(a - b) > 1e-6:
                diffs.append((key, label, a, b))
    log(f"  數值差異：{len(diffs)} 格")
    for key, label, a, b in diffs[:limit]:
        log(f"      ! {key} | {label} | 人工={clean_number(a)} 原始={clean_number(b)}")
    if len(diffs) > limit:
        log(f"      ...（其餘 {len(diffs) - limit} 筆省略）")
