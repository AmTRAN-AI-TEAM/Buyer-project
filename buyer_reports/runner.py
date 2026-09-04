"""Command-line orchestration for Buyer Reports."""

from __future__ import annotations

import argparse
import datetime as dt
import os
import re
import subprocess
import sys
import traceback
from dataclasses import dataclass
from pathlib import Path
from typing import Sequence

from openpyxl import load_workbook

from .common import (
    CTB_ETA_CONFIG_FILE_NAME,
    DEFAULT_CTB_ETA_LEAD_DAYS,
    Progress,
    clean_number,
    close_run_log,
    first_sheet_matching_keywords,
    header_date,
    is_frozen_app,
    keyword_label,
    load_ctb_eta_config,
    load_sheet_detection_config,
    log,
    parse_date_arg,
    pause_for_windows_exe,
    prevent_temp_execution,
    project_root,
    save_ctb_eta_config,
    set_log_quiet,
    setup_run_log,
    sync_ctb_eta_config,
    warn,
    write_traceback,
)
from .compare import compare
from .ctb import (
    CTB_OUTPUT_NAME,
    CTB_TEMPLATE_SHEET_NAMES,
    find_optional_workbook_with_sheets,
    find_workbook_with_sheet,
    generate_ctb,
    has_ctb_input_candidates,
    read_open_po,
)
from .dps import (
    DPS_COMPARE_SHEETS,
    DPS_PART_NUMBER_HEADERS,
    DPS_SOURCE_SHEET_KEYWORDS,
    DPS_TIDY_SHEET,
    detect_dps_tail_cutoff,
    find_dps_header,
    generate_dps,
    generate_merged_dps,
)
from .dps_pp import DPS_PP_OUTPUT_NAME, generate_dps_pp
from .pp import PP_COMPARE_SHEETS, PP_TIDY_SHEET, generate_pp
from .raken_adapter import (
    find_raken_reference_workbook,
    find_raken_shortage_workbook,
    generate_raken_ctb,
    has_raken_ctb_input_candidates,
)


@dataclass(frozen=True)
class RunContext:
    name: str
    input_dir: Path
    out_dir: Path
    dps_mode: str
    pp_mode: str
    dps_pp_dps_weeks_ahead: int
    dps_pp_dps_weeks_source: str
    dps_pp_late_dps_mode: str
    dps_drop_zero_total_rows: bool
    dps_trim_trailing_zero_dates: bool
    legacy: bool = False

    @property
    def label(self) -> str:
        return self.name if self.name else "預設"


def late_dps_mode_label(mode: str) -> str:
    if mode == "drop":
        return "截止日後排除，由 PP 接續"
    if mode == "merge_to_cutoff":
        return "截止日後併入 DPS 最後一天"
    return mode


def find_input_candidates(
    input_dir: Path,
    patterns: Sequence[str],
    kind: str,
    multiple_action: str = "將依修改時間由新到舊嘗試。",
) -> list[Path]:
    """在 input/ 底下依關鍵字找輸入檔；多個候選時依修改時間由新到舊回傳。"""
    if not input_dir.is_dir():
        raise SystemExit(f"找不到輸入資料夾：{input_dir}")
    candidates = [
        path
        for path in sorted(input_dir.glob("*.xlsx"))
        if not path.name.startswith("~$")
        and any(re.search(pattern, path.name, re.IGNORECASE) for pattern in patterns)
    ]
    if not candidates:
        raise SystemExit(
            f"在 {input_dir} 找不到 {kind} 檔（檔名需含 {' 或 '.join(patterns)}），"
            f"或請用命令列參數明確指定。"
        )
    if len(candidates) > 1:
        candidates.sort(key=lambda p: p.stat().st_mtime, reverse=True)
        warn(f"{kind} 有多個候選檔，{multiple_action}")
    return candidates


def find_input(input_dir: Path, patterns: Sequence[str], kind: str) -> Path:
    return find_input_candidates(input_dir, patterns, kind)[0]

# ---------------------------------------------------------------------------
# CLI
# ---------------------------------------------------------------------------


def build_parser(project_root: Path) -> argparse.ArgumentParser:
    parser = argparse.ArgumentParser(
        description="由 DPS / PP / CTB 原始資料生成買方報表（數值一律以原始檔為準）。",
        formatter_class=argparse.ArgumentDefaultsHelpFormatter,
    )
    parser.add_argument("--input-dir", type=Path, default=project_root / "input",
                        help="輸入資料夾")
    parser.add_argument("--out-dir", type=Path, default=project_root / "output",
                        help="輸出資料夾（不存在會自動建立）")
    parser.add_argument("--dps", type=Path, default=None,
                        help="DPS 來源檔；省略時於輸入資料夾自動尋找")
    parser.add_argument("--pp", type=Path, default=None,
                        help="PP 來源檔；省略時於輸入資料夾自動尋找")
    parser.add_argument("--skip-dps", action="store_true", help="不產生 DPS 報表")
    parser.add_argument("--skip-pp", action="store_true", help="不產生 PP 報表")
    parser.add_argument("--skip-dps-pp", action="store_true", help="不產生 DPS+PP 報表")
    parser.add_argument("--skip-ctb", action="store_true", help="不產生 CTB 報表")

    parser.add_argument("--include-star-parts", action="store_true",
                        help="保留結尾帶 * 的 DPS 料號（預設排除，與人工整理規則一致）")
    parser.add_argument("--dps-tail-cutoff", default="auto",
                        help="DPS 末欄彙總桶起始日：auto / none / YYYY-MM-DD")

    parser.add_argument("--pp-plan", default="Production Input", help="PP 的 Plan 篩選值")
    parser.add_argument("--pp-start-week", default="auto",
                        help="PP 起始週：auto（依報表日推算）或週數")
    parser.add_argument("--pp-base-year", default=None,
                        help="PP 主年度兩位數（例 26）；省略時依報表日判斷")
    parser.add_argument("--pp-report-date", type=parse_date_arg, default=None,
                        help="PP 報表基準日；省略時取樞紐快取的更新日期")
    parser.add_argument("--dps-pp-current-week", default="auto",
                        help="DPS+PP 目前週：auto（依執行日推算；週五提前用下一週）或週數")
    parser.add_argument("--raken-dps-pp-weeks", type=int, choices=(2, 3, 4), default=None,
                        help="RAKEN DPS+PP 的 DPS 保留週數（含本週）；省略時依 INI，Windows exe 會跳出選擇視窗")

    parser.add_argument("--compare", action="store_true",
                        help="與來源檔內既有的人工整理版逐格對帳並列出差異")
    parser.add_argument("--quiet", action="store_true", help="只輸出錯誤訊息")
    parser.add_argument("--no-pause", action="store_true",
                        help="Windows exe 模式下完成後不等待按 Enter")
    return parser


def _error_message(exc: BaseException) -> str:
    if isinstance(exc, SystemExit):
        return str(exc) if str(exc) else str(exc.code)
    return str(exc)


def candidate_paths(
    explicit_path: Path | None,
    input_dir: Path,
    patterns: Sequence[str],
    kind: str,
    multiple_action: str = "將依修改時間由新到舊嘗試。",
) -> list[Path]:
    if explicit_path is not None:
        return [explicit_path]
    return find_input_candidates(input_dir, patterns, kind, multiple_action)


def has_excel_files(input_dir: Path) -> bool:
    return input_dir.is_dir() and any(
        path.is_file() and path.suffix.lower() == ".xlsx" and not path.name.startswith("~$")
        for path in input_dir.glob("*.xlsx")
    )


def update_customer_dps_pp_weeks(args, customer_name: str, weeks: int, source: str) -> None:
    for customer in args.customers:
        if customer["name"].casefold() == customer_name.casefold():
            customer["dps_pp_dps_weeks_ahead"] = weeks
            customer["dps_pp_dps_weeks_source"] = source
            return


def should_prompt_raken_dps_pp_weeks(args) -> bool:
    if not is_frozen_app() or not sys.platform.startswith("win"):
        return False
    if args.raken_dps_pp_weeks is not None:
        return False
    if args.skip_dps or args.skip_pp or args.skip_dps_pp:
        return False
    if args.dps is not None or args.pp is not None:
        return False
    if getattr(args, "customer_scope", "all") == "avtc":
        return False
    return has_excel_files(args.input_dir / "RAKEN")


def _center_tk_window(root) -> None:
    root.update_idletasks()
    width = root.winfo_width()
    height = root.winfo_height()
    left = max((root.winfo_screenwidth() - width) // 2, 0)
    top = max((root.winfo_screenheight() - height) // 2, 0)
    root.geometry(f"+{left}+{top}")


def select_raken_dps_pp_weeks(default_weeks: int) -> int | None:
    try:
        import tkinter as tk
        from tkinter import ttk
    except Exception as exc:  # pragma: no cover - depends on Windows runtime
        warn(f"無法開啟 RAKEN 週數選擇視窗，已使用設定檔週數。原因：{exc}")
        return None

    options = (2, 3, 4)
    if default_weeks not in options:
        default_weeks = 3

    selected = {"value": None}

    root = tk.Tk()
    root.title("RAKEN DPS+PP 週數選擇")
    root.resizable(False, False)
    root.attributes("-topmost", True)

    frame = ttk.Frame(root, padding=20)
    frame.grid(row=0, column=0, sticky="nsew")

    ttk.Label(
        frame,
        text="請選擇 RAKEN DPS+PP 的 DPS 保留週數",
        font=("Microsoft JhengHei UI", 11, "bold"),
    ).grid(row=0, column=0, columnspan=3, sticky="w")
    ttk.Label(
        frame,
        text="此選項只影響 RAKEN 的 DPS+PP.xlsx，週數包含目前週。",
    ).grid(row=1, column=0, columnspan=3, sticky="w", pady=(6, 14))

    def choose(value: int) -> None:
        selected["value"] = value
        root.destroy()

    for col, weeks in enumerate(options):
        label = f"{weeks} 週"
        if weeks == default_weeks:
            label += "（預設）"
        ttk.Button(
            frame,
            text=label,
            command=lambda value=weeks: choose(value),
            width=14,
        ).grid(row=2, column=col, padx=4, ipadx=6, ipady=6)

    ttk.Label(
        frame,
        text="關閉視窗或按 Esc 會沿用預設值。",
    ).grid(row=3, column=0, columnspan=3, sticky="w", pady=(14, 0))

    def use_default(_event=None) -> None:
        choose(default_weeks)

    root.protocol("WM_DELETE_WINDOW", use_default)
    root.bind("<Escape>", use_default)
    _center_tk_window(root)
    root.mainloop()
    return selected["value"]


def apply_raken_dps_pp_weeks_selection(args) -> None:
    args.raken_dps_pp_weeks_source = None
    if args.raken_dps_pp_weeks is not None:
        update_customer_dps_pp_weeks(
            args,
            "RAKEN",
            args.raken_dps_pp_weeks,
            "命令列參數 --raken-dps-pp-weeks",
        )
        args.raken_dps_pp_weeks_source = "命令列參數 --raken-dps-pp-weeks"
        return

    if not should_prompt_raken_dps_pp_weeks(args):
        return

    default_weeks = next(
        (
            customer["dps_pp_dps_weeks_ahead"]
            for customer in args.customers
            if customer["name"].casefold() == "raken"
        ),
        2,
    )
    selected_weeks = select_raken_dps_pp_weeks(default_weeks)
    if selected_weeks is None:
        return
    update_customer_dps_pp_weeks(
        args,
        "RAKEN",
        selected_weeks,
        "Windows 啟動畫面選擇",
    )
    args.raken_dps_pp_weeks_source = "Windows 啟動畫面選擇"


def select_customer_scope() -> str | None:
    try:
        import tkinter as tk
        from tkinter import ttk
    except Exception as exc:  # pragma: no cover - depends on Windows runtime
        warn(f"無法開啟客戶選擇視窗，將使用全部客戶。原因：{exc}")
        return "all"

    selected = {"value": None}
    try:
        root = tk.Tk()
    except Exception as exc:  # pragma: no cover - depends on Windows runtime
        warn(f"無法開啟客戶選擇視窗，將使用全部客戶。原因：{exc}")
        return "all"
    root.title("Buyer Reports 客戶選擇")
    root.resizable(False, False)
    root.attributes("-topmost", True)

    frame = ttk.Frame(root, padding=20)
    frame.grid(row=0, column=0, sticky="nsew")
    ttk.Label(
        frame,
        text="請選擇本次要處理的客戶",
        font=("Microsoft JhengHei UI", 11, "bold"),
    ).grid(row=0, column=0, columnspan=3, sticky="w")
    ttk.Label(
        frame,
        text="選擇 RAKEN 或全部執行時，接著會設定 RAKEN DPS+PP 週次。",
    ).grid(row=1, column=0, columnspan=3, sticky="w", pady=(6, 14))

    def choose(value: str) -> None:
        selected["value"] = value
        root.destroy()

    options = (("AVTC", "avtc"), ("RAKEN", "raken"), ("全部執行", "all"))
    for col, (label, value) in enumerate(options):
        ttk.Button(
            frame,
            text=label,
            command=lambda choice=value: choose(choice),
            width=14,
        ).grid(row=2, column=col, padx=4, ipadx=6, ipady=6)

    ttk.Label(
        frame,
        text="關閉視窗會取消本次執行。",
    ).grid(row=3, column=0, columnspan=3, sticky="w", pady=(14, 0))
    root.protocol("WM_DELETE_WINDOW", root.destroy)
    root.bind("<Escape>", lambda _event: root.destroy())
    _center_tk_window(root)
    root.mainloop()
    return selected["value"]


def should_prompt_customer_scope(args) -> bool:
    return (
        is_frozen_app()
        and sys.platform.startswith("win")
        and not args.quiet
        and args.dps is None
        and args.pp is None
    )


def _ctb_task_enabled_for_context(args, context: RunContext) -> bool:
    if context.name.casefold() == "raken":
        return (
            not args.skip_ctb
            and not args.skip_dps
            and not args.skip_pp
            and not args.skip_dps_pp
            and has_raken_ctb_input_candidates(context.input_dir)
        )
    return (
        not args.skip_ctb
        and not args.skip_dps
        and not args.skip_pp
        and not args.skip_dps_pp
        and has_ctb_input_candidates(context.input_dir)
    )


def _collect_ctb_supplier_sites(args, contexts: Sequence[RunContext]) -> tuple[str, ...]:
    detected: dict[str, str] = {}
    for context in contexts:
        if not _ctb_task_enabled_for_context(args, context):
            continue
        if context.name.casefold() == "raken":
            # RAKEN 的 PO pivot 沒有 Supplier Site；不要用 AVTC 的 open po
            # 偵測方式，也不要因為此欄缺少而把 RAKEN 視為來源錯誤。
            continue
        title = f"{context.label} CTB"
        try:
            open_po_path = find_workbook_with_sheet(
                context.input_dir,
                "open po",
                f"{title} open po",
            )
            records = read_open_po(open_po_path)
        except (SystemExit, Exception) as exc:  # noqa: BLE001 - CTB 本身會再回報來源錯誤
            warn(f"{title} 無法讀取 Supplier site，將使用既有設定。原因：{_error_message(exc)}")
            continue
        for record in records:
            display = str(record.supplier_site or "").strip()
            key = display.casefold()
            if key:
                detected.setdefault(key, display)
    return tuple(detected.values())


def _show_ctb_new_supplier_site_warning(new_sites: Sequence[str]) -> None:
    if not new_sites:
        return
    message = (
        "本次偵測到新的 Supplier site：\n"
        "\n"
        + "\n".join(f"- {site}" for site in new_sites)
    )
    try:
        import tkinter as tk
        from tkinter import messagebox

        root = tk.Tk()
        root.withdraw()
        root.attributes("-topmost", True)
        try:
            messagebox.showwarning(
                "CTB ETA 設定提醒",
                message,
                parent=root,
            )
        finally:
            root.destroy()
    except Exception as exc:  # pragma: no cover - depends on local desktop/display
        warn(f"無法開啟新的 Supplier site 警告視窗，已改在終端機顯示。原因：{exc}")
        warn(message.replace("\n", "；"))


def _supplier_site_matches_query(site: str, query: str) -> bool:
    """Match a Supplier site using literal search or the custom ``**`` wildcard."""
    query = query.strip().casefold()
    if not query:
        return True

    site = site.casefold()
    if "**" not in query:
        return query in site

    # Only a pair of stars is special; single stars remain literal site text.
    pattern = ".*".join(re.escape(part) for part in query.split("**"))
    return re.fullmatch(pattern, site, flags=re.DOTALL) is not None


def _show_ctb_eta_settings_dialog(settings: dict) -> dict | None:
    try:
        import tkinter as tk
        from tkinter import messagebox, ttk
    except Exception as exc:  # pragma: no cover - depends on Windows runtime
        warn(f"無法開啟 CTB ETA 設定視窗，將沿用目前設定。原因：{exc}")
        return settings

    try:
        root = tk.Tk()
    except Exception as exc:  # pragma: no cover - depends on Windows runtime
        warn(f"無法建立 CTB ETA 設定視窗，將沿用目前設定。原因：{exc}")
        return settings

    entries: dict[str, dict] = {}
    for key, (display, days) in settings["confirmed_supplier_site_entries"].items():
        entries[key] = {
            "display": display,
            "days": days,
            "section": "confirmed",
        }
    for key, (display, days) in settings["new_supplier_site_entries"].items():
        if key not in entries:
            entries[key] = {
                "display": display,
                "days": days,
                "section": "new",
            }

    detected_keys = set(settings.get("detected_supplier_site_keys", ()))
    edited_keys: set[str] = set()
    result = {"value": None}
    sort_state = {"column": "status", "reverse": False}
    selection_anchor = {"key": None}
    active_editor = {"widget": None, "key": None, "committing": False}

    root.title("CTB ETA Supplier site 設定")
    root.geometry("820x660")
    root.minsize(720, 540)
    root.attributes("-topmost", True)

    main_frame = ttk.Frame(root, padding=16)
    main_frame.grid(row=0, column=0, sticky="nsew")
    root.columnconfigure(0, weight=1)
    root.rowconfigure(0, weight=1)
    main_frame.columnconfigure(0, weight=1)
    main_frame.rowconfigure(5, weight=1)

    ttk.Label(
        main_frame,
        text="CTB ETA Supplier site 設定",
        font=("Microsoft JhengHei UI", 12, "bold"),
    ).grid(row=0, column=0, columnspan=3, sticky="w")
    ttk.Label(
        main_frame,
        text="可修改已確認 Supplier site，也可設定本次新偵測 Supplier site；提前天數以日曆日計算。",
    ).grid(row=1, column=0, columnspan=3, sticky="w", pady=(6, 12))

    default_days_var = tk.StringVar(value=str(settings["default_lead_days"]))
    selected_days_var = tk.StringVar()
    search_var = tk.StringVar()

    search_frame = ttk.Frame(main_frame)
    search_frame.grid(row=2, column=0, columnspan=3, sticky="ew", pady=(0, 8))
    ttk.Label(search_frame, text="搜尋 Supplier site：").grid(row=0, column=0, sticky="w")
    search_entry = ttk.Entry(search_frame, textvariable=search_var, width=34)
    search_entry.grid(row=0, column=1, padx=(6, 10), sticky="w")

    default_frame = ttk.Frame(main_frame)
    default_frame.grid(row=3, column=0, columnspan=3, sticky="ew", pady=(0, 8))
    ttk.Label(default_frame, text="預設提前天數（日）：").grid(row=0, column=0, sticky="w")
    default_entry = ttk.Entry(default_frame, textvariable=default_days_var, width=10)
    default_entry.grid(row=0, column=1, padx=(6, 10), sticky="w")

    selected_frame = ttk.Frame(main_frame)
    selected_frame.grid(row=4, column=0, columnspan=3, sticky="ew", pady=(0, 8))
    ttk.Label(
        selected_frame,
        text="將選取的 Supplier site 設定提前天數：",
    ).grid(row=0, column=0, sticky="w")
    selected_days_entry = ttk.Entry(selected_frame, textvariable=selected_days_var, width=10)
    selected_days_entry.grid(row=0, column=1, padx=(6, 10), sticky="w")

    table_frame = ttk.Frame(main_frame)
    table_frame.grid(row=5, column=0, columnspan=3, sticky="nsew")
    table_frame.columnconfigure(0, weight=1)
    table_frame.columnconfigure(1, minsize=220)
    table_frame.rowconfigure(1, weight=1)
    tree = ttk.Treeview(
        table_frame,
        columns=("site", "days"),
        show="",
        selectmode="extended",
    )
    tree.heading("site", text="Supplier site")
    tree.heading("days", text="提前天數（日）")
    tree.column("site", width=320, anchor="w", stretch=True)
    tree.column("days", width=160, anchor="center", stretch=False)
    tree.grid(row=1, column=0, sticky="nsew")

    status_tree = ttk.Treeview(
        table_frame,
        columns=("status",),
        show="",
        selectmode="none",
    )
    status_tree.heading("status", text="狀態")
    status_tree.column("status", width=220, anchor="w", stretch=False)
    status_tree.grid(row=1, column=1, sticky="nsew")

    def scroll_trees(*args) -> None:
        tree.yview(*args)
        status_tree.yview(*args)

    scrollbar = ttk.Scrollbar(table_frame, orient="vertical", command=scroll_trees)
    scrollbar.grid(row=1, column=2, sticky="ns")

    def sync_status_scroll(first, last) -> None:
        status_tree.yview_moveto(first)
        scrollbar.set(first, last)

    tree.configure(yscrollcommand=sync_status_scroll)
    status_tree.configure(yscrollcommand=lambda *_args: None)

    ttk.Label(
        main_frame,
        text="新 Supplier site 會在下一次執行時移至已確認設定區。",
    ).grid(row=6, column=0, columnspan=3, sticky="w", pady=(10, 12))

    button_frame = ttk.Frame(main_frame)
    button_frame.grid(row=7, column=0, columnspan=3, sticky="e")

    row_iids: dict[str, str] = {}

    def parse_days(raw_value: str, label: str) -> int | None:
        try:
            days = int(raw_value.strip())
        except ValueError:
            messagebox.showerror("輸入錯誤", f"{label} 必須是 0 或以上的整數。", parent=root)
            return None
        if days < 0:
            messagebox.showerror("輸入錯誤", f"{label} 必須是 0 或以上的整數。", parent=root)
            return None
        return days

    def status_for(key: str) -> str:
        entry = entries[key]
        if key in edited_keys:
            return "已填寫（新 Supplier site）" if entry["section"] == "new" else "已設定"
        if entry["section"] == "new":
            return "新偵測"
        if key not in detected_keys:
            return "已設定（本次未偵測）"
        return "已設定"

    def status_tag_for(key: str) -> tuple[str, ...]:
        entry = entries[key]
        if entry["section"] == "new" and key not in edited_keys:
            return ("new_supplier_site_status",)
        return ()

    def status_sort_rank(key: str) -> int:
        if status_tag_for(key):
            return 0
        if entries[key]["section"] == "new":
            return 1
        if key in detected_keys:
            return 2
        return 3

    def filtered_keys() -> list[str]:
        query = search_var.get().strip()
        visible = [
            key
            for key, entry in entries.items()
            if _supplier_site_matches_query(entry["display"], query)
        ]
        if sort_state["column"] == "site":
            key_func = lambda key: (entries[key]["display"].casefold(),)
        elif sort_state["column"] == "days":
            key_func = lambda key: (entries[key]["days"], entries[key]["display"].casefold())
        else:
            key_func = lambda key: (status_sort_rank(key), entries[key]["display"].casefold())
        return sorted(visible, key=key_func, reverse=sort_state["reverse"])

    def selected_keys() -> set[str]:
        return {
            row_iids[item_id]
            for item_id in tree.selection()
            if item_id in row_iids
        }

    def refresh_tree() -> None:
        preserved_keys = selected_keys()
        visible_keys = filtered_keys()
        for widget in (tree, status_tree):
            for item_id in widget.get_children():
                widget.delete(item_id)
        row_iids.clear()
        for index, key in enumerate(visible_keys):
            entry = entries[key]
            item_id = f"site_{index}"
            row_iids[item_id] = key
            tree.insert(
                "",
                "end",
                iid=item_id,
                values=(entry["display"], entry["days"]),
            )
            status_tree.insert(
                "",
                "end",
                iid=item_id,
                values=(status_for(key),),
                tags=status_tag_for(key),
            )
        restored_items = [
            item_id for item_id, key in row_iids.items() if key in preserved_keys
        ]
        if restored_items:
            tree.selection_set(restored_items)
            tree.focus(restored_items[-1])
        elif visible_keys:
            tree.selection_set(())
        selection_anchor["key"] = (
            selection_anchor["key"]
            if selection_anchor["key"] in visible_keys
            else None
        )
        tree.yview_moveto(0)
        status_tree.yview_moveto(0)

    def sort_entries(column: str, reverse: bool) -> None:
        sort_state["column"] = column
        sort_state["reverse"] = reverse
        refresh_tree()

    def make_sort_header(parent, title: str, column: str):
        header = ttk.Frame(parent, relief="raised", borderwidth=1)
        header.columnconfigure(0, weight=1)
        ttk.Label(
            header,
            text=title,
            anchor="center",
            style="Treeview.Heading",
        ).grid(row=0, column=0, sticky="ew", padx=(4, 2))
        ttk.Button(
            header,
            text="↑",
            width=2,
            command=lambda: sort_entries(column, False),
        ).grid(row=0, column=1, padx=(0, 1), pady=1)
        ttk.Button(
            header,
            text="↓",
            width=2,
            command=lambda: sort_entries(column, True),
        ).grid(row=0, column=2, padx=(0, 2), pady=1)
        return header

    def clear_search() -> None:
        search_var.set("")
        search_entry.focus_set()

    def select_all(_event=None) -> str:
        visible_items = tree.get_children()
        if visible_items:
            tree.selection_set(visible_items)
            tree.focus(visible_items[0])
            selection_anchor["key"] = row_iids[visible_items[0]]
        return "break"

    def record_tree_click(event) -> None:
        item_id = tree.identify_row(event.y)
        if item_id and not (event.state & 0x0001 or event.state & 0x0004):
            selection_anchor["key"] = row_iids[item_id]

    def on_status_click(event) -> str:
        item_id = status_tree.identify_row(event.y)
        if item_id not in row_iids:
            return "break"
        tree.focus_set()
        visible_items = list(tree.get_children())
        index = visible_items.index(item_id)
        control_pressed = bool(event.state & 0x0004)
        shift_pressed = bool(event.state & 0x0001)
        anchor = next(
            (
                visible_item
                for visible_item, key in row_iids.items()
                if key == selection_anchor["key"]
            ),
            None,
        )
        if anchor not in visible_items:
            anchor = tree.focus() if tree.focus() in visible_items else item_id

        if shift_pressed:
            anchor_index = visible_items.index(anchor)
            start, end = sorted((anchor_index, index))
            range_items = visible_items[start : end + 1]
            if control_pressed:
                tree.selection_add(range_items)
            else:
                tree.selection_set(range_items)
        elif control_pressed:
            tree.selection_toggle(item_id)
        else:
            tree.selection_set(item_id)
            selection_anchor["key"] = row_iids[item_id]
        tree.focus(item_id)
        return "break"

    def on_tree_mousewheel(event) -> str:
        delta = getattr(event, "delta", 0)
        if delta:
            amount = -1 if delta > 0 else 1
        else:
            amount = -1 if getattr(event, "num", 5) == 4 else 1
        tree.yview_scroll(amount, "units")
        return "break"

    def close_inline_editor() -> None:
        editor = active_editor["widget"]
        if editor is not None:
            editor.destroy()
        active_editor["widget"] = None
        active_editor["key"] = None
        active_editor["committing"] = False

    def cancel_inline_editor(_event=None) -> str:
        if not active_editor["committing"]:
            close_inline_editor()
        return "break"

    def commit_inline_editor(key: str, editor) -> str:
        if active_editor["widget"] is not editor:
            return "break"
        days = parse_days(editor.get(), "Supplier site 提前天數")
        if days is None:
            editor.focus_set()
            return "break"
        active_editor["committing"] = True
        confirmed = messagebox.askyesno(
            "確認修改",
            f"確定將 {entries[key]['display']} 的提前天數設定為 {days} 天嗎？",
            parent=root,
        )
        if confirmed:
            entries[key]["days"] = days
            edited_keys.add(key)
        close_inline_editor()
        if confirmed:
            refresh_tree()
        return "break"

    def start_inline_editor(event) -> None:
        item_id = tree.identify_row(event.y)
        column_id = tree.identify_column(event.x)
        if item_id not in row_iids or column_id != "#2":
            return
        close_inline_editor()
        bbox = tree.bbox(item_id, column_id)
        if not bbox:
            return
        key = row_iids[item_id]
        _x, y, width, height = bbox
        editor = ttk.Entry(tree, justify="center")
        editor.insert(0, str(entries[key]["days"]))
        editor.select_range(0, "end")
        editor.place(x=_x, y=y, width=width, height=height)
        active_editor["widget"] = editor
        active_editor["key"] = key
        editor.focus_set()
        editor.bind("<Return>", lambda _event: commit_inline_editor(key, editor))
        editor.bind("<Escape>", cancel_inline_editor)
        editor.bind(
            "<FocusOut>",
            lambda _event: root.after_idle(
                lambda: cancel_inline_editor()
                if not active_editor["committing"]
                else None
            ),
        )

    def apply_selected() -> None:
        keys = selected_keys()
        if not keys:
            messagebox.showwarning(
                "未選取 Supplier site",
                "請先選取至少一個 Supplier site。",
                parent=root,
            )
            return
        days = parse_days(selected_days_var.get(), "選取 Supplier site 提前天數")
        if days is None:
            return
        if not messagebox.askyesno(
            "確認修改",
            f"確定將選取的 {len(keys)} 個 Supplier site 設定為 {days} 天嗎？",
            parent=root,
        ):
            return
        for key in keys:
            entries[key]["days"] = days
            edited_keys.add(key)
        refresh_tree()

    def apply_default_to_all() -> None:
        default_days = parse_days(default_days_var.get(), "預設提前天數")
        if default_days is None or not entries:
            return
        if not messagebox.askyesno(
            "確認套用預設值",
            f"確定要將全部 {len(entries)} 個 Supplier site 的提前天數設為 "
            f"{default_days} 天嗎？",
            parent=root,
        ):
            return
        for key, entry in entries.items():
            entry["days"] = default_days
            edited_keys.add(key)
        refresh_tree()

    def on_escape(_event=None) -> str:
        if active_editor["widget"] is not None:
            return cancel_inline_editor()
        root.destroy()
        return "break"

    def confirm() -> None:
        default_days = parse_days(default_days_var.get(), "預設提前天數")
        if default_days is None:
            return
        confirmed_entries = {
            key: (entry["display"], entry["days"])
            for key, entry in entries.items()
            if entry["section"] == "confirmed"
        }
        new_entries = {
            key: (entry["display"], entry["days"])
            for key, entry in entries.items()
            if entry["section"] == "new"
        }
        try:
            saved = save_ctb_eta_config(
                settings["path"],
                default_days,
                confirmed_entries,
                new_entries,
            )
        except Exception as exc:  # pragma: no cover - depends on filesystem permissions
            messagebox.showerror("設定檔錯誤", f"無法儲存 CTB ETA 設定：{exc}", parent=root)
            return
        result["value"] = saved
        root.destroy()

    search_var.trace_add("write", lambda *_args: refresh_tree())
    tree.bind("<Button-1>", record_tree_click, add="+")
    tree.bind("<Double-1>", start_inline_editor)
    tree.bind("<Control-a>", select_all)
    tree.bind("<Control-A>", select_all)
    status_tree.bind("<Button-1>", on_status_click)
    for widget in (tree, status_tree):
        widget.bind("<MouseWheel>", on_tree_mousewheel)
        widget.bind("<Button-4>", on_tree_mousewheel)
        widget.bind("<Button-5>", on_tree_mousewheel)
    status_tree.tag_configure(
        "new_supplier_site_status",
        background="#f4cccc",
        foreground="#9c0006",
    )

    header_main = ttk.Frame(table_frame)
    header_main.grid(row=0, column=0, sticky="nsew")
    header_main.columnconfigure(0, weight=1)
    header_main.columnconfigure(1, minsize=160)
    make_sort_header(header_main, "Supplier site", "site").grid(
        row=0,
        column=0,
        sticky="nsew",
    )
    make_sort_header(header_main, "提前天數（日）", "days").grid(
        row=0,
        column=1,
        sticky="nsew",
    )
    make_sort_header(table_frame, "狀態", "status").grid(
        row=0,
        column=1,
        sticky="nsew",
    )

    ttk.Button(search_frame, text="清除", command=clear_search).grid(row=0, column=2)
    ttk.Button(
        default_frame,
        text="全部套用預設值",
        command=apply_default_to_all,
    ).grid(row=0, column=2, padx=(0, 10))
    ttk.Button(
        selected_frame,
        text="套用選取項目",
        command=apply_selected,
    ).grid(row=0, column=2, padx=(0, 10))
    ttk.Button(button_frame, text="取消執行", command=root.destroy).grid(row=0, column=0, padx=(0, 8))
    ttk.Button(button_frame, text="確定並開始執行", command=confirm).grid(row=0, column=1)
    selected_days_entry.bind("<Return>", lambda _event: apply_selected())

    refresh_tree()
    root.protocol("WM_DELETE_WINDOW", root.destroy)
    root.bind("<Escape>", on_escape)
    root.grab_set()
    _center_tk_window(root)
    root.mainloop()
    return result["value"]


def _edit_ctb_eta_config(path: Path) -> None:
    try:
        if sys.platform.startswith("win"):
            os.startfile(str(path))  # type: ignore[attr-defined]  # pragma: no cover - Windows runtime
        elif sys.platform == "darwin":
            subprocess.Popen(
                ["open", str(path)],
                stdout=subprocess.DEVNULL,
                stderr=subprocess.DEVNULL,
            )
        else:
            subprocess.Popen(
                ["xdg-open", str(path)],
                stdout=subprocess.DEVNULL,
                stderr=subprocess.DEVNULL,
            )
    except OSError as exc:
        warn(f"無法使用作業系統預設程式開啟設定檔，請手動編輯 {path}。原因：{exc}")
    input("設定檔開啟後，編輯並關閉檔案，再按 Enter 繼續：")


def _apply_ctb_eta_settings(args, settings: dict) -> None:
    args.ctb_eta_config_path = settings["path"]
    args.ctb_eta_default_lead_days = settings["default_lead_days"]
    args.ctb_eta_lead_days_by_supplier_site = settings["lead_days_by_supplier_site"]


def prepare_ctb_eta_settings(args, contexts: Sequence[RunContext]) -> None:
    default_path = project_root() / CTB_ETA_CONFIG_FILE_NAME
    args.ctb_eta_config_path = default_path
    args.ctb_eta_default_lead_days = DEFAULT_CTB_ETA_LEAD_DAYS
    args.ctb_eta_lead_days_by_supplier_site = {}
    args.startup_cancelled = False

    ctb_contexts = [
        context
        for context in contexts
        if _ctb_task_enabled_for_context(args, context)
    ]
    if not ctb_contexts:
        return

    detected_sites = _collect_ctb_supplier_sites(args, ctb_contexts)
    settings = sync_ctb_eta_config(project_root(), detected_sites)
    _apply_ctb_eta_settings(args, settings)
    log(f"CTB ETA 設定檔：{settings['path']}")
    log(
        f"CTB ETA 天數    ：預設 {settings['default_lead_days']} 天，"
        f"Supplier site 覆寫 {len(settings['lead_days_by_supplier_site'])} 項"
    )
    if settings["promoted_supplier_sites"]:
        log(
            "CTB ETA 已移至已確認區："
            + "、".join(settings["promoted_supplier_sites"])
        )
    if settings["new_supplier_sites"]:
        log()
        log("CTB ETA 新 Supplier site（目前先使用預設天數）：")
        for site in settings["new_supplier_sites"]:
            log(f"  - {site}")
        log()

    if should_prompt_customer_scope(args):
        _show_ctb_new_supplier_site_warning(settings["new_supplier_sites"])
        edited_settings = _show_ctb_eta_settings_dialog(settings)
        if edited_settings is None:
            args.startup_cancelled = True
            log("使用者取消 CTB ETA 設定，本次執行已取消。")
            return
        _apply_ctb_eta_settings(args, edited_settings)
        log("CTB ETA 設定已由啟動畫面確認，當次 CTB 將使用畫面上的天數。")
        return

    if args.quiet or not sys.stdin.isatty():
        return

    try:
        answer = input("是否要開啟 CTB ETA Supplier site 設定檔調整天數？(Y/N)：")
    except EOFError:
        answer = "N"
    if answer.strip().casefold() not in {"y", "yes"}:
        log("CTB ETA 設定檔未開啟，沿用目前設定。")
        return

    _show_ctb_new_supplier_site_warning(settings["new_supplier_sites"])
    _edit_ctb_eta_config(settings["path"])
    _apply_ctb_eta_settings(args, load_ctb_eta_config(settings["path"]))
    log("CTB ETA 設定檔已重新讀取，當次 CTB 將使用編輯後的天數。")


def build_run_contexts(args) -> list[RunContext]:
    args.context_warnings = []
    for customer in args.customers:
        (args.input_dir / customer["name"]).mkdir(parents=True, exist_ok=True)

    if args.dps is not None or args.pp is not None:
        return [
            RunContext(
                name="",
                input_dir=args.input_dir,
                out_dir=args.out_dir,
                dps_mode="first_valid",
                pp_mode="first_valid",
                dps_pp_dps_weeks_ahead=5,
                dps_pp_dps_weeks_source="舊版根目錄預設",
                dps_pp_late_dps_mode="merge_to_cutoff",
                dps_drop_zero_total_rows=False,
                dps_trim_trailing_zero_dates=False,
                legacy=True,
            )
        ]

    contexts = []
    for customer in args.customers:
        customer_scope = getattr(args, "customer_scope", "all")
        if (
            customer_scope != "all"
            and customer["name"].casefold() != customer_scope.casefold()
        ):
            continue
        customer_dir = args.input_dir / customer["name"]
        if not has_excel_files(customer_dir):
            continue
        contexts.append(
            RunContext(
                name=customer["name"],
                input_dir=customer_dir,
                out_dir=args.out_dir / customer["name"],
                dps_mode=customer["dps_mode"],
                pp_mode=customer["pp_mode"],
                dps_pp_dps_weeks_ahead=customer["dps_pp_dps_weeks_ahead"],
                dps_pp_dps_weeks_source=customer.get(
                    "dps_pp_dps_weeks_source",
                    "buyer_reports.ini / 內建預設",
                ),
                dps_pp_late_dps_mode=customer["dps_pp_late_dps_mode"],
                dps_drop_zero_total_rows=customer["dps_drop_zero_total_rows"],
                dps_trim_trailing_zero_dates=customer["dps_trim_trailing_zero_dates"],
            )
        )

    if contexts:
        if has_excel_files(args.input_dir):
            args.context_warnings.append(
                "已偵測到客戶資料夾內有 Excel，input 根目錄下的 Excel 將不處理。"
            )
        return contexts

    if has_excel_files(args.input_dir):
        args.context_warnings.append(
            "未偵測到 AVTC/RAKEN 客戶資料夾內有 Excel，改用舊版 input 根目錄流程。"
        )
        return [
            RunContext(
                name="",
                input_dir=args.input_dir,
                out_dir=args.out_dir,
                dps_mode="first_valid",
                pp_mode="first_valid",
                dps_pp_dps_weeks_ahead=5,
                dps_pp_dps_weeks_source="舊版根目錄預設",
                dps_pp_late_dps_mode="merge_to_cutoff",
                dps_drop_zero_total_rows=False,
                dps_trim_trailing_zero_dates=False,
                legacy=True,
            )
        ]

    return []


def log_path_label(paths: Sequence[Path]) -> str:
    return "、".join(str(path) for path in paths)


def log_dropped_zero_dps_rows(rows: Sequence[dict]) -> None:
    if not rows:
        return
    log(f"  已略過 0 數量 DPS 列：{len(rows)} 列")
    for item in rows[:20]:
        source = item.get("source")
        source_name = source.name if isinstance(source, Path) else str(source)
        log(f"      - {source_name} row {item['row']}：{item['part_number']}")
    if len(rows) > 20:
        log(f"      ...（其餘 {len(rows) - 20} 列省略）")


def log_trimmed_trailing_zero_dates(dates: Sequence[dt.date], label: str = "DPS") -> None:
    if not dates:
        return
    if len(dates) == 1:
        log(f"  {label} 尾端空白日期欄：已略過 1 欄（{dates[0]}）")
    else:
        log(
            f"  {label} 尾端空白日期欄：已略過 {len(dates)} 欄"
            f"（{dates[0]} ~ {dates[-1]}）"
        )


def resolve_dps_tail_cutoff(
    dps_path: Path,
    tail_cutoff_arg: str,
    sheet_keywords: Sequence[str] = DPS_SOURCE_SHEET_KEYWORDS,
    part_number_headers: Sequence[str] = DPS_PART_NUMBER_HEADERS,
) -> dt.date | None:
    if tail_cutoff_arg.lower() == "none":
        return None
    if tail_cutoff_arg.lower() == "auto":
        wb = load_workbook(dps_path, data_only=True)
        try:
            ws = first_sheet_matching_keywords(wb, sheet_keywords)
            header_row, _pn_col, _pn_header = find_dps_header(ws, part_number_headers)
            dates = [d for d in (header_date(c) for c in ws[header_row]) if d]
            max_date = max(dates) if dates else None
        finally:
            wb.close()
        return detect_dps_tail_cutoff(dps_path, max_date) if max_date else None
    return parse_date_arg(tail_cutoff_arg)


def resolve_dps_tail_cutoff_for_paths(
    dps_paths: Sequence[Path],
    args,
    sheet_keywords: Sequence[str] = DPS_SOURCE_SHEET_KEYWORDS,
    part_number_headers: Sequence[str] = DPS_PART_NUMBER_HEADERS,
) -> dt.date | None:
    if args.dps_tail_cutoff.lower() == "none":
        return None
    if args.dps_tail_cutoff.lower() != "auto":
        return parse_date_arg(args.dps_tail_cutoff)

    cutoffs = []
    for dps_path in dps_paths:
        try:
            cutoff = resolve_dps_tail_cutoff(
                dps_path,
                args.dps_tail_cutoff,
                sheet_keywords=sheet_keywords,
                part_number_headers=part_number_headers,
            )
        except (SystemExit, Exception) as exc:  # noqa: BLE001 - 壞檔稍後合併時會再略過
            warn(f"DPS 來源檔 {dps_path.name} 無法推斷末欄彙總桶，已略過此步。原因：{exc}")
            continue
        if cutoff is not None:
            cutoffs.append(cutoff)

    unique = sorted(set(cutoffs))
    if not unique:
        return None
    if len(unique) > 1:
        warn(
            "多份 DPS 來源檔推斷出不同末欄彙總桶，"
            f"將使用最早日期 {unique[0]}；候選：{', '.join(map(str, unique))}"
        )
    return unique[0]


def log_single_dps_info(info: dict, dps_out: Path, title: str) -> None:
    log(f"\n--- {title} ---")
    log(f"  來源            ：{info['source'].name}")
    log(f"  來源工作表      ：{info['source_sheet']}")
    log(f"  料號欄          ：{info['part_number_header']}")
    log(f"  表頭列          ：第 {info['header_row']} 列")
    log(f"  日期欄          ：{info['date_columns']} 欄 / {info['dates']} 個日期"
        f"（{info['date_range'][0]} ~ {info['date_range'][1]}，D+N 兩班合併）")
    if info["tail_cutoff"]:
        log(f"  末欄彙總桶      ：{info['tail_cutoff']} 起之日期併入同一欄"
            f"（沿用既有整理後版面）")
    else:
        log("  末欄彙總桶      ：未使用，每個日期各自成欄")
    log(f"  輸出            ：{info['rows']} 列 x {info['out_columns']} 個日期欄，"
        f"合計 {info['grand_total']:,} pcs")
    if info["excluded"]:
        total = clean_number(sum(info["excluded"].values()))
        log(f"  已排除 * 料號   ：{len(info['excluded'])} 個，合計 {total:,} pcs"
            f"（用 --include-star-parts 可保留）")
        for pn, qty in sorted(info["excluded"].items()):
            log(f"      - {pn}  {clean_number(qty):,}")
    if info["text_cells"]:
        log(f"  日期區文字格    ：{info['text_cells']} 格（已當 0 計）")
    log_dropped_zero_dps_rows(info.get("dropped_zero_rows", []))
    log_trimmed_trailing_zero_dates(info.get("trimmed_trailing_zero_dates", []))
    log("  輸出格式        ：料號文字；數據與 total 為數值")
    log(f"  產出檔          ：{dps_out}")


def log_merged_dps_info(info: dict, dps_out: Path, title: str) -> None:
    log(f"\n--- {title} ---")
    log(f"  模式            ：合併所有可用 DPS 檔")
    log(f"  成功併入        ：{len(info['sources'])} 份")
    for detail in info["source_details"]:
        log(
            f"      - {detail['source'].name}；sheet={detail['source_sheet']}；"
            f"料號欄={detail['part_number_header']}；"
            f"日期={detail['date_range'][0]} ~ {detail['date_range'][1]}"
        )
    if info["skipped"]:
        log(f"  已略過          ：{len(info['skipped'])} 份")
        for path, reason in info["skipped"]:
            log(f"      - {path.name}：{reason}")
    log(f"  日期欄          ：{info['date_columns']} 欄 / {info['dates']} 個日期"
        f"（{info['date_range'][0]} ~ {info['date_range'][1]}，同料號同日期累加）")
    if info["tail_cutoff"]:
        log(f"  末欄彙總桶      ：{info['tail_cutoff']} 起之日期併入同一欄")
    else:
        log("  末欄彙總桶      ：未使用，每個日期各自成欄")
    log(f"  輸出            ：{info['rows']} 列 x {info['out_columns']} 個日期欄，"
        f"合計 {info['grand_total']:,} pcs")
    if info["excluded"]:
        total = clean_number(sum(info["excluded"].values()))
        log(f"  已排除 * 料號   ：{len(info['excluded'])} 個，合計 {total:,} pcs"
            f"（用 --include-star-parts 可保留）")
    if info["text_cells"]:
        log(f"  日期區文字格    ：{info['text_cells']} 格（已當 0 計）")
    log_dropped_zero_dps_rows(info.get("dropped_zero_rows", []))
    log_trimmed_trailing_zero_dates(info.get("trimmed_trailing_zero_dates", []))
    log("  輸出格式        ：料號文字；數據與 total 為數值")
    log(f"  產出檔          ：{dps_out}")


def run_dps_report(args, context: RunContext, progress: Progress | None = None) -> dict:
    last_error = None
    dps_out = context.out_dir / f"{DPS_TIDY_SHEET}.xlsx"
    title = f"{context.label} DPS" if context.name else "DPS"
    context.out_dir.mkdir(parents=True, exist_ok=True)
    if progress is not None:
        progress.step(f"{title}: 檢查來源")
    output_step_done = False
    try:
        multiple_action = (
            "將全部合併；格式不符者會略過。"
            if context.dps_mode == "merge_all"
            else "將依修改時間由新到舊嘗試。"
        )
        paths = candidate_paths(args.dps, context.input_dir, [r"DPS"], title, multiple_action)
    except (SystemExit, Exception) as exc:  # noqa: BLE001 - 找不到候選檔也只略過 DPS
        last_error = exc
        warn(f"{title} 無法產出，已略過。原因：{_error_message(exc)}")
        paths = []
    if context.dps_mode == "merge_all" and paths:
        try:
            cutoff = resolve_dps_tail_cutoff_for_paths(
                paths,
                args,
                sheet_keywords=args.dps_sheet_keywords,
                part_number_headers=args.dps_part_number_headers,
            )
            if progress is not None and not output_step_done:
                progress.step(f"{title}: 產出報表")
                output_step_done = True
            info = generate_merged_dps(
                paths,
                dps_out,
                include_star_parts=args.include_star_parts,
                tail_cutoff=cutoff,
                sheet_keywords=args.dps_sheet_keywords,
                part_number_headers=args.dps_part_number_headers,
                drop_zero_total_rows=context.dps_drop_zero_total_rows,
                trim_trailing_zero_date_columns=context.dps_trim_trailing_zero_dates,
            )
            log_merged_dps_info(info, dps_out, title)
            if args.compare:
                warn(f"{title} 是多檔合併模式，沒有單一人工整理表可逐格對帳，已略過 --compare。")
            return {
                "customer": context.name,
                "kind": "DPS",
                "ok": True,
                "source": paths,
                "output": dps_out,
            }
        except (SystemExit, Exception) as exc:  # noqa: BLE001 - 合併失敗仍要進摘要
            last_error = exc
            warn(f"{title} 無法產出，已略過。原因：{_error_message(exc)}")
            paths = []

    for dps_path in paths:
        try:
            if not dps_path.is_file():
                raise SystemExit(f"找不到 DPS 檔案：{dps_path}")

            cutoff = resolve_dps_tail_cutoff(
                dps_path,
                args.dps_tail_cutoff,
                sheet_keywords=args.dps_sheet_keywords,
                part_number_headers=args.dps_part_number_headers,
            )
            if progress is not None and not output_step_done:
                progress.step(f"{title}: 產出報表")
                output_step_done = True
            info = generate_dps(
                dps_path,
                dps_out,
                include_star_parts=args.include_star_parts,
                tail_cutoff=cutoff,
                sheet_keywords=args.dps_sheet_keywords,
                part_number_headers=args.dps_part_number_headers,
                drop_zero_total_rows=context.dps_drop_zero_total_rows,
                trim_trailing_zero_date_columns=context.dps_trim_trailing_zero_dates,
            )
            log_single_dps_info(info, dps_out, title)

            if args.compare:
                compare(
                    "DPS", dps_path, DPS_COMPARE_SHEETS, dps_out, DPS_TIDY_SHEET,
                    key_col=1, first_data_col_manual=2, first_data_col_generated=2,
                )
            return {
                "customer": context.name,
                "kind": "DPS",
                "ok": True,
                "source": dps_path,
                "output": dps_out,
            }
        except (SystemExit, Exception) as exc:  # noqa: BLE001 - 單一報表失敗要能繼續下一份
            last_error = exc
            warn(f"DPS 來源檔 {dps_path.name} 無法產出，已略過。原因：{_error_message(exc)}")
            if args.dps is not None:
                break
    if progress is not None and not output_step_done:
        progress.step(f"{title}: 略過報表")
    return {
        "customer": context.name,
        "kind": "DPS",
        "ok": False,
        "error": _error_message(last_error) if last_error else "未知錯誤",
        "output": dps_out,
        "stale_output": dps_out.exists(),
    }


def run_pp_report(args, context: RunContext, progress: Progress | None = None) -> dict:
    last_error = None
    pp_out = context.out_dir / f"{PP_TIDY_SHEET}.xlsx"
    title = f"{context.label} PP" if context.name else "PP"
    context.out_dir.mkdir(parents=True, exist_ok=True)
    if context.pp_mode != "first_valid":
        warn(f"{title} 目前不支援 {context.pp_mode}，已改用 first_valid。")
    if progress is not None:
        progress.step(f"{title}: 檢查來源")
    output_step_done = False
    try:
        paths = candidate_paths(args.pp, context.input_dir, [r"\bPP\b", r"PP"], title)
    except (SystemExit, Exception) as exc:  # noqa: BLE001 - 找不到候選檔也只略過 PP
        last_error = exc
        warn(f"{title} 無法產出，已略過。原因：{_error_message(exc)}")
        paths = []
    for pp_path in paths:
        try:
            if not pp_path.is_file():
                raise SystemExit(f"找不到 PP 檔案：{pp_path}")

            start_week = None if args.pp_start_week.lower() == "auto" else int(args.pp_start_week)
            if progress is not None and not output_step_done:
                progress.step(f"{title}: 產出報表")
                output_step_done = True
            info = generate_pp(
                pp_path,
                pp_out,
                plan=args.pp_plan,
                start_week=start_week,
                base_year=args.pp_base_year,
                report_date=args.pp_report_date,
                sheet_keywords=args.pp_sheet_keywords,
                part_number_keywords=args.pp_part_number_field_keywords,
            )
            log(f"\n--- {title} ---")
            log(f"  來源            ：{info['source'].name}")
            log(f"  版面工作表      ：{info['layout_sheet']}")
            log(f"  樞紐分析表      ：{info['pivot_table'] or '未知'}")
            log(f"  樞紐快取        ：{info['cache_part']}"
                f"（原始表 {info['cache_source_sheet'] or '未知'}，{info['records']} 筆）")
            log(f"  料號欄          ：{info['part_number_field']}")
            log(f"  快取更新        ：{info['refreshed_date']} by {info['refreshed_by'] or '未知'}")
            log(f"  報表基準日      ：{info['report_date']}"
                f"（主年度 20{info['base_year']}，起始週 WK{info['start_week']:02d}）")
            log(f"  欄位版面        ：{'取自可見樞紐報表' if info['layout_found'] else '推導模式'}")
            if info["historical_cache_periods"]:
                log(
                    "  快取補齊歷史週  ："
                    f"{', '.join(info['historical_cache_periods'])}"
                    "（起始週以前，整理後已顯示，不納入 total）"
                )
            log(
                "  total 加總範圍  ："
                f"{info['total_start_label']} 起，共 {len(info['total_periods'])} 欄"
            )
            if info["hidden_source_periods"]:
                log(
                    "  來源隱藏期間欄  ："
                    f"{', '.join(info['hidden_source_periods'])}（整理後已顯示）"
                )
            log(f"  Plan 篩選       ：{info['plan']}（{info['plan_rows']} 筆料號）")
            log(f"  期間欄          ：{len(info['periods'])} 欄 → {', '.join(info['periods'])}")
            log(f"  輸出            ：{info['rows']} 列"
                f"（已略過期間內全為 0 的 {info['dropped_zero']} 個料號），"
                f"合計 {info['grand_total']:,} pcs")
            log("  輸出格式        ：料號文字；數據與 total 為數值")
            log(f"  產出檔          ：{pp_out}")

            if info["refreshed_date"] and info["refreshed_date"] < dt.date.today() - dt.timedelta(days=45):
                warn(
                    f"PP 樞紐快取最後更新於 {info['refreshed_date']}，距今已超過 45 天，"
                    "數字可能是舊快照，請向提供者確認是否已 refresh。"
                )

            if args.compare:
                compare(
                    "PP", pp_path, PP_COMPARE_SHEETS, pp_out, PP_TIDY_SHEET,
                    key_col=2, first_data_col_manual=4, first_data_col_generated=4,
                )
            return {
                "customer": context.name,
                "kind": "PP",
                "ok": True,
                "source": pp_path,
                "output": pp_out,
            }
        except (SystemExit, Exception) as exc:  # noqa: BLE001 - 單一報表失敗要能繼續下一份
            last_error = exc
            warn(f"PP 來源檔 {pp_path.name} 無法產出，已略過。原因：{_error_message(exc)}")
            if args.pp is not None:
                break
    if progress is not None and not output_step_done:
        progress.step(f"{title}: 略過報表")
    return {
        "customer": context.name,
        "kind": "PP",
        "ok": False,
        "error": _error_message(last_error) if last_error else "未知錯誤",
        "output": pp_out,
        "stale_output": pp_out.exists(),
    }


def run_dps_pp_report(args, context: RunContext, progress: Progress | None = None) -> dict:
    last_error = None
    out_path = context.out_dir / DPS_PP_OUTPUT_NAME
    title = f"{context.label} DPS+PP" if context.name else "DPS+PP"
    context.out_dir.mkdir(parents=True, exist_ok=True)
    if progress is not None:
        progress.step(f"{title}: 檢查來源")
    output_step_done = False

    try:
        multiple_action = (
            "DPS+PP 將全部合併；格式不符者會略過。"
            if context.dps_mode == "merge_all"
            else "DPS+PP 將依修改時間由新到舊嘗試。"
        )
        dps_paths = candidate_paths(args.dps, context.input_dir, [r"DPS"], title, multiple_action)
        pp_paths = candidate_paths(args.pp, context.input_dir, [r"\bPP\b", r"PP"], title)
    except (SystemExit, Exception) as exc:  # noqa: BLE001 - 找不到候選檔也只略過 DPS+PP
        last_error = exc
        warn(f"{title} 無法產出，已略過。原因：{_error_message(exc)}")
        dps_paths = []
        pp_paths = []

    try:
        current_week = (
            None
            if args.dps_pp_current_week.lower() == "auto"
            else int(args.dps_pp_current_week)
        )
    except ValueError as exc:
        warn(f"{title} 無法產出，已略過。原因：DPS+PP 目前週必須是 auto 或週數。")
        return {
            "customer": context.name,
            "kind": "DPS+PP",
            "ok": False,
            "error": str(exc),
            "output": out_path,
            "stale_output": out_path.exists(),
        }

    for pp_path in pp_paths:
        try:
            if progress is not None and not output_step_done:
                progress.step(f"{title}: 產出報表")
                output_step_done = True
            start_week = None if args.pp_start_week.lower() == "auto" else int(args.pp_start_week)
            info = generate_dps_pp(
                dps_paths,
                pp_path,
                out_path,
                dps_mode=context.dps_mode,
                dps_weeks_ahead=context.dps_pp_dps_weeks_ahead,
                current_week=current_week,
                include_star_parts=args.include_star_parts,
                pp_plan=args.pp_plan,
                pp_start_week=start_week,
                pp_base_year=args.pp_base_year,
                pp_report_date=args.pp_report_date,
                dps_sheet_keywords=args.dps_sheet_keywords,
                dps_part_number_headers=args.dps_part_number_headers,
                pp_sheet_keywords=args.pp_sheet_keywords,
                pp_part_number_keywords=args.pp_part_number_field_keywords,
                drop_zero_total_rows=context.dps_drop_zero_total_rows,
                trim_trailing_zero_date_columns=context.dps_trim_trailing_zero_dates,
                late_dps_mode=context.dps_pp_late_dps_mode,
            )

            log(f"\n--- {title} ---")
            log(f"  DPS 來源        ：{len(info['dps_sources'])} 份")
            for detail in info["dps_source_details"]:
                log(
                    f"      - {detail['source'].name}；sheet={detail['source_sheet']}；"
                    f"料號欄={detail['part_number_header']}；"
                    f"日期={detail['date_range'][0]} ~ {detail['date_range'][1]}"
                )
            if info["skipped_dps"]:
                log(f"  DPS 已略過      ：{len(info['skipped_dps'])} 份")
                for path, reason in info["skipped_dps"]:
                    log(f"      - {path.name}：{reason}")
            log_dropped_zero_dps_rows(info.get("dps_dropped_zero_rows", []))
            log_trimmed_trailing_zero_dates(
                info.get("dps_trimmed_trailing_zero_dates", []),
                label="DPS+PP 的 DPS 區段",
            )
            log(f"  PP 來源         ：{info['pp_source'].name}")
            log(f"  PP 工作表       ：{info['pp_sheet']}；料號欄={info['pp_part_number_field']}")
            log(f"  PP 樞紐快取     ：{info['pp_cache']}")
            if info["current_week_auto"]:
                if info["current_week_base_date"] != info["current_date"]:
                    log(
                        f"  第一週起點      ：{info['current_week_range'][0]} "
                        f"（執行日 {info['current_date']} 為週五，提前使用下一週）"
                    )
                else:
                    log(
                        f"  第一週起點      ：{info['current_week_range'][0]} "
                        f"（依執行日 {info['current_date']}）"
                    )
            else:
                log(
                    f"  第一週起點      ：{info['current_week_range'][0]} "
                    "（手動指定 --dps-pp-current-week）"
                )
            log(
                f"  目前週          ：20{info['current_week_year'] % 100:02d} "
                f"WK{info['current_week']:02d}"
            )
            log(
                f"  DPS 保留週數    ：含本週共 {context.dps_pp_dps_weeks_ahead} 週"
                f"（{context.dps_pp_dps_weeks_source}）"
            )
            log(
                f"  DPS 使用範圍    ：到 WK{info['dps_cutoff_week']:02d} "
                f"({info['dps_cutoff_range'][0]} ~ {info['dps_cutoff_range'][1]})"
            )
            log(
                f"  PP 接續範圍     ：WK{info['pp_start_week']:02d} 起，"
                f"{len(info['pp_periods'])} 欄 → {', '.join(info['pp_periods'])}"
            )
            if info["dps_late_total"]:
                if info["dps_late_mode"] == "drop":
                    log(
                        f"  DPS 後段處理    ：{info['dps_cutoff_range'][1]} 之後合計 "
                        f"{info['dps_late_total']:,} pcs 已排除，由 PP 接續"
                    )
                else:
                    log(
                        f"  DPS 後段處理    ：{info['dps_cutoff_range'][1]} 之後合計 "
                        f"{info['dps_late_total']:,} pcs 已併入該日"
                    )
            else:
                log("  DPS 後段處理    ：無 cutoff 之後 DPS")
            if info["bom_found"]:
                log(f"  BOM             ：已讀取 BOM1（{info['bom_count']} 個料號）")
            else:
                log("  BOM             ：未找到 BOM1，BOM 欄留空")
            log(
                f"  輸出            ：{info['rows']} 列，"
                f"DPS 日期欄 {info['dps_dates']} 欄，"
                f"PP 期間欄 {len(info['pp_periods'])} 欄，"
                f"合計 {info['grand_total']:,} pcs"
            )
            log("  輸出格式        ：料號文字；數據與 total 為數值")
            log(f"  產出檔          ：{out_path}")
            if args.compare:
                warn(f"{title} 目前沒有單一人工整理表可逐格對帳，已略過 --compare。")
            return {
                "customer": context.name,
                "kind": "DPS+PP",
                "ok": True,
                "source": (info["dps_sources"], pp_path),
                "output": out_path,
                "dps_cutoff_end": info["dps_cutoff_range"][1],
            }
        except (SystemExit, Exception) as exc:  # noqa: BLE001 - 單一 PP 壞檔可嘗試下一個
            last_error = exc
            warn(f"DPS+PP 使用 PP 來源檔 {pp_path.name} 無法產出，已略過。原因：{_error_message(exc)}")
            if args.pp is not None:
                break

    if progress is not None and not output_step_done:
        progress.step(f"{title}: 略過報表")
    return {
        "customer": context.name,
        "kind": "DPS+PP",
        "ok": False,
        "error": _error_message(last_error) if last_error else "未知錯誤",
        "output": out_path,
        "stale_output": out_path.exists(),
    }


def run_ctb_report(
    args,
    context: RunContext,
    progress: Progress | None = None,
    *,
    dps_cutoff_end: dt.date | None = None,
) -> dict:
    out_path = context.out_dir / CTB_OUTPUT_NAME
    dps_pp_path = context.out_dir / DPS_PP_OUTPUT_NAME
    title = f"{context.label} CTB" if context.name else "CTB"
    context.out_dir.mkdir(parents=True, exist_ok=True)
    if progress is not None:
        progress.step(f"{title}: 檢查來源")

    try:
        if not dps_pp_path.is_file():
            raise SystemExit(f"找不到 {dps_pp_path}，請先成功產出 DPS+PP.xlsx")
        if dps_cutoff_end is None:
            raise SystemExit(
                "CTB 找不到本次 DPS+PP 的 cutoff 日期，"
                "為避免套用錯誤 Balance 規則已停止產出"
            )

        if context.name.casefold() == "raken":
            reference_path = find_raken_reference_workbook(context.input_dir)
            shortage_path = find_raken_shortage_workbook(context.input_dir)
            if progress is not None:
                progress.step(f"{title}: 產出報表")
            info = generate_raken_ctb(
                dps_pp_path=dps_pp_path,
                reference_path=reference_path,
                shortage_path=shortage_path,
                output_path=out_path,
                dps_cutoff_end=dps_cutoff_end,
                default_eta_lead_days=args.ctb_eta_default_lead_days,
                eta_lead_days_by_supplier_site=args.ctb_eta_lead_days_by_supplier_site,
            )
            log(f"\n--- {title} ---")
            log(f"  DPS+PP 來源     ：{info['dps_pp_source'].name}")
            log(
                f"  CTB 參考來源    ：{info['reference_source'].name}"
                "（讀取 demand / CTB / PO，並沿用 CTB 版型；不複製原始內容）"
            )
            log(f"  shortage 來源   ：{info['shortage_source'].name}（over shortage）")
            log("  B 欄料號        ：以 DPS+PP FG → demand PART_NO 為主，CTB sheet 僅補用量/搭配料展開")
            log("  料號排序        ：可計算料號依 input CTB 群組/來源列順序；缺 mapping 成品置於末端")
            log("  ERP 預留欄      ：A/C/E/I/J 僅保留欄名，資料列留白")
            log("  BOM 用量        ：使用光學 CTB CTB sheet 的 F 欄；demand 特別用量忽略")
            log("  Open PO         ：使用 PO sheet 實際子件數量，寫入 CTB H 欄；不建立外部輔助 sheet")
            log(f"  Balance 初始需求：只加總至 DPS cutoff {dps_cutoff_end}")
            shortage_count = str(info["over_shortage_rows"])
            if info.get("over_shortage_source_rows") != info["over_shortage_rows"]:
                shortage_count = f"{info['over_shortage_rows']}/{info['over_shortage_source_rows']}"
            open_po_count = str(info["open_po_rows"])
            if info.get("open_po_source_rows") != info["open_po_rows"]:
                open_po_count = f"{info['open_po_rows']}/{info['open_po_source_rows']}"
            log(
                f"  輸入資料        ：BOM {info['bom_rows']} 列，"
                f"over shortage {shortage_count} 列，"
                f"open po {open_po_count} 列"
            )
            log(
                f"  輸出            ：{info['parts']} 個料號，"
                f"Demand {info['demand_rows']} 列，"
                f"ETA {info['eta_rows']} 列，"
                f"other {info['other_rows']} 列，"
                f"Balance {info['balance_rows']} 列，"
                f"期間 {info['periods']} 欄"
            )
            if info.get("placeholder_parts"):
                log(
                    "  缺 mapping 顯示 ："
                    f"{info['placeholder_parts']} 個 DPS+PP 成品以空白計算列輸出，方便補資料"
                )
            log("  輸出內容        ：僅 CTB 工作表")
            log(f"  產出檔          ：{out_path}")
            for warning_message in info.get("warnings", []):
                warn(f"{title}：{warning_message}")
            if args.compare:
                warn(f"{title} 目前尚未支援 CTB 逐格對帳，已略過 --compare。")
            return {
                "customer": context.name,
                "kind": "CTB",
                "ok": True,
                "output": out_path,
            }

        bom_path = find_workbook_with_sheet(context.input_dir, "BOM1", f"{title} BOM1")
        open_po_path = find_workbook_with_sheet(context.input_dir, "open po", f"{title} open po")
        over_shortage_path = find_workbook_with_sheet(
            context.input_dir,
            "over shortage",
            f"{title} over shortage",
        )
        template_path = find_optional_workbook_with_sheets(
            context.input_dir,
            CTB_TEMPLATE_SHEET_NAMES,
        )
        if progress is not None:
            progress.step(f"{title}: 產出報表")
        info = generate_ctb(
            dps_pp_path=dps_pp_path,
            bom_path=bom_path,
            open_po_path=open_po_path,
            over_shortage_path=over_shortage_path,
            output_path=out_path,
            template_path=template_path,
            dps_cutoff_end=dps_cutoff_end,
            default_eta_lead_days=args.ctb_eta_default_lead_days,
            eta_lead_days_by_supplier_site=args.ctb_eta_lead_days_by_supplier_site,
        )
        log(f"\n--- {title} ---")
        log(f"  DPS+PP 來源     ：{info['dps_pp_source'].name}")
        log(f"  BOM1 來源       ：{info['bom_source'].name}")
        log(f"  open po 來源    ：{info['open_po_source'].name}")
        log(f"  over shortage 來源：{info['over_shortage_source'].name}")
        log(f"  Balance 初始需求：只加總至 DPS cutoff {dps_cutoff_end}")
        if info.get("template_source") is not None:
            log(
                f"  CTB 版型來源    ：{info['template_source'].name}"
                f"（工作表 {info.get('template_sheet', 'CTB')}；只沿用列結構與格式，數值重算）"
            )
        else:
            log("  CTB 版型來源    ：未提供，使用程式新建版面")
        log(
            f"  輸入資料        ：BOM {info['bom_rows']} 列，"
            f"over shortage {info['over_shortage_rows']} 列，"
            f"open po {info['open_po_rows']} 列"
        )
        other_rows_text = (
            f"，other {info['other_rows']} 列"
            if info.get("other_rows") is not None
            else ""
        )
        template_rows_text = (
            f"，版型 {info['template_rows']} 列"
            if info.get("template_rows") is not None
            else ""
        )
        log(
            f"  輸出            ：{info['parts']} 個料號，"
            f"Demand {info['demand_rows']} 列，"
            f"ETA {info['eta_rows']} 列"
            f"{other_rows_text}，"
            f"Balance {info['balance_rows']} 列，"
            f"期間 {info['periods']} 欄"
            f"{template_rows_text}"
        )
        log(
            f"  ETA 日期規則    ：逐筆模擬 Balance，第一個負值期間依 Supplier site 往前"
            f" {args.ctb_eta_default_lead_days} 個日曆日（未覆寫時）；"
            "若無負值則 fallback 至 open po Need By Date"
        )
        log(
            f"  ETA Supplier site 覆寫：{len(args.ctb_eta_lead_days_by_supplier_site)} 項，"
            f"設定檔 {args.ctb_eta_config_path}"
        )
        log(f"  產出檔          ：{out_path}")
        if args.compare:
            warn(f"{title} 目前尚未支援 CTB 逐格對帳，已略過 --compare。")
        return {
            "customer": context.name,
            "kind": "CTB",
            "ok": True,
            "output": out_path,
        }
    except (SystemExit, Exception) as exc:  # noqa: BLE001 - 單一客戶 CTB 失敗不阻斷其他報表
        if progress is not None:
            progress.step(f"{title}: 略過報表")
        warn(f"{title} 無法產出，已略過。原因：{_error_message(exc)}")
        return {
            "customer": context.name,
            "kind": "CTB",
            "ok": False,
            "error": _error_message(exc),
            "output": out_path,
            "stale_output": out_path.exists(),
        }


def run_reports(args) -> bool:
    args.input_dir.mkdir(parents=True, exist_ok=True)
    contexts = build_run_contexts(args)
    if contexts:
        run_log_paths = setup_run_log([context.out_dir for context in contexts])
    else:
        run_log_paths = setup_run_log(args.out_dir)

    log("=" * 72)
    log("Buyer Reports 產生器")
    log(f"輸入資料夾：{args.input_dir}")
    log(f"輸出資料夾：{args.out_dir}")
    log(f"執行記錄  ：{log_path_label(run_log_paths)}")
    config_status = "已讀取" if args.sheet_config_loaded else "未找到，使用內建預設"
    log(f"設定檔    ：{args.sheet_config_path}（{config_status}）")
    log(f"DPS sheet 關鍵字：{keyword_label(args.dps_sheet_keywords)}")
    log(f"PP sheet 關鍵字 ：{keyword_label(args.pp_sheet_keywords)}")
    log(f"DPS 料號欄別名 ：{keyword_label(args.dps_part_number_headers)}")
    log(f"PP 料號欄關鍵字：{keyword_label(args.pp_part_number_field_keywords)}")
    log("客戶設定        ：" + "、".join(
        f"{customer['name']}（DPS={customer['dps_mode']}，"
        f"PP={customer['pp_mode']}，"
        f"DPS+PP 的 DPS 保留週數=含本週共{customer['dps_pp_dps_weeks_ahead']}週，"
        f"週數來源={customer.get('dps_pp_dps_weeks_source', 'buyer_reports.ini / 內建預設')}，"
        f"DPS後段={late_dps_mode_label(customer['dps_pp_late_dps_mode'])}，"
        f"DPS零數量列={'略過' if customer['dps_drop_zero_total_rows'] else '保留'}，"
        f"DPS尾端空白日期={'略過' if customer['dps_trim_trailing_zero_dates'] else '保留'}）"
        for customer in args.customers
    ))
    log("=" * 72)
    for message in getattr(args, "context_warnings", []):
        warn(message)

    if not contexts:
        log("\n找不到任何可處理的 Excel。請把檔案放入 input/AVTC 或 input/RAKEN。")
        return False

    log("本次處理資料夾：")
    for context in contexts:
        log(f"  {context.label}：{context.input_dir} → {context.out_dir}")

    prepare_ctb_eta_settings(args, contexts)
    if getattr(args, "startup_cancelled", False):
        return True

    tasks = []
    for context in contexts:
        if not args.skip_dps:
            tasks.append((context, "DPS", run_dps_report))
        if not args.skip_pp:
            tasks.append((context, "PP", run_pp_report))
        if not args.skip_dps and not args.skip_pp and not args.skip_dps_pp:
            tasks.append((context, "DPS+PP", run_dps_pp_report))
        if _ctb_task_enabled_for_context(args, context):
            tasks.append((context, "CTB", run_ctb_report))

    results = []
    success_by_context_kind = {}
    dps_pp_cutoff_by_context: dict[str, dt.date] = {}
    with Progress(total=len(tasks) * 2) as progress:
        for context, _name, runner in tasks:
            if runner is run_ctb_report and not success_by_context_kind.get((context.name, "DPS+PP")):
                if progress is not None:
                    progress.step(f"{context.label} CTB: 檢查來源")
                    progress.step(f"{context.label} CTB: 略過報表")
                result = {
                    "customer": context.name,
                    "kind": "CTB",
                    "ok": False,
                    "error": "本次 DPS+PP 未成功產出，已略過 CTB，避免使用舊檔。",
                    "output": context.out_dir / CTB_OUTPUT_NAME,
                    "stale_output": (context.out_dir / CTB_OUTPUT_NAME).exists(),
                }
            elif runner is run_ctb_report:
                result = runner(
                    args,
                    context,
                    progress,
                    dps_cutoff_end=dps_pp_cutoff_by_context.get(context.name),
                )
            else:
                result = runner(args, context, progress)
            results.append(result)
            success_by_context_kind[(context.name, result["kind"])] = result["ok"]
            if (
                result["kind"] == "DPS+PP"
                and result["ok"]
                and result.get("dps_cutoff_end") is not None
            ):
                dps_pp_cutoff_by_context[context.name] = result["dps_cutoff_end"]

    log("\n--- 執行摘要 ---")
    for result in results:
        label = f"{result['customer']} {result['kind']}" if result.get("customer") else result["kind"]
        if result["ok"]:
            log(f"  {label}：成功 → {result['output']}")
        else:
            log(f"  {label}：失敗 / 已略過 → {result['error']}")
            if result.get("stale_output"):
                log(f"      注意：{result['output']} 已存在，可能是前次執行留下的舊檔。")

    success = any(result["ok"] for result in results)
    failed = [result for result in results if not result["ok"]]
    if not results:
        log("\n沒有啟用任何報表。")
        return True
    if failed:
        log("[警告] 部分報表未產出，請查看上方警告或對應 output 子資料夾的 log。")
    log("\n完成。")
    return success and not failed


def main() -> int:
    root = project_root()
    args = build_parser(root).parse_args()
    set_log_quiet(args.quiet)
    pause_after_run = is_frozen_app() and not args.no_pause

    try:
        prevent_temp_execution()
        args.out_dir.mkdir(parents=True, exist_ok=True)
        sheet_config = load_sheet_detection_config(root)
        args.sheet_config_path = sheet_config["path"]
        args.sheet_config_loaded = sheet_config["loaded"]
        args.dps_sheet_keywords = sheet_config["dps_sheet_keywords"]
        args.pp_sheet_keywords = sheet_config["pp_sheet_keywords"]
        args.dps_part_number_headers = sheet_config["dps_part_number_headers"]
        args.pp_part_number_field_keywords = sheet_config["pp_part_number_field_keywords"]
        args.customers = sheet_config["customers"]
        args.customer_scope = "all"
        if should_prompt_customer_scope(args):
            selected_scope = select_customer_scope()
            if selected_scope is None:
                print("使用者取消本次執行。", flush=True)
                return 0
            args.customer_scope = selected_scope
        apply_raken_dps_pp_weeks_selection(args)
        ok = run_reports(args)
        return 0 if ok else 1
    except SystemExit as exc:
        if exc.code not in (0, None):
            warn(str(exc))
        return exc.code if isinstance(exc.code, int) else 1
    except KeyboardInterrupt:
        warn("使用者中止執行。")
        return 130
    except Exception as exc:  # noqa: BLE001 - exe 模式需要把未知錯誤寫進 log
        warn(f"執行失敗：{exc}")
        write_traceback(traceback.format_exc())
        return 1
    finally:
        close_run_log()
        pause_for_windows_exe(pause_after_run)
