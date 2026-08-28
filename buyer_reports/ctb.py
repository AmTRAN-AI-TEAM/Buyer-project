"""CTB report generation."""

from __future__ import annotations

import datetime as dt
import re
from collections import defaultdict
from dataclasses import dataclass, field
from pathlib import Path
from typing import Mapping, Sequence

from openpyxl import Workbook, load_workbook
from openpyxl.formatting.rule import CellIsRule
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import column_index_from_string, get_column_letter

from .common import (
    autosize,
    clean_number,
    copy_cell_format,
    copy_column_layout,
    copy_row_layout,
    DEFAULT_CTB_ETA_LEAD_DAYS,
    normalize_label,
    normalize_part_number,
    numeric,
    set_filter_to_used_range,
    write_number_cell,
    write_text_cell,
)
from .dps_pp import week_label_for_date

CTB_OUTPUT_NAME = "CTB.xlsx"
CTB_SHEET = "CTB"
CTB_BOM_SHEET = "BOM1"
CTB_OVER_SHORTAGE_SHEET = "over shortage"
CTB_OPEN_PO_SHEET = "open po"
CTB_DPS_PP_SHEET = "DPS+PP"
CTB_ROW_TYPE_COL = 13
CTB_FIRST_PERIOD_COL = 14
ETA_LEAD_DAYS = DEFAULT_CTB_ETA_LEAD_DAYS
CTB_OVER_SHORTAGE_COL = 10
CTB_PO_REMAIN_COL = 11
CTB_TOTAL_COL = 12
NEGATIVE_FONT_COLOR = "FFFF0000"
WEEKDAY_LABELS = {
    "mon",
    "monday",
    "tue",
    "tues",
    "tuesday",
    "wed",
    "wednesday",
    "thu",
    "thur",
    "thurs",
    "thursday",
    "fri",
    "friday",
    "sat",
    "saturday",
    "sun",
    "sunday",
}
MONTH_PERIOD_KEY_RE = re.compile(r"^[a-z]{3}\d{2}(?:fcst)?$")
WEEK_PERIOD_KEY_RE = re.compile(r"^wk\d{1,2}$")


@dataclass(frozen=True)
class Period:
    label: str
    start: dt.date | None
    header1: str
    header2: str
    header3: str
    header4: str
    source_col: int


@dataclass
class BomRow:
    source_row: int
    category: str
    parent: str
    child: str
    use: float
    remark: str
    vendor: str
    demand: list[float]


@dataclass
class ShortageRecord:
    part: str
    description: str = ""
    buyer: str = ""
    planner: str = ""
    lead_time: str = ""
    po_remain: float = 0.0
    over_shortage: float = 0.0
    hld: float = 0.0
    bor_mm: float = 0.0
    overshortage1: float = 0.0


@dataclass
class OpenPoRecord:
    source_row: int
    key: str
    item: str
    supplier: str
    supplier_site: str
    quantity_due: float
    need_by_date: dt.date | None


@dataclass
class CtbPart:
    part: str
    category: str = ""
    model: str = ""
    vendor: str = ""
    demand: list[float] = field(default_factory=list)
    bom_rows: list[BomRow] = field(default_factory=list)
    shortage: ShortageRecord | None = None
    open_po: list[OpenPoRecord] = field(default_factory=list)


def _sheet_name(wb, target: str) -> str | None:
    target_norm = normalize_label(target)
    return next((name for name in wb.sheetnames if normalize_label(name) == target_norm), None)


def workbook_has_sheet(path: Path, sheet_name: str) -> bool:
    try:
        wb = load_workbook(path, read_only=True, data_only=True)
    except Exception:  # noqa: BLE001 - caller treats unreadable files as non-candidates
        return False
    try:
        return _sheet_name(wb, sheet_name) is not None
    finally:
        wb.close()


def workbook_has_any_ctb_sheet(path: Path) -> bool:
    try:
        wb = load_workbook(path, read_only=True, data_only=True)
    except Exception:  # noqa: BLE001
        return False
    try:
        return any(
            _sheet_name(wb, sheet_name) is not None
            for sheet_name in (CTB_BOM_SHEET, CTB_OPEN_PO_SHEET, CTB_OVER_SHORTAGE_SHEET)
        )
    finally:
        wb.close()


def has_ctb_input_candidates(input_dir: Path) -> bool:
    if not input_dir.is_dir():
        return False
    return any(
        path.is_file()
        and path.suffix.lower() == ".xlsx"
        and not path.name.startswith("~$")
        and workbook_has_any_ctb_sheet(path)
        for path in input_dir.glob("*.xlsx")
    )


def find_workbook_with_sheet(input_dir: Path, sheet_name: str, label: str) -> Path:
    candidates = [
        path
        for path in input_dir.glob("*.xlsx")
        if path.is_file()
        and not path.name.startswith("~$")
        and workbook_has_sheet(path, sheet_name)
    ]
    if not candidates:
        raise SystemExit(f"找不到 {label} 來源檔：{input_dir} 內需有工作表 {sheet_name!r}")
    candidates.sort(key=lambda path: path.stat().st_mtime, reverse=True)
    return candidates[0]


def find_optional_workbook_with_sheet(input_dir: Path, sheet_name: str) -> Path | None:
    candidates = [
        path
        for path in input_dir.glob("*.xlsx")
        if path.is_file()
        and not path.name.startswith("~$")
        and workbook_has_sheet(path, sheet_name)
    ]
    if not candidates:
        return None
    candidates.sort(key=lambda path: path.stat().st_mtime, reverse=True)
    return candidates[0]


def _parse_date(value) -> dt.date | None:
    if isinstance(value, dt.datetime):
        return value.date()
    if isinstance(value, dt.date):
        return value
    if value is None:
        return None
    text = str(value).strip()
    if not text:
        return None
    for fmt in ("%Y-%m-%d", "%Y/%m/%d", "%Y%m%d", "%Y-%m-%d %H:%M:%S"):
        try:
            return dt.datetime.strptime(text, fmt).date()
        except ValueError:
            continue
    return None


def _date_header_text(value) -> str:
    date = _parse_date(value)
    return date.isoformat() if date else ("" if value is None else str(value))


def _month_label(date: dt.date | None, fallback: str = "") -> str:
    return date.strftime("%b").upper() if date else fallback


def read_dps_pp(dps_pp_path: Path) -> tuple[list[Period], dict[str, list[float]]]:
    wb = load_workbook(dps_pp_path, data_only=True)
    try:
        sheet_name = _sheet_name(wb, CTB_DPS_PP_SHEET)
        if sheet_name is None:
            raise SystemExit(f"{dps_pp_path.name} 內找不到 {CTB_DPS_PP_SHEET} 工作表")
        ws = wb[sheet_name]
        total_col = None
        for cell in ws[4]:
            text = normalize_label(cell.value)
            if text == "total":
                total_col = cell.column
                break
        if total_col is None:
            raise SystemExit(f"{dps_pp_path.name} 的 DPS+PP 找不到 total 欄")

        periods: list[Period] = []
        for col in range(2, total_col):
            raw_date = ws.cell(4, col).value
            start = _parse_date(raw_date)
            row2 = "" if ws.cell(2, col).value is None else str(ws.cell(2, col).value)
            row3 = "" if ws.cell(3, col).value is None else str(ws.cell(3, col).value)
            row4 = _date_header_text(raw_date)
            label = row4 or row2 or get_column_letter(col)
            periods.append(
                Period(
                    label=label,
                    start=start,
                    header1="" if ws.cell(1, col).value is None else str(ws.cell(1, col).value),
                    header2=row2,
                    header3=row3,
                    header4=row4,
                    source_col=col,
                )
            )

        demand_by_parent: dict[str, list[float]] = {}
        for row in ws.iter_rows(min_row=5, values_only=True):
            parent = normalize_part_number(row[0])
            if not parent:
                continue
            values = [
                numeric(row[period.source_col - 1] if period.source_col - 1 < len(row) else None)
                for period in periods
            ]
            existing = demand_by_parent.setdefault(parent, [0.0] * len(periods))
            for idx, value in enumerate(values):
                existing[idx] += value
        return periods, demand_by_parent
    finally:
        wb.close()


def _find_header_row(ws, labels: Sequence[str], max_row: int = 20) -> int | None:
    targets = {normalize_label(label) for label in labels}
    for row_idx in range(1, min(ws.max_row, max_row) + 1):
        values = {
            normalize_label(ws.cell(row_idx, col).value)
            for col in range(1, ws.max_column + 1)
            if ws.cell(row_idx, col).value is not None
        }
        if targets.issubset(values):
            return row_idx
    return None


def _header_cols(ws, row_idx: int) -> dict[str, int]:
    cols = {}
    for col in range(1, ws.max_column + 1):
        value = ws.cell(row_idx, col).value
        if value is None:
            continue
        text = normalize_label(value)
        if text:
            cols.setdefault(text, col)
    return cols


def _find_col(headers: dict[str, int], aliases: Sequence[str]) -> int | None:
    normalized = [(normalize_label(alias), alias) for alias in aliases]
    for key, col in headers.items():
        if any(alias == key for alias, _raw in normalized):
            return col
    for key, col in headers.items():
        if any(alias in key for alias, _raw in normalized):
            return col
    return None


def _find_over_shortage_col(headers: dict[str, int]) -> int | None:
    return _find_col(headers, ["OVER SHORTAGE", "Over Shortage", "OVER_SHORTAGE"])


def read_bom_rows(
    bom_path: Path,
    periods: Sequence[Period],
    demand_by_parent: dict[str, list[float]],
) -> list[BomRow]:
    wb = load_workbook(bom_path, data_only=True)
    formula_wb = load_workbook(bom_path, data_only=False)
    try:
        sheet_name = _sheet_name(wb, CTB_BOM_SHEET)
        formula_sheet_name = _sheet_name(formula_wb, CTB_BOM_SHEET)
        if sheet_name is None or formula_sheet_name is None:
            raise SystemExit(f"{bom_path.name} 內找不到 {CTB_BOM_SHEET} 工作表")
        ws = wb[sheet_name]
        formula_ws = formula_wb[formula_sheet_name]
        header_row = _find_header_row(ws, ["Child P/N", "USE"], max_row=10)
        if header_row is None:
            raise SystemExit(f"{bom_path.name} 的 BOM1 找不到 Child P/N / USE 表頭")
        headers = _header_cols(ws, header_row)
        child_col = _find_col(headers, ["Child P/N", "Child PN", "Child"])
        use_col = _find_col(headers, ["USE"])
        vendor_col = _find_col(headers, ["vendor"])
        remark_col = _find_col(headers, ["Remark"])
        model_col = _find_col(headers, ["Model"])
        if child_col is None or use_col is None:
            raise SystemExit(f"{bom_path.name} 的 BOM1 缺少 Child P/N 或 USE 欄")
        parent_col = child_col - 1 if child_col > 1 else None

        rows = []
        for row_idx in range(header_row + 1, ws.max_row + 1):
            child = ws.cell(row_idx, child_col).value
            parent = ws.cell(row_idx, parent_col).value if parent_col else None
            if child is None or parent is None:
                continue
            child_text = normalize_part_number(child)
            parent_text = normalize_part_number(parent)
            if not child_text or not parent_text:
                continue
            use = numeric(ws.cell(row_idx, use_col).value)
            if use == 0:
                continue
            parent_demand = demand_by_parent.get(parent_text, [0.0] * len(periods))
            vendor_value = "" if vendor_col is None else ws.cell(row_idx, vendor_col).value
            vendor_formula = "" if vendor_col is None else formula_ws.cell(row_idx, vendor_col).value
            vendor_text = "" if _is_ctb_lookup_formula(vendor_formula) or vendor_value is None else str(vendor_value).strip()
            rows.append(
                BomRow(
                    source_row=row_idx,
                    category="" if model_col is None or ws.cell(row_idx, model_col).value is None else str(ws.cell(row_idx, model_col).value).strip(),
                    parent=parent_text,
                    child=child_text,
                    use=use,
                    remark="" if remark_col is None or ws.cell(row_idx, remark_col).value is None else str(ws.cell(row_idx, remark_col).value).strip(),
                    vendor=vendor_text,
                    demand=[value * use for value in parent_demand],
                )
            )
        return rows
    finally:
        wb.close()
        formula_wb.close()


def _merge_shortage_record(existing: ShortageRecord, incoming: ShortageRecord) -> None:
    if not existing.description:
        existing.description = incoming.description
    if not existing.buyer:
        existing.buyer = incoming.buyer
    if not existing.planner:
        existing.planner = incoming.planner
    if not existing.lead_time:
        existing.lead_time = incoming.lead_time
    existing.po_remain += incoming.po_remain
    existing.over_shortage += incoming.over_shortage
    existing.hld += incoming.hld
    existing.bor_mm += incoming.bor_mm
    existing.overshortage1 += incoming.overshortage1


def read_over_shortage(over_shortage_path: Path) -> dict[str, ShortageRecord]:
    wb = load_workbook(over_shortage_path, data_only=True)
    try:
        sheet_name = _sheet_name(wb, CTB_OVER_SHORTAGE_SHEET)
        if sheet_name is None:
            raise SystemExit(f"{over_shortage_path.name} 內找不到 {CTB_OVER_SHORTAGE_SHEET} 工作表")
        ws = wb[sheet_name]
        header_row = _find_header_row(ws, ["Part No"], max_row=10)
        if header_row is None:
            raise SystemExit(f"{over_shortage_path.name} 的 {CTB_OVER_SHORTAGE_SHEET} 找不到 Part No 表頭")
        headers = _header_cols(ws, header_row)
        part_col = _find_col(headers, ["Part No"])
        if part_col is None:
            raise SystemExit(f"{over_shortage_path.name} 的 {CTB_OVER_SHORTAGE_SHEET} 缺少 Part No 欄")

        description_col = _find_col(headers, ["Description", "DESCRIPTION"])
        planner_col = _find_col(headers, ["Planner", "PLANNER"])
        buyer_col = _find_col(headers, ["Buyer"])
        lead_time_col = _find_col(headers, ["LT", "Lead Time"])
        po_remain_col = _find_col(headers, ["PO_REMAIN", "Po Remain"])
        over_col = _find_over_shortage_col(headers)
        if over_col is None:
            raise SystemExit(
                f"{over_shortage_path.name} 的 {CTB_OVER_SHORTAGE_SHEET} 缺少 OVER SHORTAGE 欄"
            )
        hld_col = _find_col(headers, ["HLD"])
        bor_col = _find_col(headers, ["BOR MM"])
        overshortage1_col = _find_col(headers, ["Overshortage1", "overshortage"])

        records = {}
        for row_idx in range(header_row + 1, ws.max_row + 1):
            part = ws.cell(row_idx, part_col).value
            part_text = normalize_part_number(part)
            if not part_text:
                continue
            record = ShortageRecord(
                part=part_text,
                description="" if description_col is None or ws.cell(row_idx, description_col).value is None else str(ws.cell(row_idx, description_col).value).strip(),
                buyer="" if buyer_col is None or ws.cell(row_idx, buyer_col).value is None else str(ws.cell(row_idx, buyer_col).value).strip(),
                planner="" if planner_col is None or ws.cell(row_idx, planner_col).value is None else str(ws.cell(row_idx, planner_col).value).strip(),
                lead_time="" if lead_time_col is None or ws.cell(row_idx, lead_time_col).value is None else str(ws.cell(row_idx, lead_time_col).value).strip(),
                po_remain=numeric(ws.cell(row_idx, po_remain_col).value if po_remain_col else None),
                over_shortage=numeric(ws.cell(row_idx, over_col).value if over_col else None),
                hld=numeric(ws.cell(row_idx, hld_col).value if hld_col else None),
                bor_mm=numeric(ws.cell(row_idx, bor_col).value if bor_col else None),
                overshortage1=numeric(ws.cell(row_idx, overshortage1_col).value if overshortage1_col else None),
            )
            existing = records.get(part_text)
            if existing is None:
                records[part_text] = record
            else:
                _merge_shortage_record(existing, record)
        return records
    finally:
        wb.close()


def read_open_po(open_po_path: Path) -> list[OpenPoRecord]:
    wb = load_workbook(open_po_path, data_only=True)
    try:
        sheet_name = _sheet_name(wb, CTB_OPEN_PO_SHEET)
        if sheet_name is None:
            raise SystemExit(f"{open_po_path.name} 內找不到 {CTB_OPEN_PO_SHEET} 工作表")
        ws = wb[sheet_name]
        header_row = _find_header_row(ws, ["Item", "Quantity Due"], max_row=10)
        if header_row is None:
            raise SystemExit(f"{open_po_path.name} 的 open po 找不到 Item / Quantity Due 表頭")
        headers = _header_cols(ws, header_row)
        key_col = _find_col(headers, ["料号+厂商", "料號+廠商", "part supplier key"])
        item_col = _find_col(headers, ["Item"])
        quantity_col = _find_col(headers, ["Quantity Due"])
        supplier_col = _find_col(headers, ["Supplier"])
        supplier_site_col = _find_col(headers, ["Supplier Site"])
        need_by_col = _find_col(headers, ["Need By Date"])
        if item_col is None or quantity_col is None:
            raise SystemExit(f"{open_po_path.name} 的 open po 缺少 Item 或 Quantity Due 欄")

        records = []
        for row_idx in range(header_row + 1, ws.max_row + 1):
            item = ws.cell(row_idx, item_col).value
            item_text = normalize_part_number(item)
            if not item_text:
                continue
            quantity_due = numeric(ws.cell(row_idx, quantity_col).value)
            if quantity_due == 0:
                continue
            supplier_site = (
                ""
                if supplier_site_col is None or ws.cell(row_idx, supplier_site_col).value is None
                else str(ws.cell(row_idx, supplier_site_col).value).strip()
            )
            key = ""
            if key_col is not None and ws.cell(row_idx, key_col).value is not None:
                key = normalize_part_number(ws.cell(row_idx, key_col).value)
            if not key:
                key = f"{item_text}{supplier_site}"
            records.append(
                OpenPoRecord(
                    source_row=row_idx,
                    key=key,
                    item=item_text,
                    supplier="" if supplier_col is None or ws.cell(row_idx, supplier_col).value is None else str(ws.cell(row_idx, supplier_col).value).strip(),
                    supplier_site=supplier_site,
                    quantity_due=quantity_due,
                    need_by_date=_parse_date(ws.cell(row_idx, need_by_col).value if need_by_col else None),
                )
            )
        return records
    finally:
        wb.close()


def period_index_for_date(periods: Sequence[Period], date: dt.date | None) -> int | None:
    if not periods:
        return None
    dated = [(idx, period.start) for idx, period in enumerate(periods) if period.start is not None]
    if not dated:
        return None
    if date is None:
        return dated[-1][0]
    for idx, start in dated:
        if start >= date:
            return idx
    return dated[-1][0]


def _is_ctb_lookup_formula(value) -> bool:
    return isinstance(value, str) and value.startswith("=") and "CTB!" in value.upper()


def _format_use(value: float) -> str:
    cleaned = clean_number(value)
    return str(cleaned)


def _model_from_bom_rows(rows: Sequence[BomRow]) -> str:
    values = []
    seen = set()
    for row in rows:
        model = row.category.strip()
        if not model or model.casefold() == "#n/a":
            continue
        value = f"{model}*{_format_use(row.use)}"
        if value in seen:
            continue
        seen.add(value)
        values.append(value)
    return "/".join(values)


def _vendor_from_bom_rows(rows: Sequence[BomRow]) -> str:
    values = []
    seen = set()
    for row in rows:
        vendor = row.vendor.strip()
        if not vendor or vendor.casefold() == "#n/a":
            continue
        if vendor in seen:
            continue
        seen.add(vendor)
        values.append(vendor)
    return "+".join(values)


def _eta_key(part_no: str, supplier_site: str) -> str:
    return f"{part_no}{supplier_site}" if part_no or supplier_site else ""


def _supplier_site_key(value) -> str:
    return "" if value is None else str(value).strip().casefold()


def build_part_map(
    periods: Sequence[Period],
    bom_rows: Sequence[BomRow],
    shortage: dict[str, ShortageRecord],
    open_po: Sequence[OpenPoRecord],
) -> tuple[dict[str, CtbPart], list[str]]:
    parts: dict[str, CtbPart] = {}
    order = []

    def get_part(part: str) -> CtbPart:
        if part not in parts:
            parts[part] = CtbPart(part=part, demand=[0.0] * len(periods))
            order.append(part)
        return parts[part]

    for row in bom_rows:
        part = get_part(row.child)
        part.bom_rows.append(row)
        for idx, value in enumerate(row.demand):
            part.demand[idx] += value

    for record in open_po:
        part = get_part(record.item)
        part.open_po.append(record)

    for part_no, record in shortage.items():
        part = get_part(part_no)
        part.shortage = record
        if not part.category:
            part.category = record.description

    for part in parts.values():
        if part.bom_rows:
            part.model = _model_from_bom_rows(part.bom_rows)
            part.vendor = _vendor_from_bom_rows(part.bom_rows)

    return parts, order


def filter_ctb_parts(parts: dict[str, CtbPart], order: Sequence[str]) -> list[CtbPart]:
    return [
        parts[part]
        for part in order
        if any(parts[part].demand)
        or parts[part].open_po
        or (
            parts[part].shortage is not None
            and (
                parts[part].shortage.over_shortage
                or parts[part].shortage.po_remain
                or parts[part].shortage.hld
                or parts[part].shortage.bor_mm
            )
        )
    ]


def build_parts(
    periods: Sequence[Period],
    bom_rows: Sequence[BomRow],
    shortage: dict[str, ShortageRecord],
    open_po: Sequence[OpenPoRecord],
) -> list[CtbPart]:
    parts, order = build_part_map(periods, bom_rows, shortage, open_po)
    return filter_ctb_parts(parts, order)


def _project_balance_values(
    periods: Sequence[Period],
    demand: Sequence[float],
    eta: Sequence[float],
    other: Sequence[float],
    over_shortage: float,
    initial_sum_cols: tuple[int, int] | None = None,
    period_start_col: int = CTB_FIRST_PERIOD_COL,
) -> list[float]:
    """Project Balance with the same period relationships written to Excel."""
    if not periods:
        return []

    def value_at(values: Sequence[float], idx: int) -> float:
        return numeric(values[idx]) if idx < len(values) else 0.0

    if initial_sum_cols is None:
        initial_indices = range(1, len(periods))
    else:
        initial_indices = (
            idx
            for idx in range(len(periods))
            if initial_sum_cols[0] <= period_start_col + idx <= initial_sum_cols[1]
        )
    balances = [numeric(over_shortage)]
    balances[0] += sum(value_at(demand, idx) for idx in initial_indices)
    for idx in range(1, len(periods)):
        balances.append(
            balances[idx - 1]
            + value_at(eta, idx - 1)
            - value_at(demand, idx)
            - value_at(other, idx)
        )
    return balances


def eta_schedule_for_records(
    periods: Sequence[Period],
    records: Sequence[OpenPoRecord],
    *,
    demand: Sequence[float] | None = None,
    over_shortage: float = 0.0,
    other: Sequence[float] | None = None,
    initial_sum_cols: tuple[int, int] | None = None,
    period_start_col: int = CTB_FIRST_PERIOD_COL,
    default_lead_days: int = ETA_LEAD_DAYS,
    lead_days_by_supplier_site: Mapping[str, int] | None = None,
) -> dict[str, list[float]]:
    schedules: dict[str, list[float]] = defaultdict(lambda: [0.0] * len(periods))
    if not periods:
        return schedules

    demand_values = list(demand) if demand is not None else [0.0] * len(periods)
    other_values = list(other) if other is not None else [0.0] * len(periods)
    scheduled_eta = [0.0] * len(periods)

    for record in records:
        balance_values = _project_balance_values(
            periods,
            demand_values,
            scheduled_eta,
            other_values,
            over_shortage,
            initial_sum_cols,
            period_start_col,
        )
        negative_idx = next(
            (idx for idx, balance in enumerate(balance_values) if balance < 0),
            None,
        )
        if negative_idx is None:
            idx = period_index_for_date(periods, record.need_by_date)
        else:
            shortage_date = periods[negative_idx].start
            lead_days = default_lead_days
            if lead_days_by_supplier_site:
                lead_days = lead_days_by_supplier_site.get(
                    _supplier_site_key(record.supplier_site),
                    default_lead_days,
                )
            target_date = (
                shortage_date - dt.timedelta(days=lead_days)
                if shortage_date is not None
                else record.need_by_date
            )
            idx = period_index_for_date(periods, target_date)
        if idx is None:
            continue
        schedules[record.key][idx] += record.quantity_due
        scheduled_eta[idx] += record.quantity_due
    return schedules


def write_auxiliary_sheets(
    wb: Workbook,
    periods: Sequence[Period],
    bom_rows: Sequence[BomRow],
    shortage_records: dict[str, ShortageRecord],
    open_po_records: Sequence[OpenPoRecord],
) -> None:
    ws = wb.create_sheet(CTB_BOM_SHEET)
    headers = ["Model", "Parent P/N", "Child P/N", "USE", "Remark", "vendor"]
    for col, value in enumerate(headers, start=1):
        write_text_cell(ws.cell(1, col), value)
    for offset, period in enumerate(periods, start=len(headers) + 1):
        write_text_cell(ws.cell(1, offset), period.header4 or period.label)
    total_col = len(headers) + len(periods) + 1
    write_text_cell(ws.cell(1, total_col), "total")
    for row_idx, row in enumerate(bom_rows, start=2):
        values = [row.category, row.parent, row.child, row.use, row.remark, row.vendor]
        for col, value in enumerate(values, start=1):
            if col == 4:
                write_number_cell(ws.cell(row_idx, col), value)
            else:
                write_text_cell(ws.cell(row_idx, col), value)
        for offset, value in enumerate(row.demand, start=len(headers) + 1):
            write_number_cell(ws.cell(row_idx, offset), clean_number(value))
        write_number_cell(ws.cell(row_idx, total_col), clean_number(sum(row.demand)))
    autosize(ws, maximum=22)

    ws = wb.create_sheet(CTB_OVER_SHORTAGE_SHEET)
    headers = [
        "Part No", "Description", "Buyer", "Planner", "Lead Time",
        "Po Remain", "Over Shortage", "HLD", "BOR MM", "Overshortage1",
    ]
    for col, value in enumerate(headers, start=1):
        write_text_cell(ws.cell(1, col), value)
    for row_idx, record in enumerate(shortage_records.values(), start=2):
        values = [
            record.part, record.description, record.buyer, record.planner, record.lead_time,
            record.po_remain, record.over_shortage, record.hld, record.bor_mm,
            record.overshortage1,
        ]
        for col, value in enumerate(values, start=1):
            if col in {6, 7, 8, 9, 10}:
                write_number_cell(ws.cell(row_idx, col), clean_number(numeric(value)))
            else:
                write_text_cell(ws.cell(row_idx, col), value)
    if shortage_records:
        _add_negative_font_rule(ws, f"G2:G{len(shortage_records) + 1}")
    autosize(ws, maximum=24)

    ws = wb.create_sheet(CTB_OPEN_PO_SHEET)
    headers = ["key", "Item", "Supplier", "Supplier Site", "Quantity Due", "Need By Date"]
    for col, value in enumerate(headers, start=1):
        write_text_cell(ws.cell(1, col), value)
    for row_idx, record in enumerate(open_po_records, start=2):
        values = [
            record.key,
            record.item,
            record.supplier,
            record.supplier_site,
            record.quantity_due,
            record.need_by_date.isoformat() if record.need_by_date else "",
        ]
        for col, value in enumerate(values, start=1):
            if col == 5:
                write_number_cell(ws.cell(row_idx, col), clean_number(numeric(value)))
            else:
                write_text_cell(ws.cell(row_idx, col), value)
    autosize(ws, maximum=28)


def copy_dps_pp_sheet(wb: Workbook, dps_pp_path: Path) -> None:
    source_wb = load_workbook(dps_pp_path, data_only=True)
    try:
        source_name = _sheet_name(source_wb, CTB_DPS_PP_SHEET)
        if source_name is None:
            return
        source_ws = source_wb[source_name]
        ws = wb.create_sheet(CTB_DPS_PP_SHEET)
        for row in source_ws.iter_rows():
            for cell in row:
                target = ws.cell(cell.row, cell.column)
                if isinstance(cell.value, (int, float)) and not isinstance(cell.value, bool):
                    write_number_cell(target, cell.value)
                else:
                    write_text_cell(target, cell.value)
        autosize(ws, maximum=18)
    finally:
        source_wb.close()


def _write_formula_cell(cell, formula: str) -> None:
    cell.value = formula
    cell.number_format = "General"


def _cell_ref(row_idx: int, col_idx: int, *, absolute_col: bool = False) -> str:
    col = get_column_letter(col_idx)
    return f"{'$' if absolute_col else ''}{col}{row_idx}"


def _sheet_ref(sheet_name: str) -> str:
    return "'" + sheet_name.replace("'", "''") + "'"


def _row_range_ref(row_idx: int, start_col: int, end_col: int) -> str:
    if start_col > end_col:
        return ""
    start_ref = _cell_ref(row_idx, start_col)
    end_ref = _cell_ref(row_idx, end_col)
    return start_ref if start_col == end_col else f"{start_ref}:{end_ref}"


def _period_row_range_ref(
    row_idx: int,
    period_columns: Sequence[tuple[int, Period]],
    col_range: tuple[int, int] | None = None,
) -> str:
    cols = [
        col
        for col, _period in period_columns
        if col_range is None or col_range[0] <= col <= col_range[1]
    ]
    if not cols:
        return ""
    return _row_range_ref(row_idx, cols[0], cols[-1])


def _sum_expression(ref: str) -> str:
    return f"SUM({ref})" if ref else "0"


def _sum_period_formula(
    row_idx: int,
    period_columns: Sequence[tuple[int, Period]],
    col_range: tuple[int, int] | None = None,
) -> str:
    return f"={_sum_expression(_period_row_range_ref(row_idx, period_columns, col_range))}"


def _sum_rows_in_col_expression(row_indices: Sequence[int], col_idx: int) -> str:
    if not row_indices:
        return "0"
    sorted_rows = sorted(row_indices)
    if len(sorted_rows) == 1:
        return _cell_ref(sorted_rows[0], col_idx)
    if sorted_rows == list(range(sorted_rows[0], sorted_rows[-1] + 1)):
        return f"SUM({_cell_ref(sorted_rows[0], col_idx)}:{_cell_ref(sorted_rows[-1], col_idx)})"
    refs = ",".join(_cell_ref(row_idx, col_idx) for row_idx in sorted_rows)
    return f"SUM({refs})"


def _over_shortage_lookup_formula(row_idx: int) -> str:
    sheet = _sheet_ref(CTB_OVER_SHORTAGE_SHEET)
    return f"=SUMIF({sheet}!$A:$A,{_cell_ref(row_idx, 3, absolute_col=True)},{sheet}!$G:$G)"


def _open_po_lookup_formula(row_idx: int) -> str:
    sheet = _sheet_ref(CTB_OPEN_PO_SHEET)
    return f"=SUMIF({sheet}!$A:$A,{_cell_ref(row_idx, 4, absolute_col=True)},{sheet}!$E:$E)"


def _first_balance_formula(
    balance_row: int,
    demand_row: int,
    period_columns: Sequence[tuple[int, Period]],
    initial_sum_cols: tuple[int, int] | None = None,
) -> str:
    if initial_sum_cols is None:
        demand_ref = _period_row_range_ref(demand_row, period_columns[1:])
    else:
        demand_ref = _period_row_range_ref(demand_row, period_columns, initial_sum_cols)
    demand_expr = _sum_expression(demand_ref)
    return f"={_cell_ref(balance_row, CTB_OVER_SHORTAGE_COL)}+{demand_expr}"


def _next_balance_formula(
    balance_row: int,
    demand_row: int,
    eta_rows: Sequence[int],
    other_rows: Sequence[int],
    prev_col: int,
    current_col: int,
) -> str:
    eta_expr = _sum_rows_in_col_expression(eta_rows, prev_col)
    other_expr = _sum_rows_in_col_expression(other_rows, current_col)
    return (
        f"={_cell_ref(balance_row, prev_col)}"
        f"+{eta_expr}"
        f"-{_cell_ref(demand_row, current_col)}"
        f"-{other_expr}"
    )


def _write_balance_formulas(
    ws,
    balance_row: int,
    demand_row: int,
    eta_rows: Sequence[int],
    other_rows: Sequence[int],
    period_columns: Sequence[tuple[int, Period]],
    initial_sum_cols: tuple[int, int] | None = None,
) -> None:
    for idx, (col, _period) in enumerate(period_columns):
        if idx == 0:
            formula = _first_balance_formula(balance_row, demand_row, period_columns, initial_sum_cols)
        else:
            formula = _next_balance_formula(
                balance_row,
                demand_row,
                eta_rows,
                other_rows,
                period_columns[idx - 1][0],
                col,
            )
        _write_formula_cell(ws.cell(balance_row, col), formula)


def _po_remain_formula(
    balance_row: int,
    period_columns: Sequence[tuple[int, Period]],
    po_remain_period_idx: int | None,
) -> str | None:
    if po_remain_period_idx is None or po_remain_period_idx >= len(period_columns):
        return None
    balance_col = period_columns[po_remain_period_idx][0]
    return (
        f"={_cell_ref(balance_row, CTB_OVER_SHORTAGE_COL)}"
        f"-{_cell_ref(balance_row, balance_col)}"
    )


def _negative_font_rule() -> CellIsRule:
    return CellIsRule(operator="lessThan", formula=["0"], font=Font(color=NEGATIVE_FONT_COLOR))


def _add_negative_font_rule(ws, cell_range: str) -> None:
    if cell_range:
        ws.conditional_formatting.add(cell_range, _negative_font_rule())


def _apply_ctb_negative_formatting(
    ws,
    balance_rows: Sequence[int],
    period_columns: Sequence[tuple[int, Period]],
) -> None:
    for row_idx in balance_rows:
        _add_negative_font_rule(ws, f"{_cell_ref(row_idx, CTB_OVER_SHORTAGE_COL)}")
        if period_columns:
            _add_negative_font_rule(
                ws,
                _row_range_ref(row_idx, period_columns[0][0], period_columns[-1][0]),
            )


def _enable_formula_recalculation(wb: Workbook) -> None:
    calc = getattr(wb, "calculation", None)
    if calc is None:
        return
    calc.calcMode = "auto"
    calc.fullCalcOnLoad = True
    calc.forceFullCalc = True


def _write_header(ws, periods: Sequence[Period]) -> int:
    static_headers = [
        "Category", "model", "Part No", "", "Code", "",
        "Vendor", "", "Rate", "OVER SHORTAGE", "PO Remain", "total", "ETA目标",
    ]
    for col, value in enumerate(static_headers, start=1):
        write_text_cell(ws.cell(4, col), value)
    start_col = len(static_headers) + 1
    for offset, period in enumerate(periods, start=start_col):
        forecast_label = _forecast_period_label(period)
        if forecast_label:
            write_text_cell(ws.cell(1, offset), _month_label(period.start, period.header1))
            write_text_cell(ws.cell(2, offset), forecast_label)
            _clear_cell(ws.cell(3, offset))
            header4 = period.start.isoformat() if period.start else period.header4 or period.label
            write_text_cell(ws.cell(4, offset), header4)
        elif period.start is not None:
            write_text_cell(ws.cell(1, offset), _month_label(period.start, period.header1))
            write_text_cell(ws.cell(2, offset), week_label_for_date(period.start))
            write_text_cell(ws.cell(3, offset), period.start.strftime("%a"))
            write_text_cell(ws.cell(4, offset), period.start.isoformat())
        else:
            write_text_cell(ws.cell(1, offset), period.header1)
            write_text_cell(ws.cell(2, offset), period.header2)
            write_text_cell(ws.cell(3, offset), period.header3)
            write_text_cell(ws.cell(4, offset), period.header4 or period.label)
    return start_col


def _write_generated_summary_headers(ws, first_col: int) -> int:
    write_text_cell(ws.cell(4, first_col), "demand")
    write_text_cell(ws.cell(4, first_col + 1), "PO")
    _clear_cell(ws.cell(4, first_col + 2))
    return first_col + 2


def _write_optional_text_cell(cell, value) -> None:
    text = _cell_text(value)
    if text:
        write_text_cell(cell, text)
    else:
        _clear_cell(cell)


def _write_static_cells(ws, row_idx: int, part: CtbPart, row_type: str, key: str = "", po: OpenPoRecord | None = None) -> None:
    shortage = part.shortage
    supplier_site = po.supplier_site if po else ""
    key_value = _eta_key(part.part, supplier_site) if row_type.casefold() == "eta" and po is not None else key

    _clear_cell(ws.cell(row_idx, 1))
    _write_optional_text_cell(ws.cell(row_idx, 2), part.model)
    write_text_cell(ws.cell(row_idx, 3), part.part)
    if row_type.casefold() == "eta" and po is not None:
        _write_formula_cell(ws.cell(row_idx, 4), f"={_cell_ref(row_idx, 3)}&{_cell_ref(row_idx, 5)}")
    else:
        _write_optional_text_cell(ws.cell(row_idx, 4), key_value)
    _write_optional_text_cell(ws.cell(row_idx, 5), supplier_site)
    _write_optional_text_cell(ws.cell(row_idx, 6), part.vendor)
    _clear_cell(ws.cell(row_idx, 7))
    _clear_cell(ws.cell(row_idx, 8))
    _clear_cell(ws.cell(row_idx, 9))
    if shortage and row_type.lower().startswith("balance"):
        _write_formula_cell(ws.cell(row_idx, CTB_OVER_SHORTAGE_COL), _over_shortage_lookup_formula(row_idx))
    else:
        _clear_cell(ws.cell(row_idx, CTB_OVER_SHORTAGE_COL))
    _clear_cell(ws.cell(row_idx, CTB_PO_REMAIN_COL))
    _clear_cell(ws.cell(row_idx, CTB_TOTAL_COL))
    write_text_cell(ws.cell(row_idx, 13), row_type)


def write_ctb_sheet(
    wb: Workbook,
    periods: Sequence[Period],
    parts: Sequence[CtbPart],
    *,
    default_eta_lead_days: int = ETA_LEAD_DAYS,
    eta_lead_days_by_supplier_site: Mapping[str, int] | None = None,
) -> dict:
    ws = wb.active
    ws.title = CTB_SHEET
    start_col = _write_header(ws, periods)
    last_period_col = start_col + len(periods) - 1
    end_col = _write_generated_summary_headers(ws, last_period_col + 1)
    summary_demand_col = last_period_col + 1
    summary_po_col = last_period_col + 2
    period_columns = [(start_col + idx, period) for idx, period in enumerate(periods)]
    po_remain_period_idx = _last_by_day_period_index(period_columns)

    fills = {
        "Demand": PatternFill("solid", fgColor="EAF4FF"),
        "ETA": PatternFill("solid", fgColor="FFF4D6"),
        "Balance": PatternFill("solid", fgColor="EAF7EA"),
    }
    row_idx = 5
    demand_rows = eta_rows = other_rows = balance_rows = 0
    balance_row_indices: list[int] = []
    for part in parts:
        demand_row = row_idx
        _write_static_cells(ws, row_idx, part, "Demand")
        for offset, value in enumerate(part.demand, start=start_col):
            write_number_cell(ws.cell(row_idx, offset), clean_number(value))
        _write_formula_cell(ws.cell(row_idx, summary_demand_col), _sum_period_formula(row_idx, period_columns))
        demand_rows += 1
        row_idx += 1

        schedules = eta_schedule_for_records(
            periods,
            part.open_po,
            demand=part.demand,
            over_shortage=part.shortage.over_shortage if part.shortage else 0.0,
            period_start_col=start_col,
            default_lead_days=default_eta_lead_days,
            lead_days_by_supplier_site=eta_lead_days_by_supplier_site,
        )
        po_by_key: dict[str, OpenPoRecord] = {}
        for record in part.open_po:
            po_by_key.setdefault(record.key, record)

        eta_row_indices: list[int] = []
        eta_items = list(schedules.items())
        if not eta_items:
            eta_items = [("", [0.0] * len(periods))]
        for key, schedule in eta_items:
            eta_row = row_idx
            eta_row_indices.append(eta_row)
            record = po_by_key.get(key)
            _write_static_cells(ws, row_idx, part, "ETA", key=key, po=record)
            _write_formula_cell(ws.cell(row_idx, CTB_PO_REMAIN_COL), _open_po_lookup_formula(row_idx))
            _write_formula_cell(ws.cell(row_idx, CTB_TOTAL_COL), _sum_period_formula(row_idx, period_columns))
            _write_formula_cell(ws.cell(row_idx, summary_po_col), _sum_period_formula(row_idx, period_columns))
            for idx, value in enumerate(schedule):
                write_number_cell(ws.cell(row_idx, start_col + idx), clean_number(value))
            eta_rows += 1
            row_idx += 1

        other_row = row_idx
        _write_static_cells(ws, row_idx, part, "other")
        for offset in range(start_col, last_period_col + 1):
            write_number_cell(ws.cell(row_idx, offset), 0)
        other_rows += 1
        row_idx += 1

        balance_row = row_idx
        _write_static_cells(ws, row_idx, part, "Balance")
        if part.shortage:
            po_remain_formula = _po_remain_formula(balance_row, period_columns, po_remain_period_idx)
            if po_remain_formula is not None:
                _write_formula_cell(ws.cell(row_idx, CTB_PO_REMAIN_COL), po_remain_formula)
            else:
                write_number_cell(ws.cell(row_idx, CTB_PO_REMAIN_COL), clean_number(part.shortage.po_remain))
        _write_balance_formulas(
            ws,
            balance_row,
            demand_row,
            eta_row_indices,
            [other_row],
            period_columns,
        )
        balance_row_indices.append(balance_row)
        balance_rows += 1
        row_idx += 1

    for row in ws.iter_rows(min_row=1, max_row=4, max_col=end_col):
        for cell in row:
            if cell.value is not None:
                cell.font = Font(bold=True)
                cell.alignment = Alignment(horizontal="center")
    for row in ws.iter_rows(min_row=5, max_row=row_idx - 1, max_col=end_col):
        row_type = str(ws.cell(row[0].row, 13).value)
        fill = fills.get("Balance" if row_type.startswith("Balance") else row_type)
        if fill:
            for cell in row:
                cell.fill = fill

    _apply_ctb_negative_formatting(ws, balance_row_indices, period_columns)
    ws.freeze_panes = None
    set_filter_to_used_range(ws, end_col, row_idx - 1)
    autosize(ws, maximum=20)
    ws.column_dimensions["C"].width = 18
    ws.column_dimensions["D"].width = 26
    return {
        "parts": len(parts),
        "demand_rows": demand_rows,
        "eta_rows": eta_rows,
        "other_rows": other_rows,
        "balance_rows": balance_rows,
        "periods": len(periods),
    }


def _copy_template_sheet(source_ws, target_ws) -> None:
    target_ws.title = CTB_SHEET
    for col in range(1, source_ws.max_column + 1):
        copy_column_layout(source_ws, target_ws, col, col)
    for row in range(1, source_ws.max_row + 1):
        copy_row_layout(source_ws, target_ws, row, row)
        last_col = source_ws.max_column if row <= 4 else CTB_ROW_TYPE_COL
        for col in range(1, last_col + 1):
            source = source_ws.cell(row, col)
            target = target_ws.cell(row, col)
            target.value = source.value
            copy_cell_format(source, target)

    for merged_range in source_ws.merged_cells.ranges:
        target_ws.merge_cells(str(merged_range))
    target_ws.freeze_panes = None
    target_ws.auto_filter.ref = source_ws.auto_filter.ref
    target_ws.sheet_view.showGridLines = source_ws.sheet_view.showGridLines


def _template_period_columns(ws) -> list[tuple[int, Period]]:
    periods: list[tuple[int, Period]] = []
    for col in range(CTB_FIRST_PERIOD_COL, ws.max_column + 1):
        start = _parse_date(ws.cell(4, col).value)
        if start is None:
            if periods:
                break
            continue
        periods.append(
            (
                col,
                Period(
                    label=start.isoformat(),
                    start=start,
                    header1="" if ws.cell(1, col).value is None else str(ws.cell(1, col).value),
                    header2="" if ws.cell(2, col).value is None else str(ws.cell(2, col).value),
                    header3="" if ws.cell(3, col).value is None else str(ws.cell(3, col).value),
                    header4=start.isoformat(),
                    source_col=col,
                ),
            )
        )
    return periods


def _template_summary_columns(ws, last_period_col: int) -> dict[str, int]:
    columns = {}
    for col in range(last_period_col + 1, ws.max_column + 1):
        label = normalize_label(ws.cell(4, col).value)
        if label:
            columns.setdefault(label, col)
    return columns


def _cell_text(value) -> str:
    return "" if value is None else str(value).strip()


def _ctb_row_type(ws, row_idx: int) -> str:
    return _cell_text(ws.cell(row_idx, CTB_ROW_TYPE_COL).value)


def _is_balance_row_type(row_type: str) -> bool:
    return row_type.casefold().startswith("balance")


def _template_part_for_row(
    ws,
    row_idx: int,
    periods: Sequence[Period],
    parts_by_part: dict[str, CtbPart],
    shortage: dict[str, ShortageRecord],
) -> CtbPart | None:
    part_no = normalize_part_number(ws.cell(row_idx, 3).value)
    if not part_no:
        return None
    part = parts_by_part.get(part_no)
    if part is not None:
        if part.shortage is None and part_no in shortage:
            part.shortage = shortage[part_no]
        return part
    return CtbPart(
        part=part_no,
        demand=[0.0] * len(periods),
        shortage=shortage.get(part_no),
    )


def _compact_period_text(value: str) -> str:
    return re.sub(r"[^0-9a-z]+", "", str(value).strip().casefold())


def _is_by_day_period(period: Period) -> bool:
    return normalize_label(period.header3) in WEEKDAY_LABELS


def _month_period_key(period: Period) -> str | None:
    for text in (period.header2, period.label, period.header4):
        key = _compact_period_text(text)
        if MONTH_PERIOD_KEY_RE.match(key):
            return key.removesuffix("fcst")
    return None


def _forecast_period_label(period: Period) -> str | None:
    for text in (period.header2, period.label, period.header4):
        key = _compact_period_text(text)
        if key.endswith("fcst") and MONTH_PERIOD_KEY_RE.match(key):
            month_key = key.removesuffix("fcst")
            return f"{month_key[:3].title()}{month_key[3:]}FCST"
    return None


def _week_period_key(period: Period) -> str | None:
    if _is_by_day_period(period):
        return None
    for text in (period.header2, period.label):
        key = _compact_period_text(text)
        if WEEK_PERIOD_KEY_RE.match(key):
            return key
    return None


def _values_for_template_periods(
    source_periods: Sequence[Period],
    source_values: Sequence[float],
    template_periods: Sequence[Period],
) -> list[float]:
    source_by_date: dict[dt.date, float] = {}
    source_month: dict[str, float] = {}
    source_week: dict[str, float] = {}
    source_days: list[tuple[dt.date, float]] = []
    for idx, period in enumerate(source_periods):
        if idx >= len(source_values):
            continue
        value = source_values[idx]
        month_key = _month_period_key(period)
        if month_key:
            source_month[month_key] = source_month.get(month_key, 0.0) + value
            continue
        week_key = _week_period_key(period)
        if week_key:
            source_week[week_key] = source_week.get(week_key, 0.0) + value
            continue
        if period.start is not None:
            source_by_date[period.start] = source_by_date.get(period.start, 0.0) + value
            if _is_by_day_period(period):
                source_days.append((period.start, value))

    values = []
    for period in template_periods:
        month_key = _month_period_key(period)
        if month_key:
            values.append(source_month.get(month_key, 0.0))
            continue
        week_key = _week_period_key(period)
        if week_key:
            if week_key in source_week:
                values.append(source_week[week_key])
            elif period.start is not None:
                period_end = period.start + dt.timedelta(days=6)
                values.append(
                    sum(value for date, value in source_days if period.start <= date <= period_end)
                )
            else:
                values.append(0.0)
            continue
        values.append(source_by_date.get(period.start, 0.0) if period.start is not None else 0.0)
    return values


def _clear_cell(cell) -> None:
    cell.value = None


def _write_period_values(
    ws,
    row_idx: int,
    period_columns: Sequence[tuple[int, Period]],
    values: Sequence[float],
    *,
    write_zero: bool = False,
) -> None:
    for idx, (col, _period) in enumerate(period_columns):
        value = values[idx] if idx < len(values) else 0.0
        if value or write_zero:
            write_number_cell(ws.cell(row_idx, col), clean_number(value))


def _clear_tail_values(ws, row_idx: int, first_col: int) -> None:
    # Data rows only copy A:M from the template, so computed tail cells have no stale values.
    return


def _write_template_part_static(ws, row_idx: int, part: CtbPart, row_type: str) -> None:
    _write_static_cells(ws, row_idx, part, row_type)


def _write_template_eta_static(ws, row_idx: int, record: OpenPoRecord | None) -> None:
    if record is None:
        _clear_cell(ws.cell(row_idx, 4))
        _clear_cell(ws.cell(row_idx, 5))
        _clear_cell(ws.cell(row_idx, 7))
        return
    _write_formula_cell(ws.cell(row_idx, 4), f"={_cell_ref(row_idx, 3)}&{_cell_ref(row_idx, 5)}")
    _write_optional_text_cell(ws.cell(row_idx, 5), record.supplier_site)
    _clear_cell(ws.cell(row_idx, 7))


def _formula_sum_col_range(formula) -> tuple[int, int] | None:
    if not isinstance(formula, str):
        return None
    match = re.search(
        r"(?:SUM|SUBTOTAL)\(\s*(?:\d+\s*,\s*)?\$?([A-Z]+)\$?\d+:\$?([A-Z]+)\$?\d+",
        formula,
        re.IGNORECASE,
    )
    if match is None:
        return None
    return (
        column_index_from_string(match.group(1)),
        column_index_from_string(match.group(2)),
    )


def _formula_negative_ref_col(formula) -> int | None:
    if not isinstance(formula, str):
        return None
    match = re.search(r"-\$?([A-Z]+)\$?\d+\s*$", formula)
    if match is None:
        return None
    return column_index_from_string(match.group(1))


def _period_index_for_column(period_columns: Sequence[tuple[int, Period]], target_col: int | None) -> int | None:
    if target_col is None:
        return None
    for idx, (col, _period) in enumerate(period_columns):
        if col == target_col:
            return idx
    return None


def _last_by_day_period_index(period_columns: Sequence[tuple[int, Period]]) -> int | None:
    last_idx = None
    for idx, (_col, period) in enumerate(period_columns):
        if _is_by_day_period(period):
            last_idx = idx
    return last_idx


def _process_template_group(
    ws,
    formula_ws,
    group_rows: Sequence[int],
    source_periods: Sequence[Period],
    period_columns: Sequence[tuple[int, Period]],
    summary_columns: dict[str, int],
    parts_by_part: dict[str, CtbPart],
    shortage_records: dict[str, ShortageRecord],
    open_po_by_key: dict[str, list[OpenPoRecord]],
    *,
    default_eta_lead_days: int = ETA_LEAD_DAYS,
    eta_lead_days_by_supplier_site: Mapping[str, int] | None = None,
) -> dict[str, int]:
    first_row = group_rows[0]
    template_periods = [period for _col, period in period_columns]
    part = _template_part_for_row(ws, first_row, source_periods, parts_by_part, shortage_records)
    if part is None:
        return {"demand_rows": 0, "eta_rows": 0, "other_rows": 0, "balance_rows": 0}

    demand_values = _values_for_template_periods(source_periods, part.demand, template_periods)
    stats = {"demand_rows": 0, "eta_rows": 0, "other_rows": 0, "balance_rows": 0}
    first_tail_col = period_columns[-1][0] + 1 if period_columns else ws.max_column + 1
    po_remain_period_idx = _last_by_day_period_index(period_columns)
    demand_row_indices: list[int] = []
    eta_row_indices: list[int] = []
    other_row_indices: list[int] = []

    initial_sum_cols = None
    for row_idx in group_rows:
        if _is_balance_row_type(_ctb_row_type(ws, row_idx)):
            initial_sum_cols = _formula_sum_col_range(
                formula_ws.cell(row_idx, period_columns[0][0]).value
            )
            break

    eta_records: list[OpenPoRecord] = []
    eta_keys: dict[int, str] = {}
    for row_idx in group_rows:
        if _ctb_row_type(ws, row_idx).casefold() != "eta":
            continue
        key = normalize_part_number(ws.cell(row_idx, 4).value)
        eta_keys[row_idx] = key
        eta_records.extend(open_po_by_key.get(key, []))
    eta_schedules = eta_schedule_for_records(
        template_periods,
        eta_records,
        demand=demand_values,
        over_shortage=part.shortage.over_shortage if part.shortage else 0.0,
        initial_sum_cols=initial_sum_cols,
        period_start_col=period_columns[0][0],
        default_lead_days=default_eta_lead_days,
        lead_days_by_supplier_site=eta_lead_days_by_supplier_site,
    )

    for row_idx in group_rows:
        row_type = _ctb_row_type(ws, row_idx)
        if row_type.casefold() == "demand":
            demand_row_indices.append(row_idx)
            _write_template_part_static(ws, row_idx, part, row_type)
            _clear_cell(ws.cell(row_idx, 10))
            _clear_cell(ws.cell(row_idx, 11))
            _clear_cell(ws.cell(row_idx, 12))
            _write_period_values(ws, row_idx, period_columns, demand_values, write_zero=True)
            _clear_tail_values(ws, row_idx, first_tail_col)
            demand_col = summary_columns.get("demand")
            if demand_col:
                demand_sum_range = _formula_sum_col_range(formula_ws.cell(row_idx, demand_col).value)
                _write_formula_cell(
                    ws.cell(row_idx, demand_col),
                    _sum_period_formula(row_idx, period_columns, demand_sum_range),
                )
            stats["demand_rows"] += 1
        elif row_type.casefold() == "eta":
            eta_row_indices.append(row_idx)
            key = eta_keys[row_idx]
            records = open_po_by_key.get(key, [])
            schedule = eta_schedules.get(key, [0.0] * len(template_periods))
            _write_template_part_static(ws, row_idx, part, row_type)
            _write_template_eta_static(ws, row_idx, records[0] if records else None)
            _write_formula_cell(ws.cell(row_idx, CTB_PO_REMAIN_COL), _open_po_lookup_formula(row_idx))
            _write_formula_cell(ws.cell(row_idx, CTB_TOTAL_COL), _sum_period_formula(row_idx, period_columns))
            _write_period_values(ws, row_idx, period_columns, schedule)
            _clear_tail_values(ws, row_idx, first_tail_col)
            po_col = summary_columns.get("po")
            if po_col:
                po_sum_range = _formula_sum_col_range(formula_ws.cell(row_idx, po_col).value)
                _write_formula_cell(
                    ws.cell(row_idx, po_col),
                    _sum_period_formula(row_idx, period_columns, po_sum_range),
                )
            stats["eta_rows"] += 1
        elif row_type.casefold() == "other":
            other_row_indices.append(row_idx)
            _write_template_part_static(ws, row_idx, part, row_type)
            _clear_cell(ws.cell(row_idx, 10))
            _clear_cell(ws.cell(row_idx, 11))
            _clear_cell(ws.cell(row_idx, 12))
            _write_period_values(ws, row_idx, period_columns, [0.0] * len(template_periods))
            _clear_tail_values(ws, row_idx, first_tail_col)
            stats["other_rows"] += 1

    demand_row = demand_row_indices[0] if demand_row_indices else first_row
    for row_idx in group_rows:
        row_type = _ctb_row_type(ws, row_idx)
        if not _is_balance_row_type(row_type):
            continue
        initial_sum_cols = _formula_sum_col_range(formula_ws.cell(row_idx, period_columns[0][0]).value)
        _write_template_part_static(ws, row_idx, part, row_type)
        po_remain_idx = po_remain_period_idx
        if po_remain_idx is None:
            po_remain_ref_col = _formula_negative_ref_col(formula_ws.cell(row_idx, 11).value)
            po_remain_idx = _period_index_for_column(period_columns, po_remain_ref_col)
        po_remain_formula = _po_remain_formula(row_idx, period_columns, po_remain_idx)
        if part.shortage is not None and po_remain_formula is not None:
            _write_formula_cell(ws.cell(row_idx, CTB_PO_REMAIN_COL), po_remain_formula)
        elif part.shortage is not None:
            write_number_cell(ws.cell(row_idx, 11), clean_number(part.shortage.po_remain))
        else:
            _clear_cell(ws.cell(row_idx, 11))
        _clear_cell(ws.cell(row_idx, 12))
        _write_balance_formulas(
            ws,
            row_idx,
            demand_row,
            eta_row_indices,
            other_row_indices,
            period_columns,
            initial_sum_cols,
        )
        _clear_tail_values(ws, row_idx, first_tail_col)
        stats["balance_rows"] += 1

    return stats


def write_ctb_from_template(
    wb: Workbook,
    template_path: Path,
    source_periods: Sequence[Period],
    parts_by_part: dict[str, CtbPart],
    shortage_records: dict[str, ShortageRecord],
    open_po_records: Sequence[OpenPoRecord],
    *,
    default_eta_lead_days: int = ETA_LEAD_DAYS,
    eta_lead_days_by_supplier_site: Mapping[str, int] | None = None,
) -> dict:
    template_wb = load_workbook(template_path, data_only=True)
    formula_wb = load_workbook(template_path, data_only=False)
    try:
        template_name = _sheet_name(template_wb, CTB_SHEET)
        formula_name = _sheet_name(formula_wb, CTB_SHEET)
        if template_name is None or formula_name is None:
            raise SystemExit(f"{template_path.name} 內找不到 {CTB_SHEET} 工作表")
        template_ws = template_wb[template_name]
        formula_ws = formula_wb[formula_name]
        ws = wb.active
        _copy_template_sheet(template_ws, ws)

        period_columns = _template_period_columns(ws)
        if not period_columns:
            raise SystemExit(f"{template_path.name} 的 CTB 找不到日期期間欄")
        summary_columns = _template_summary_columns(ws, period_columns[-1][0])
        open_po_by_key: dict[str, list[OpenPoRecord]] = defaultdict(list)
        for record in open_po_records:
            open_po_by_key[record.key].append(record)

        row_idx = 5
        demand_rows = eta_rows = other_rows = balance_rows = 0
        template_parts = set()
        while row_idx <= ws.max_row:
            row_type = _ctb_row_type(ws, row_idx)
            if row_type.casefold() != "demand":
                row_idx += 1
                continue
            group_start = row_idx
            row_idx += 1
            while row_idx <= ws.max_row and _ctb_row_type(ws, row_idx).casefold() != "demand":
                row_idx += 1
            group_rows = list(range(group_start, row_idx))
            part_no = normalize_part_number(ws.cell(group_start, 3).value)
            if part_no:
                template_parts.add(part_no)
            stats = _process_template_group(
                ws,
                formula_ws,
                group_rows,
                source_periods,
                period_columns,
                summary_columns,
                parts_by_part,
                shortage_records,
                open_po_by_key,
                default_eta_lead_days=default_eta_lead_days,
                eta_lead_days_by_supplier_site=eta_lead_days_by_supplier_site,
            )
            demand_rows += stats["demand_rows"]
            eta_rows += stats["eta_rows"]
            other_rows += stats["other_rows"]
            balance_rows += stats["balance_rows"]

        balance_row_indices = [
            row
            for row in range(5, ws.max_row + 1)
            if _is_balance_row_type(_ctb_row_type(ws, row))
        ]
        _apply_ctb_negative_formatting(ws, balance_row_indices, period_columns)
        set_filter_to_used_range(ws, ws.max_column, ws.max_row)
        return {
            "mode": "template",
            "parts": len(template_parts),
            "demand_rows": demand_rows,
            "eta_rows": eta_rows,
            "other_rows": other_rows,
            "balance_rows": balance_rows,
            "periods": len(period_columns),
            "template_rows": ws.max_row,
            "template_source": template_path,
        }
    finally:
        template_wb.close()
        formula_wb.close()


def generate_ctb(
    dps_pp_path: Path,
    bom_path: Path,
    open_po_path: Path,
    over_shortage_path: Path,
    output_path: Path,
    template_path: Path | None = None,
    *,
    default_eta_lead_days: int = ETA_LEAD_DAYS,
    eta_lead_days_by_supplier_site: Mapping[str, int] | None = None,
) -> dict:
    periods, demand_by_parent = read_dps_pp(dps_pp_path)
    bom_rows = read_bom_rows(bom_path, periods, demand_by_parent)
    shortage = read_over_shortage(over_shortage_path)
    open_po = read_open_po(open_po_path)
    parts_by_part, part_order = build_part_map(periods, bom_rows, shortage, open_po)
    parts = filter_ctb_parts(parts_by_part, part_order)

    wb = Workbook()
    if template_path is not None and workbook_has_sheet(template_path, CTB_SHEET):
        stats = write_ctb_from_template(
            wb,
            template_path,
            periods,
            parts_by_part,
            shortage,
            open_po,
            default_eta_lead_days=default_eta_lead_days,
            eta_lead_days_by_supplier_site=eta_lead_days_by_supplier_site,
        )
    else:
        stats = {
            "mode": "generated",
            **write_ctb_sheet(
                wb,
                periods,
                parts,
                default_eta_lead_days=default_eta_lead_days,
                eta_lead_days_by_supplier_site=eta_lead_days_by_supplier_site,
            ),
        }
    copy_dps_pp_sheet(wb, dps_pp_path)
    write_auxiliary_sheets(wb, periods, bom_rows, shortage, open_po)

    _enable_formula_recalculation(wb)
    output_path.parent.mkdir(parents=True, exist_ok=True)
    wb.save(output_path)
    return {
        **stats,
        "dps_pp_source": dps_pp_path,
        "bom_source": bom_path,
        "open_po_source": open_po_path,
        "over_shortage_source": over_shortage_path,
        "template_source": template_path if stats.get("mode") == "template" else None,
        "bom_rows": len(bom_rows),
        "over_shortage_rows": len(shortage),
        "open_po_rows": len(open_po),
        "output": output_path,
    }
