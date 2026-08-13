"""PP report generation and pivot cache parsing."""

from __future__ import annotations

import datetime as dt
import posixpath
import re
import xml.etree.ElementTree as ET
import zipfile
from collections import OrderedDict, defaultdict
from pathlib import Path
from typing import Sequence

from openpyxl import Workbook, load_workbook
from openpyxl.styles import Alignment, Font
from openpyxl.utils import get_column_letter

from .common import (
    autosize,
    clean_number,
    copy_cell_format,
    copy_column_layout,
    copy_row_layout,
    DEFAULT_PP_PART_NUMBER_FIELD_KEYWORDS,
    DEFAULT_PP_SHEET_KEYWORDS,
    find_total_col,
    first_existing_sheet,
    keyword_label,
    normalize_label,
    numeric,
    serial_to_date,
    set_filter_to_used_range,
    sheet_name_matches_keywords,
    warn,
)

SPREADSHEET_NS = "http://schemas.openxmlformats.org/spreadsheetml/2006/main"
OFFICE_REL_NS = "http://schemas.openxmlformats.org/officeDocument/2006/relationships"

MONTH_ABBR = [
    "Jan", "Feb", "Mar", "Apr", "May", "Jun",
    "Jul", "Aug", "Sep", "Oct", "Nov", "Dec",
]
MONTH_INDEX = {name.lower(): i for i, name in enumerate(MONTH_ABBR, start=1)}
PP_SOURCE_SHEET_KEYWORDS = DEFAULT_PP_SHEET_KEYWORDS
PP_PART_NUMBER_FIELD_KEYWORDS = DEFAULT_PP_PART_NUMBER_FIELD_KEYWORDS

# ---------------------------------------------------------------------------
# PP：樞紐快取解析
# ---------------------------------------------------------------------------


def _rels_part_name(part_name: str) -> str:
    folder, filename = posixpath.split(part_name)
    return posixpath.join(folder, "_rels", f"{filename}.rels")


def _resolve_part(source_part: str, target: str) -> str:
    if target.startswith("/"):
        return posixpath.normpath(target.lstrip("/"))
    return posixpath.normpath(posixpath.join(posixpath.dirname(source_part), target))


def _relationships(xlsx: zipfile.ZipFile, rels_name: str) -> list[dict[str, str]]:
    if rels_name not in xlsx.namelist():
        return []
    rels = ET.fromstring(xlsx.read(rels_name))
    return [dict(rel.attrib) for rel in rels]


def _workbook_sheets(xlsx: zipfile.ZipFile) -> list[dict[str, str]]:
    workbook_part = "xl/workbook.xml"
    workbook = ET.fromstring(xlsx.read(workbook_part))
    rels = {
        rel["Id"]: rel["Target"]
        for rel in _relationships(xlsx, _rels_part_name(workbook_part))
        if "Id" in rel and "Target" in rel
    }
    sheets_root = workbook.find(f"{{{SPREADSHEET_NS}}}sheets")
    sheets: list[dict[str, str]] = []
    if sheets_root is None:
        return sheets
    for sheet in sheets_root:
        rid = sheet.attrib.get(f"{{{OFFICE_REL_NS}}}id")
        if not rid or rid not in rels:
            continue
        sheets.append({
            "name": sheet.attrib.get("name", ""),
            "part": _resolve_part(workbook_part, rels[rid]),
        })
    return sheets


def _workbook_pivot_cache_definitions(xlsx: zipfile.ZipFile) -> dict[str, str]:
    workbook_part = "xl/workbook.xml"
    workbook = ET.fromstring(xlsx.read(workbook_part))
    rels = {
        rel["Id"]: rel["Target"]
        for rel in _relationships(xlsx, _rels_part_name(workbook_part))
        if "Id" in rel and "Target" in rel
    }
    pivot_caches = workbook.find(f"{{{SPREADSHEET_NS}}}pivotCaches")
    cache_parts: dict[str, str] = {}
    if pivot_caches is None:
        return cache_parts
    for pivot_cache in pivot_caches:
        cache_id = pivot_cache.attrib.get("cacheId")
        rid = pivot_cache.attrib.get(f"{{{OFFICE_REL_NS}}}id")
        if cache_id and rid in rels:
            cache_parts[cache_id] = _resolve_part(workbook_part, rels[rid])
    return cache_parts


def _pivot_table_cache_definition(
    xlsx: zipfile.ZipFile,
    pivot_table_part: str,
    workbook_cache_defs: dict[str, str],
) -> tuple[str | None, str | None]:
    cache_id = None
    if pivot_table_part in xlsx.namelist():
        pivot_table = ET.fromstring(xlsx.read(pivot_table_part))
        cache_id = pivot_table.attrib.get("cacheId")

    for rel in _relationships(xlsx, _rels_part_name(pivot_table_part)):
        target = rel.get("Target", "")
        rel_type = rel.get("Type", "")
        if "pivotCacheDefinition" in rel_type or "pivotCacheDefinition" in target:
            return _resolve_part(pivot_table_part, target), cache_id

    if cache_id is not None:
        return workbook_cache_defs.get(cache_id), cache_id
    return None, cache_id


def _sheet_pivot_tables(
    xlsx: zipfile.ZipFile,
    sheet_part: str,
    workbook_cache_defs: dict[str, str],
) -> list[dict[str, str | None]]:
    pivots: list[dict[str, str | None]] = []
    for rel in _relationships(xlsx, _rels_part_name(sheet_part)):
        target = rel.get("Target", "")
        rel_type = rel.get("Type", "")
        if "pivotTable" not in rel_type and "pivotTable" not in target:
            continue
        pivot_table_part = _resolve_part(sheet_part, target)
        cache_definition, cache_id = _pivot_table_cache_definition(
            xlsx, pivot_table_part, workbook_cache_defs
        )
        pivots.append({
            "pivot_table": pivot_table_part,
            "cache_definition": cache_definition,
            "cache_id": cache_id,
        })
    return pivots


def _cache_field_names(xlsx: zipfile.ZipFile, definition_part: str) -> list[str]:
    definition = ET.fromstring(xlsx.read(definition_part))
    cache_fields = definition.find(f"{{{SPREADSHEET_NS}}}cacheFields")
    if cache_fields is None:
        return []
    return [cf.attrib.get("name", "") for cf in cache_fields]


def field_matches_keywords(field_name: str, keywords: Sequence[str]) -> bool:
    normalized = normalize_label(field_name)
    return any(normalize_label(keyword) in normalized for keyword in keywords)


def select_part_number_field(fields: Sequence[str], keywords: Sequence[str]) -> str | None:
    matches = [
        normalize_field(field)
        for field in fields
        if field_matches_keywords(normalize_field(field), keywords)
    ]
    if not matches:
        return None

    fg_matches = [field for field in matches if re.search(r"\bFG\b", field, re.IGNORECASE)]
    if fg_matches:
        if len(fg_matches) > 1:
            warn(
                "PP 樞紐快取中有多個符合料號關鍵字且包含 FG 的欄位，"
                f"將使用第一個：{fg_matches[0]}；候選：{', '.join(fg_matches)}"
            )
        return fg_matches[0]

    if len(matches) > 1:
        warn(
            "PP 樞紐快取中有多個符合料號關鍵字的欄位，"
            f"將使用第一個：{matches[0]}；候選：{', '.join(matches)}"
        )
    return matches[0]


def select_pp_pivot_source(
    pp_path: Path,
    sheet_keywords: Sequence[str] = PP_SOURCE_SHEET_KEYWORDS,
    part_number_keywords: Sequence[str] = PP_PART_NUMBER_FIELD_KEYWORDS,
) -> dict[str, str]:
    keyword_text = keyword_label(sheet_keywords)
    part_keyword_text = keyword_label(part_number_keywords)
    with zipfile.ZipFile(pp_path) as xlsx:
        workbook_cache_defs = _workbook_pivot_cache_definitions(xlsx)
        matched_sheet_names: list[str] = []
        skipped_no_pivot: list[str] = []
        for sheet in _workbook_sheets(xlsx):
            sheet_name = sheet["name"]
            if not sheet_name_matches_keywords(sheet_name, sheet_keywords):
                continue
            matched_sheet_names.append(sheet_name)
            pivots = _sheet_pivot_tables(xlsx, sheet["part"], workbook_cache_defs)
            if not pivots:
                skipped_no_pivot.append(sheet_name)
                continue

            for pivot in pivots:
                cache_definition = pivot["cache_definition"]
                if cache_definition is None:
                    continue
                fields = _cache_field_names(xlsx, cache_definition)
                part_number_field = select_part_number_field(fields, part_number_keywords)
                if part_number_field is not None:
                    if skipped_no_pivot:
                        warn(
                            f"下列工作表名稱包含 {keyword_text}，但沒有樞紐分析表，已略過："
                            + ", ".join(skipped_no_pivot)
                        )
                    return {
                        "sheet_name": sheet_name,
                        "sheet_part": sheet["part"],
                        "pivot_table": pivot["pivot_table"] or "",
                        "cache_definition": cache_definition,
                        "cache_id": pivot["cache_id"] or "",
                        "part_number_field": part_number_field,
                    }
            raise SystemExit(
                f"{pp_path.name} 的 {sheet_name} 工作表有樞紐分析表，"
                f"但其樞紐快取找不到名稱包含 {part_keyword_text} 的料號欄位。"
            )

    if matched_sheet_names:
        raise SystemExit(
            f"{pp_path.name} 找到名稱包含 {keyword_text} 的工作表"
            f"（{', '.join(matched_sheet_names)}），"
            "但沒有任何一張含樞紐分析表。"
        )
    raise SystemExit(f"{pp_path.name} 找不到名稱包含 {keyword_text} 且含樞紐分析表的工作表。")


def _pivot_cache_parts(xlsx: zipfile.ZipFile) -> list[tuple[str, str]]:
    """列出活頁簿內所有 pivotCacheDefinition 及其對應的 records 檔。

    舊版寫死 pivotCacheDefinition1.xml；一旦來源檔含多個樞紐快取，編號就會
    對不上。這裡改成從 zip 目錄動態列舉，並用 rels 找到對應的 records。
    """
    names = xlsx.namelist()
    parts: list[tuple[str, str]] = []
    for definition in sorted(n for n in names if re.fullmatch(
        r"xl/pivotCache/pivotCacheDefinition\d+\.xml", n
    )):
        records = None
        rels_name = definition.replace(
            "pivotCache/", "pivotCache/_rels/"
        ) + ".rels"
        if rels_name in names:
            for rel in _relationships(xlsx, rels_name):
                target = rel.get("Target", "")
                if "pivotCacheRecords" in target:
                    records = _resolve_part(definition, target)
                    break
        if records is None:
            guess = definition.replace("Definition", "Records")
            records = guess if guess in names else None
        if records:
            parts.append((definition, records))
    return parts


def parse_pivot_cache(
    pp_path: Path,
    definition_part: str | None = None,
    part_number_keywords: Sequence[str] = PP_PART_NUMBER_FIELD_KEYWORDS,
) -> dict:
    with zipfile.ZipFile(pp_path) as xlsx:
        parts = _pivot_cache_parts(xlsx)
        if not parts:
            raise SystemExit(f"{pp_path.name} 內找不到樞紐快取（pivotCache），無法取得逐料號明細")
        if definition_part is not None:
            parts = [part for part in parts if part[0] == definition_part]
            if not parts:
                raise SystemExit(f"{pp_path.name} 找不到指定的樞紐快取：{definition_part}")

        chosen = None
        for definition_name, records_name in parts:
            definition = ET.fromstring(xlsx.read(definition_name))
            cache_fields = definition.find(f"{{{SPREADSHEET_NS}}}cacheFields")
            if cache_fields is None:
                fields = []
            else:
                fields = [cf.attrib.get("name", "") for cf in cache_fields]
            part_number_field = select_part_number_field(fields, part_number_keywords)
            if part_number_field is not None:
                chosen = (definition_name, records_name, definition, fields, part_number_field)
                break
        if chosen is None:
            raise SystemExit(
                f"樞紐快取中找不到名稱包含 {keyword_label(part_number_keywords)} 的料號欄位，"
                "來源檔格式可能已變更"
            )
        definition_name, records_name, definition, fields, part_number_field = chosen

        shared_by_field: list[list[str]] = []
        cache_fields = definition.find(f"{{{SPREADSHEET_NS}}}cacheFields")
        if cache_fields is None:
            raise SystemExit("樞紐快取缺少 cacheFields，來源檔格式可能已變更")
        for cache_field in cache_fields:
            shared_items: list[str] = []
            shared_root = cache_field.find(f"{{{SPREADSHEET_NS}}}sharedItems")
            if shared_root is not None:
                for item in shared_root:
                    tag = item.tag.split("}", 1)[-1]
                    shared_items.append("" if tag == "m" else item.attrib.get("v", ""))
            shared_by_field.append(shared_items)

        records_root = ET.fromstring(xlsx.read(records_name))
        records: list[list[str]] = []
        for record in records_root.findall(f"{{{SPREADSHEET_NS}}}r"):
            row: list[str] = []
            for index, child in enumerate(record):
                tag = child.tag.split("}", 1)[-1]
                if tag == "x":
                    shared_index = int(child.attrib.get("v", "0"))
                    values = shared_by_field[index]
                    row.append(values[shared_index] if shared_index < len(values) else "")
                else:
                    row.append("" if tag == "m" else child.attrib.get("v", ""))
            records.append(row)

        refreshed_raw = definition.attrib.get("refreshedDate")
        refreshed_by = definition.attrib.get("refreshedBy", "")
        refreshed_date = None
        if refreshed_raw:
            try:
                refreshed_date = serial_to_date(float(refreshed_raw))
            except ValueError:
                refreshed_date = None

        source = definition.find(f"{{{SPREADSHEET_NS}}}cacheSource")
        source_sheet = ""
        if source is not None:
            ws_source = source.find(f"{{{SPREADSHEET_NS}}}worksheetSource")
            if ws_source is not None:
                source_sheet = ws_source.attrib.get("sheet", "")

        return {
            "fields": fields,
            "records": records,
            "part_number_field": part_number_field,
            "definition_part": definition_name,
            "refreshed_date": refreshed_date,
            "refreshed_by": refreshed_by,
            "source_sheet": source_sheet,
        }


# ---------------------------------------------------------------------------
# PP：期間欄位推導
# ---------------------------------------------------------------------------

WEEK_LABEL_RE = re.compile(r"^WK\s*(\d{1,2})(?:\s+([A-Za-z]{3,9}))?\s*'?$", re.IGNORECASE)
MONTH_FCST_LABEL_RE = re.compile(r"^([A-Za-z]{3})\s*'?\s*(\d{2})\s*FCST$", re.IGNORECASE)
MONTH_PLAIN_LABEL_RE = re.compile(r"^([A-Za-z]{3})\s*(?:-|')\s*(\d{2})\s*'?$", re.IGNORECASE)
TOTAL_LABEL_RE = re.compile(r"total", re.IGNORECASE)

CACHE_WEEK_OLD_RE = re.compile(r"^WK\s*(\d{1,2})\s+(\d{2})'([A-Za-z]{3,9})", re.IGNORECASE)
CACHE_WEEK_MONTH_YEAR_RE = re.compile(
    r"^WK\s*(\d{1,2})\s+([A-Za-z]{3,9})\s*'?\s*(\d{2})",
    re.IGNORECASE,
)
CACHE_MONTH_RE = re.compile(r"^([A-Za-z]{3})\s*(?:'|-)\s*(\d{2})\s*(?:FCST)?$", re.IGNORECASE)


def normalize_field(name: str) -> str:
    return name.replace("_x000a_", " ").replace("\n", " ").strip()


def parse_cache_week(name: str) -> tuple[str, int, str] | None:
    text = normalize_field(name)
    m = CACHE_WEEK_OLD_RE.match(text)
    if m:
        return m.group(2), int(m.group(1)), m.group(3)[:3].title()
    m = CACHE_WEEK_MONTH_YEAR_RE.match(text)
    if m:
        return m.group(3), int(m.group(1)), m.group(2)[:3].title()
    return None


def index_cache_periods(fields: Sequence[str]) -> tuple[dict, dict]:
    """建立 (年,週) -> 欄位索引清單、(年,月) -> 欄位索引 兩張對照表。"""
    weeks: dict[tuple[str, int], list[int]] = defaultdict(list)
    months: dict[tuple[str, str], int] = {}
    for idx, raw in enumerate(fields):
        name = normalize_field(raw)
        week = parse_cache_week(name)
        if week:
            year, week_num, _month = week
            weeks[(year, week_num)].append(idx)
            continue
        m = CACHE_MONTH_RE.match(name)
        if m and m.group(1).lower() in MONTH_INDEX:
            months.setdefault((m.group(2), m.group(1).title()), idx)
    return dict(weeks), months


def find_layout_row(wb, sheet_name: str | None = None) -> tuple[object, int, int] | None:
    """找出可見樞紐報表中，描述期間欄位版面的那一列。

    這一列（例：`WK27 Jul | WK28 | ... | Oct-26 | Nov26FCST | 2026 TOTAL | Jan'27 FCST`）
    就是人工整理版的欄位藍本，靠它推導「週明細到哪個月為止、之後改用月預測」，
    不必把 30~44 週、Nov/Dec、'27 這些寫死在程式裡。
    """
    best = None
    worksheets = [wb[sheet_name]] if sheet_name is not None else wb.worksheets
    for ws in worksheets:
        for row in ws.iter_rows(min_row=1, max_row=40):
            hits = 0
            first_col = None
            for cell in row:
                text = str(cell.value).strip() if cell.value is not None else ""
                if not text:
                    continue
                if (
                    WEEK_LABEL_RE.match(text)
                    or MONTH_FCST_LABEL_RE.match(text)
                    or MONTH_PLAIN_LABEL_RE.match(text)
                ):
                    hits += 1
                    if first_col is None:
                        first_col = cell.column
            if hits >= 8 and (best is None or hits > best[3]):
                best = (ws, row[0].row, first_col, hits)
    if best is None:
        return None
    return best[0], best[1], best[2]


def build_pp_periods(
    fields: Sequence[str],
    layout: list[str] | None,
    base_year: str,
    start_week: int,
) -> list[tuple[str, list[int]]]:
    """回傳 [(輸出欄名, [快取欄位索引...])]。"""
    weeks, months = index_cache_periods(fields)

    if not layout:
        # 沒有可見版面可參考時的退路：週明細只取到最後一個「有對應月份且
        # 月份在起始週所屬月份之後 3 個月內」的週；其餘用月預測。
        warn("找不到可見樞紐報表的欄位版面，改用推導模式（週明細取 15 週）。")
        layout = [f"WK{w:02d}" for w in range(start_week, start_week + 15)]
        covered_months = set()
        for w in range(start_week, start_week + 15):
            for idx in weeks.get((base_year, w), []):
                week = parse_cache_week(normalize_field(fields[idx]))
                if week:
                    covered_months.add(week[2])
        for month in MONTH_ABBR:
            if (base_year, month) in months and month not in covered_months:
                if MONTH_INDEX[month.lower()] > max(
                    (MONTH_INDEX[c.lower()] for c in covered_months), default=0
                ):
                    layout.append(f"{month}{base_year}FCST")
        next_year = f"{int(base_year) + 1:02d}"
        for month in MONTH_ABBR:
            if (next_year, month) in months:
                layout.append(f"{month}'{next_year} FCST")

    # 第一輪：走一遍版面，判斷每個週欄屬於哪個月、哪些月份有週明細
    tokens: list[tuple[str, object, str]] = []  # (kind, key, label)
    current_month: str | None = None
    weekly_months: set[tuple[str, str]] = set()

    for label in layout:
        text = label.strip()
        if not text:
            continue
        if TOTAL_LABEL_RE.search(text) and not WEEK_LABEL_RE.match(text):
            tokens.append(("total", None, text))
            current_month = None
            continue
        m = WEEK_LABEL_RE.match(text)
        if m:
            week = int(m.group(1))
            hint = m.group(2)
            month = None
            if hint and hint[:3].lower() in MONTH_INDEX:
                month = hint[:3].title()
            else:
                candidates = []
                for idx in weeks.get((base_year, week), []):
                    cache_week = parse_cache_week(normalize_field(fields[idx]))
                    if cache_week:
                        candidates.append(cache_week[2])
                if candidates:
                    month = current_month if current_month in candidates else candidates[0]
            if month:
                current_month = month
                weekly_months.add((base_year, month))
            tokens.append(("week", week, f"WK{week:02d}"))
            continue
        m = MONTH_FCST_LABEL_RE.match(text) or MONTH_PLAIN_LABEL_RE.match(text)
        if m:
            month = m.group(1).title()
            year = m.group(2)
            tokens.append(("month", (year, month), text))
            current_month = None
            continue
        tokens.append(("other", None, text))

    # 第二輪：從起始週開始輸出，跳過月小計欄與年度合計欄
    periods: "OrderedDict[str, list[int]]" = OrderedDict()
    started = False
    for kind, key, label in tokens:
        if not started:
            if kind == "week" and key == start_week:
                started = True
            else:
                continue
        if kind == "week":
            idxs = weeks.get((base_year, key), [])
            if not idxs:
                warn(f"快取中找不到 {base_year}' 年的 WK{key:02d}，該欄以 0 輸出。")
            periods.setdefault(label, [])
            for idx in idxs:
                if idx not in periods[label]:
                    periods[label].append(idx)
        elif kind == "month":
            if key in weekly_months:
                continue  # 該月已有週明細，這是小計欄，不重複輸出
            idx = months.get(key)
            if idx is None:
                warn(f"快取中找不到月份欄位 {key[1]}-{key[0]}，略過。")
                continue
            periods.setdefault(label, [idx])

    if not started:
        raise SystemExit(
            f"版面中找不到起始週 WK{start_week:02d}；請用 --pp-start-week 指定正確的起始週。"
        )
    return list(periods.items())


def read_layout(pp_path: Path, sheet_name: str | None = None) -> tuple[list[str] | None, list[str]]:
    """回傳 (期間欄位標籤序列, 客戶顯示順序)。"""
    wb = load_workbook(pp_path, data_only=True)
    try:
        found = find_layout_row(wb, sheet_name)
        if found is None:
            return None, []
        ws, row_idx, first_col = found
        labels = [
            str(cell.value).strip()
            for cell in ws[row_idx]
            if cell.column >= first_col and cell.value is not None and str(cell.value).strip()
        ]
        customers: list[str] = []
        for row in ws.iter_rows(min_row=row_idx + 1, max_col=max(1, first_col - 1)):
            for cell in row:
                text = str(cell.value).strip() if cell.value is not None else ""
                if text and "TTL" not in text.upper() and text not in customers:
                    customers.append(text)
        return labels, customers
    finally:
        wb.close()


def generate_pp(
    pp_path: Path,
    output_path: Path,
    plan: str = "Production Input",
    start_week: int | None = None,
    base_year: str | None = None,
    report_date: dt.date | None = None,
    sheet_keywords: Sequence[str] = PP_SOURCE_SHEET_KEYWORDS,
    part_number_keywords: Sequence[str] = PP_PART_NUMBER_FIELD_KEYWORDS,
) -> dict:
    pivot_source = select_pp_pivot_source(
        pp_path,
        sheet_keywords=sheet_keywords,
        part_number_keywords=part_number_keywords,
    )
    cache = parse_pivot_cache(
        pp_path,
        definition_part=pivot_source["cache_definition"],
        part_number_keywords=part_number_keywords,
    )
    fields = cache["fields"]
    records = cache["records"]

    field_index = {normalize_field(name): idx for idx, name in enumerate(fields)}
    part_number_field = cache["part_number_field"]
    required = ["Customer", "Model", "Plan"]
    missing = [name for name in required if name not in field_index]
    if missing:
        raise SystemExit(f"樞紐快取缺少必要欄位：{', '.join(missing)}")
    if part_number_field not in field_index:
        raise SystemExit(f"樞紐快取找不到選定的料號欄位：{part_number_field}")

    customer_idx = field_index["Customer"]
    model_idx = field_index["Model"]
    pn_idx = field_index[part_number_field]
    plan_idx = field_index["Plan"]

    if report_date is None:
        report_date = cache["refreshed_date"] or dt.date.today()
    if base_year is None:
        base_year = f"{report_date.year % 100:02d}"
    if start_week is None:
        # 取報表日所在週的前一週開始（與人工整理慣例一致：保留上一週作參照）
        start_week = max(1, report_date.isocalendar()[1] - 1)

    layout_labels, customer_order = read_layout(pp_path, sheet_name=pivot_source["sheet_name"])
    periods = build_pp_periods(fields, layout_labels, base_year, start_week)
    if not periods:
        raise SystemExit("推導不出任何輸出期間欄位")

    plans_seen = {r[plan_idx].strip() for r in records if len(r) > plan_idx}
    if plan not in plans_seen:
        raise SystemExit(
            f"快取中沒有 Plan = {plan!r}；可選值：{', '.join(sorted(plans_seen))}"
        )

    aggregate: dict[str, dict[str, float]] = defaultdict(lambda: defaultdict(float))
    metadata: dict[str, tuple[str, str]] = {}
    source_order: list[str] = []
    plan_rows = 0

    for record in records:
        if len(record) <= max(customer_idx, model_idx, pn_idx, plan_idx):
            continue
        if record[plan_idx].strip() != plan:
            continue
        pn = record[pn_idx].strip()
        if not pn:
            continue
        plan_rows += 1
        if pn not in metadata:
            metadata[pn] = (record[customer_idx].strip(), record[model_idx].strip())
            source_order.append(pn)
        for label, source_indexes in periods:
            aggregate[pn][label] += sum(
                numeric(record[i]) for i in source_indexes if i < len(record)
            )

    labels = [label for label, _ in periods]
    kept = [pn for pn in source_order if any(aggregate[pn][label] for label in labels)]

    rank = {name: i for i, name in enumerate(customer_order)}
    fallback = len(rank)
    kept.sort(key=lambda pn: (rank.get(metadata[pn][0], fallback), metadata[pn][0], pn))

    rows = []
    for pn in kept:
        customer, model = metadata[pn]
        rows.append(
            [customer, pn, model]
            + [clean_number(aggregate[pn][label]) for label in labels]
        )

    write_pp_workbook(
        output_path,
        labels,
        rows,
        part_number_header=part_number_field,
        template_path=pp_path,
    )

    return {
        "source": pp_path,
        "layout_sheet": pivot_source["sheet_name"],
        "pivot_table": pivot_source["pivot_table"],
        "cache_id": pivot_source["cache_id"],
        "cache_part": cache["definition_part"],
        "cache_source_sheet": cache["source_sheet"],
        "part_number_field": part_number_field,
        "refreshed_date": cache["refreshed_date"],
        "refreshed_by": cache["refreshed_by"],
        "records": len(records),
        "plan": plan,
        "plan_rows": plan_rows,
        "base_year": base_year,
        "start_week": start_week,
        "report_date": report_date,
        "layout_found": layout_labels is not None,
        "periods": labels,
        "rows": len(rows),
        "dropped_zero": len(metadata) - len(kept),
        "grand_total": clean_number(sum(sum(r[3:]) for r in rows)),
    }


PP_TIDY_SHEET = "PP整理後"
PP_COMPARE_SHEETS = (PP_TIDY_SHEET, "整理后PP")
PP_SPACER_COLS = 3  # 人工版在 total 前留了三個空白欄，這裡照樣保留以維持版面一致


def write_pp_workbook(
    output_path: Path,
    labels: Sequence[str],
    rows: Sequence[list],
    part_number_header: str = "AVTC FG Part Number",
    template_path: Path | None = None,
) -> None:
    wb = Workbook()
    ws = wb.active
    ws.title = PP_TIDY_SHEET

    ws.cell(1, 1, "Customer")
    ws.cell(1, 2, part_number_header)
    ws.cell(1, 3, "Model")
    for offset, label in enumerate(labels):
        ws.cell(1, 4 + offset, label)
    first_col = 4
    last_col = 3 + len(labels)
    total_col = last_col + PP_SPACER_COLS + 1
    ws.cell(1, total_col, "total")

    first_letter = get_column_letter(first_col)
    last_letter = get_column_letter(last_col)

    for r_offset, row in enumerate(rows):
        r = 2 + r_offset
        ws.cell(r, 1, row[0])
        ws.cell(r, 2, row[1])
        ws.cell(r, 3, row[2])
        for offset, value in enumerate(row[3:]):
            cell = ws.cell(r, 4 + offset, value)
            cell.number_format = "#,##0_ "
        total = ws.cell(r, total_col, f"=SUM({first_letter}{r}:{last_letter}{r})")
        total.number_format = "#,##0_ "

    for cell in ws[1]:
        if cell.value is not None:
            cell.font = Font(bold=True)
            cell.alignment = Alignment(horizontal="center")
    ws.freeze_panes = "D2"
    autosize(ws)
    ws.column_dimensions["A"].width = 17.0
    ws.column_dimensions["B"].width = 28.6
    ws.column_dimensions["C"].width = 20.7
    for offset in range(len(labels)):
        ws.column_dimensions[get_column_letter(4 + offset)].width = 13.0
    last_row = len(rows) + 1
    apply_pp_layout(ws, template_path, last_col, total_col, last_row)

    output_path.parent.mkdir(parents=True, exist_ok=True)
    wb.save(output_path)


def apply_pp_layout(
    ws,
    template_path: Path | None,
    last_period_col: int,
    total_col: int,
    last_row: int,
) -> None:
    set_filter_to_used_range(ws, total_col, last_row)
    if template_path is None:
        return

    template_wb = None
    try:
        template_wb = load_workbook(template_path, data_only=False)
        sheet_name = first_existing_sheet(template_wb, PP_COMPARE_SHEETS)
        if sheet_name is None:
            return
        template_ws = template_wb[sheet_name]
        template_total_col = find_total_col(template_ws) or min(total_col, template_ws.max_column)
        template_last_period_col = max(4, template_total_col - PP_SPACER_COLS - 1)

        def source_col(target_col: int) -> int:
            if target_col <= 3:
                return min(target_col, template_ws.max_column)
            if target_col == total_col:
                return template_total_col
            if last_period_col < target_col < total_col:
                spacer_offset = target_col - last_period_col
                return min(template_last_period_col + spacer_offset, template_ws.max_column)
            if target_col <= template_last_period_col:
                return min(target_col, template_ws.max_column)
            return min(template_last_period_col, template_ws.max_column)

        copy_row_layout(template_ws, ws, 1, 1)
        for col in range(1, total_col + 1):
            src_col = source_col(col)
            copy_column_layout(template_ws, ws, src_col, col)
            copy_cell_format(template_ws.cell(1, src_col), ws.cell(1, col))

        if template_ws.max_row < 2:
            return
        for row_idx in range(2, last_row + 1):
            copy_row_layout(template_ws, ws, 2, row_idx)
            for col in range(1, total_col + 1):
                src_col = source_col(col)
                copy_cell_format(template_ws.cell(2, src_col), ws.cell(row_idx, col))
    except Exception as exc:  # noqa: BLE001 - 樣式套用失敗不應阻斷數值報表產生
        warn(f"PP 樣式樣板套用失敗，已改用內建版面：{exc}")
    finally:
        if template_wb is not None:
            template_wb.close()
